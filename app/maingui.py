import os
import subprocess

import tempfile
from builtins import getattr

from PyQt5.QtWidgets import (
    QMainWindow,
    QTableView,
    QMessageBox,
)
from pipe import index

from sqlalchemy.exc import IntegrityError

from clipboard import ClipBoard

import advanced_definition_form
import rmas_incoming_form
import advanced_sale_proforma_form
import agentgui
import db
import expedition_form
import expenses_form
import inventory_form
import models
import partner_form
import payments_form
import product_form
import purchase_proforma_form
import reception_order
import sale_proforma_form
import documents_form
from db import PurchaseProforma, PurchaseDocument, \
    SaleDocument, SaleProforma, correct_mask
from models import SaleProformaModel, PurchaseProformaModel
from pdfbuilder import build_document
from ui_maingui import Ui_MainGui
from utils import (
    getTracking,
    getNote,
    get_directory
)


PASSWORD = '0010'

PREFIXES = [
    'agents_',
    'partners_',
    'proformas_purchases_',
    'proformas_sales_',
    'invoices_purchases_',
    'invoices_sales_',
    'warehouse_expeditions_',
    'warehouse_receptions_',
    'warehouse_incoming_rmas_',
    'warehouse_outgoing_rmas_',
    'tools_',
    'rmas_incoming_',
    'rmas_outgoing_',
]

ACTIONS = [
    'apply',
    'new',
    'edit',
    'delete',
    'cancel',
    'print',
    'export',
    'payments',
    'expenses',
    'docs',
    'toinv',
    'towh',
    'ship',
    'newadv',
    'mail',
    'options',
    'dhl',
    'process',
    'double_click',
    'search',
    'create_warehouse',
    'change_warehouse',
    'create_spec',
    'change_spec',
    'create_condition',
    'change_condition',
    'change_description',
    'check_inventory',
    'create_product',
    'view_pdf',
    'advnorm',
    'ready',
    'import',
    'dump',
    'template',
    'create_courier',
    'export_excel',
    'available_stock',
    'trace',
    'tocn',
    'sii',
    'harvest',
    'facks',
    'stock_valuation',
    'issued_invoices',
    'switch',
    'top_partners',
    'owing',
    'whatsapp'
]


def clean_up_directories():
    from contextlib import suppress
    import os
    with suppress(FileNotFoundError):
        for file in filter(lambda s: s.endswith('.pdf'), os.listdir()):
            os.remove(file)

    for file in os.listdir(os.path.join(os.curdir, 'temp')):
        os.remove(os.path.join(os.curdir, 'temp', file))


def get_prefix(model, document):
    if isinstance(model, (SaleProformaModel, PurchaseProformaModel)):
        return 'PI '
    return 'INV'

def view_document(view, model):
    ixs = view.selectedIndexes()
    if not ixs:
        return
    row = {i.row() for i in ixs}.pop()
    document = model[row]
    filename = get_prefix(model, document) + document.doc_repr + '.pdf'
    pdf_object = build_document(document)

    with tempfile.NamedTemporaryFile('wb', dir='.', delete=False, suffix='.pdf') as file:
        pdf_object.output(file.name)
        subprocess.Popen((file.name, ), shell=True)



def view_document_old(view, model):
    ixs = view.selectedIndexes()
    if not ixs:
        return
    row = {i.row() for i in ixs}.pop()
    document = model[row]
    filename = get_prefix(model, document) + document.doc_repr + '.pdf'
    pdf = build_document(document)




    pdf.output(filename)

    subprocess.Popen((filename,), shell=True)


def dump_rma(candidates, xlsx_file_path):
    from openpyxl import load_workbook
    try:
        workbook = load_workbook(xlsx_file_path)
        ws = workbook.active
        max_row = ws.max_row
        accepted_lines = filter(
            lambda l: l.accepted != 'n' and not models.exists_credit_line(l.sn),
            candidates
        )

        for i, line in enumerate(accepted_lines, start=1):
            for j, value in enumerate(line.as_excel_row, start=1):
                ws.cell(max_row + i, j, value=value)

        workbook.save(xlsx_file_path)

    except: # If does not work, ignore completely
        pass


class MainGui(Ui_MainGui, QMainWindow):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.setCentralWidget(self.main_tab)

        self.clipboard = ClipBoard(data='', form=self)

        # For closing
        self.opened_windows_instances = set()
        # Prevent creating multiple
        self.opened_windows_classes = set()

        self.main_tab.setCurrentIndex(0)

        self.init_models()

        self.set_handlers()

        self.main_tab.currentChanged.connect(self.tab_changed)

    def export_documents(self, view, model, ask_directory=True):
        rows = {i.row() for i in view.selectedIndexes()}
        if not rows:
            return

        if not ask_directory:
            from utils import get_desktop
            directory = get_desktop()

        else:
            directory = get_directory(self)

        if not directory:
            return

        for row in rows:
            doc = model[row]
            if isinstance(doc, (db.SaleInvoice, db.PurchaseInvoice)):
                name = 'INV ' + doc.doc_repr + '.pdf'
            elif isinstance(doc, db.PurchaseProforma):
                name = 'PO ' + doc.doc_repr + '.pdf'
            else:
                name = 'PI ' + doc.doc_repr + '.pdf'

            pdf = build_document(doc)
            path = os.path.join(directory, name)
            pdf.output(path)
        return path

    def selection_changed_generic(self, view):
        rows = {i.row() for i in view.selectedIndexes()}
        total, paid, device_count, owing = 0, 0, 0, 0

        for row in rows:
            doc = view.model()[row]
            paid += sum(p.amount for p in doc.payments)
            total += doc.total_debt
            device_count += doc.device_count
            owing += doc.owing

        name = view.objectName()
        prefix = name[0:name.rfind('_') + 1]

        getattr(self, prefix + 'total').setText('Total: ' + str(round(total, 2)))
        getattr(self, prefix + 'paid').setText('Paid: ' + str(round(paid, 2)))

        try:
            label = getattr(self, prefix + 'owing')
            text = 'Owes: ' if prefix in ('invoices_sales_', 'proformas_sales_') else 'Owing: '
            label.setText(text + str(owing))

            getattr(self, prefix + 'device_count').\
                setText('Processed Devices: ' + str(device_count))
        except AttributeError:
            pass

    def init_models(self):
        for prefix in PREFIXES:
            self.set_mv(prefix, last=0)

    def set_mv(self, prefix, search_key=None, filters=None, last=10):
        from utils import setCommonViewConfig
        from delegates import WarningEditDelegate

        try:
            last_field = getattr(self, prefix + 'last')
        except AttributeError:
            pass
        else:
            try:
                last = int(last_field.text())
            except ValueError:
                if last_field.text() == '':
                    last = 366

        if prefix == 'agents_':
            self.agents_model = models.AgentModel(search_key=search_key)
            self.agents_view.setModel(self.agents_model)
            setCommonViewConfig(self.agents_view)
            self.agents_view.resizeColumnToContents(1)
            self.agents_view.resizeColumnToContents(3)

        elif prefix == 'partners_':
            self.partners_model = models.PartnerModel(search_key=search_key)
            self.partners_view.setModel(self.partners_model)
            setCommonViewConfig(self.partners_view)
            self.partners_view.resizeColumnToContents(1)
            self.partners_view.resizeColumnToContents(2)
            self.partners_view.resizeColumnToContents(3)
            self.partners_view.resizeColumnToContents(7)

        elif prefix == 'proformas_purchases_':

            self.proformas_purchases_model = \
                models.PurchaseProformaModel(filters=filters, search_key=search_key, last=last)
            self.proformas_purchases_view.setModel(self.proformas_purchases_model)

            # Inmediately after setting model you need to put this code
            self.proformas_purchases_view.selectionModel(). \
                selectionChanged.connect(self.proformas_purchases_selection_changed)

            self.proformas_purchases_view.setSelectionBehavior(QTableView.SelectRows)
            self.proformas_purchases_view.setSortingEnabled(True)
            self.proformas_purchases_view.setAlternatingRowColors(True)
            self.proformas_purchases_view.resizeColumnToContents(3)
            self.proformas_purchases_view.resizeColumnToContents(4)
            self.proformas_purchases_view.resizeColumnToContents(5)
            self.proformas_purchases_view.resizeColumnToContents(6)

        elif prefix == 'invoices_purchases_':
            self.invoices_purchases_model = \
                models.PurchaseInvoiceModel(filters=filters, search_key=search_key, last=last)
            self.invoices_purchases_view.setModel(self.invoices_purchases_model)

            self.invoices_purchases_view.selectionModel(). \
                selectionChanged.connect(self.invoices_purchases_selection_changed)

            self.invoices_purchases_view.setSelectionBehavior(QTableView.SelectRows)
            self.invoices_purchases_view.setSortingEnabled(True)
            self.invoices_purchases_view.setAlternatingRowColors(True)

            self.invoices_purchases_view.resizeColumnToContents(3)
            self.invoices_purchases_view.resizeColumnToContents(4)
            self.invoices_purchases_view.resizeColumnToContents(13)


        elif prefix == 'invoices_sales_':
            self.invoices_sales_model = \
                models.SaleInvoiceModel(filters=filters, search_key=search_key, last=last)
            self.invoices_sales_view.setModel(self.invoices_sales_model)

            self.invoices_sales_view.selectionModel(). \
                selectionChanged.connect(self.invoices_sales_selection_changed)

            self.invoices_sales_view.setItemDelegate(
                WarningEditDelegate(
                    parent=self,
                    column=self.invoices_sales_model.WARNING
                )
            )

            self.invoices_sales_view.setSelectionBehavior(QTableView.SelectRows)
            self.invoices_sales_view.setSortingEnabled(True)
            self.invoices_sales_view.setAlternatingRowColors(True)

            self.invoices_sales_view.resizeColumnsToContents()


        elif prefix == 'proformas_sales_':

            self.proformas_sales_model = \
                models.SaleProformaModel(filters=filters, search_key=search_key, last=last)
            self.proformas_sales_view.setModel(
                self.proformas_sales_model
            )
            self.proformas_sales_view_delegate = WarningEditDelegate(
                parent=self, column=self.proformas_sales_model.WARNING
            )
            self.proformas_sales_view.setItemDelegate(
                self.proformas_sales_view_delegate
            )

            self.proformas_sales_view.selectionModel(). \
                selectionChanged.connect(self.proformas_sales_selection_changed)

            self.proformas_sales_view.setSelectionBehavior(QTableView.SelectRows)
            self.proformas_sales_view.setAlternatingRowColors(True)
            self.proformas_sales_view.setSortingEnabled(True)

            self.proformas_sales_view.resizeColumnsToContents()

            correct_mask()


        elif prefix == 'warehouse_receptions_':
            self.warehouse_receptions_model = \
                models.ReceptionModel(filters=filters, search_key=search_key, last=last)
            self.warehouse_receptions_view.setModel(self.warehouse_receptions_model)
            self.warehouse_receptions_view.setSelectionBehavior(QTableView.SelectRows)
            self.warehouse_receptions_view.setSortingEnabled(True)
            self.warehouse_receptions_view.setAlternatingRowColors(True)

            self.warehouse_receptions_view.resizeColumnToContents(6)
            self.warehouse_receptions_view.resizeColumnToContents(7)
            self.warehouse_receptions_view.resizeColumnToContents(8)
            self.warehouse_receptions_view.resizeColumnToContents(4)

        elif prefix == 'warehouse_expeditions_':
            self.warehouse_expeditions_model = \
                models.ExpeditionModel(filters=filters, search_key=search_key, last=last)
            self.warehouse_expeditions_view.setModel(self.warehouse_expeditions_model)
            self.warehouse_expeditions_view.setSelectionBehavior(QTableView.SelectRows)
            self.warehouse_expeditions_view.setAlternatingRowColors(True)
            self.warehouse_expeditions_view.setSortingEnabled(True)

            self.warehouse_expeditions_view.resizeColumnToContents(6)
            self.warehouse_expeditions_view.resizeColumnToContents(7)
            self.warehouse_expeditions_view.resizeColumnToContents(8)
            self.warehouse_expeditions_view.resizeColumnToContents(4)

        elif prefix == 'rmas_incoming_':
            self.rmas_incoming_model = models.RmaIncomingModel(
                search_key=search_key, filters=filters, last=last)
            self.rmas_incoming_view.setModel(self.rmas_incoming_model)

            self.rmas_incoming_view.setSelectionBehavior(QTableView.SelectRows)
            self.rmas_incoming_view.setSortingEnabled(True)
            self.rmas_incoming_view.setAlternatingRowColors(True)

        elif prefix == 'warehouse_incoming_rmas_':
            self.warehouse_rma_incoming_model = models.WhRmaIncomingModel(
                search_key=search_key, filters=filters, last=last
            )
            self.warehouse_incoming_rmas_view.setModel(self.warehouse_rma_incoming_model)
            self.warehouse_incoming_rmas_view.setSelectionBehavior(QTableView.SelectRows)
            self.warehouse_incoming_rmas_view.setSortingEnabled(True)
            self.warehouse_incoming_rmas_view.setAlternatingRowColors(True)

    def set_handlers(self):
        from itertools import product
        for prefix, action in product(PREFIXES, ACTIONS):
            try:
                widget_name = prefix + action

                self.attach_handler(prefix, action)
            except AttributeError:
                continue

    def attach_handler(self, prefix, action):
        try:
            if action == 'apply':
                handler = getattr(self, action + '_handler')
                widget = getattr(self, prefix + action)
                widget.clicked.connect(handler)
            elif action == 'double_click':
                if prefix == 'proformas_sales_':
                    return
                else:
                    handler = getattr(self, prefix + action + '_handler')
                    widget = getattr(self, prefix + 'view')
                    widget.doubleClicked.connect(handler)

            elif action == 'search':
                handler = getattr(self, action + '_handler')
                widget = getattr(self, prefix + action)
                widget.returnPressed.connect(handler)
            else:
                handler = getattr(self, prefix + action + '_handler')
                widget = getattr(self, prefix + action)
                widget.clicked.connect(handler)

        except AttributeError:
            raise

    def get_filters(self, *, prefix):
        filters = {}
        for name in [
            'types', 'financial', 'shipment',
            'logistic', 'cancelled', 'rmas_incoming'
        ]:
            try:
                filters[name] = getattr(self, name)(prefix=prefix)
            except AttributeError:
                continue
        return filters

    def rmas_incoming(self, prefix):
        return [
            name for name in (
                'empty', 'rejected', 'partially_accepted',
                'accepted'
            ) if getattr(self, prefix + name).isChecked()
        ]

    def cancelled(self, prefix):
        return [
            name for name in (
                'cancelled', 'notcancelled'
            ) if getattr(self, prefix + name).isChecked()
        ]

    def types(self, *, prefix):
        return [
            i for i in range(1, 7)
            if getattr(self, prefix + 'serie' + str(i)).isChecked()
        ]

    def financial(self, *, prefix):
        return [
            name for name in (
                'notpaid', 'partiallypaid', 'fullypaid'
            ) if getattr(self, prefix + name).isChecked()
        ]

    def shipment(self, *, prefix):
        return [
            name for name in ('sent', 'notsent')
            if getattr(self, prefix + name).isChecked()
        ]

    def logistic(self, *, prefix):
        return [
            name for name in (
                'empty', 'partially_processed',
                'completed', 'overflowed'
            ) if getattr(self, prefix + name).isChecked()
        ]

    def apply_handler(self):
        object_name = self.sender().objectName()
        prefix = object_name[0:object_name.rfind('_') + 1]
        filters = self.get_filters(prefix=prefix)
        search_key = getattr(self, prefix + 'search').text()
        self.set_mv(prefix, search_key=search_key, filters=filters)

    def search_handler(self):
        widget = self.sender()
        object_name = widget.objectName()
        prefix = object_name[0:object_name.rfind('_') + 1]
        filters = self.get_filters(prefix=prefix)
        search_key = widget.text()
        self.set_mv(prefix, filters=filters, search_key=search_key)

     # Agents Handlers:

    def agents_double_click_handler(self, index):
        self.launch_agent_form(index)

    def agents_new_handler(self):
        self.launch_agent_form()

    def agents_edit_handler(self):
        indexes = self.agents_view.selectedIndexes()
        if indexes:
            index = indexes[0]
            self.launch_agent_form(index)

    def launch_agent_form(self, index=None):
        if agentgui.AgentGui not in self.opened_windows_classes:
            self.a = agentgui.AgentGui(self, self.agents_view, index)
            self.a.show()

            self.opened_windows_instances.add(self.a)
            self.opened_windows_classes.add(agentgui.AgentGui)

    def agents_search_handler(self):
        key = self.agent_search_line_edit.text()
        self.setUpAgentsModelAndView(search_key=key)

    def agents_delete_handler(self):
        indexes = self.agents_view.selectedIndexes()
        if indexes:
            index = indexes[0]
            try:
                self.agentModel.delete(index)
                self.agents_view.clearSelection()
            except IntegrityError as e:
                if e.orig.args[0] == 1451:
                    QMessageBox.critical(self, 'Error - Delete', \
                                         "Could not delete Agent with data associated. Deactivate it.")

                    # Partners Handlers:

    def partners_double_click_handler(self, index):
        self.launch_partners_form(index)

    def partners_new_handler(self):
        self.launch_partners_form()

    def partners_edit_handler(self):
        indexes = self.partners_view.selectedIndexes()
        if indexes:
            index = indexes[0]
            self.launch_partners_form(index)

    def partners_delete_handler(self):
        indexes = self.partners_view.selectedIndexes()
        if indexes:
            index = indexes[0]
            try:
                self.partnerModel.delete(index)
                self.partners_view.clearSelection()
            except IntegrityError as e:
                QMessageBox.critical(self, "Error - Delete",
                                     "Could not delete Partner with data associated. Deactivate it.")

    def partners_search_handler(self):
        key = self.partner_search_line_edit.text()
        self.setUpPartnersModelAndView(search_key=key)

    def launch_partners_form(self, index=None):
        if partner_form.PartnerForm not in self.opened_windows_classes:
            self.p = partner_form.PartnerForm(self, self.partners_view, index)
            self.p.show()

            self.opened_windows_instances.add(self.p)
            self.opened_windows_classes.add(partner_form.PartnerForm)

    def proformas_purchases_view_pdf_handler(self):
        try:
            view_document(
                self.proformas_purchases_view,
                self.proformas_purchases_model
            )
        except Exception:
            # QMessageBox.critical(self, 'Error', 'Error viewing the file')
            raise


    def proformas_purchases_selection_changed(self, selection_model):
        self.selection_changed_generic(self.proformas_purchases_view)

    def proformas_purchases_double_click_handler(self, index):
        row = index.row()
        proforma = self.proformas_purchases_model[index.row()]
        self.epp = purchase_proforma_form.EditableForm(proforma)
        self.epp.show()

    def proformas_purchases_new_handler(self):
        self.ppf = purchase_proforma_form.Form(self.proformas_purchases_view)
        self.ppf.show()


    def proformas_purchases_cancel_handler(self):
        indexes = self.proformas_purchases_view.selectedIndexes()
        if not indexes:
            return
        if QMessageBox.question(self, 'Proformas - Cancel', 'Cancel proforma/s ?',\
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
            return
        self.proformas_purchases_model.cancel(indexes)

    def proformas_purchases_print_handler(self):
        print('print')

    def proformas_purchases_export_handler(self):
        exported_path = self.export_documents(
            self.proformas_purchases_view,
            self.proformas_purchases_model
        )
        if exported_path:

            QMessageBox.information(self, 'Success', 'Document Exported successfully')


    def proformas_purchases_payments_handler(self, invoice=None):
        if invoice:
            proforma = invoice
        else:
            proforma = self.get_proforma_purchase()
        if proforma:
            payments_form.PaymentForm(self, proforma).exec_()

    def proformas_purchases_expenses_handler(self, invoice=None):
        if invoice:
            proforma = invoice
        else:
            proforma = self.get_proforma_purchase()
        if proforma:
            expenses_form.ExpenseForm(self, proforma).exec_()

    def proformas_purchases_docs_handler(self):
        # documents_form.Form(self, self.get_proforma_purchase()).exec_()
        pass

    def proformas_purchases_toinv_handler(self):
        rows = {index.row() for index in self.proformas_purchases_view.selectedIndexes()}
        if not rows:
            return

        model = self.proformas_purchases_model
        if any(model[row].cancelled for row in rows):
            QMessageBox.critical(self, 'Error', 'Cancelled proforma/s')
            return

        if any(model[row].invoice is not None for row in rows):
            QMessageBox.critical(self, 'Error', 'Invoice already associated')
            return

        try:
            invoice = model.associateInvoice(rows)
        except ValueError as ex:
            QMessageBox.critical(self, 'Error', str(ex))
        else:
            QMessageBox.information(self, 'Success', f'{invoice.doc_repr} created.')

    def proformas_purchases_towh_handler(self, invoice=None):
        if invoice:
            proforma = invoice
        else:
            proforma = self.get_proforma_purchase()

        if proforma:
            try:
                ok, note = getNote(self, proforma)
                if ok:
                    self.proformas_purchases_model.toWarehouse(proforma, note)
                    QMessageBox.information(self, 'Information', 'Successfully created warehouse reception')
            except IntegrityError as ex:
                if ex.orig.args[0] == 1048:
                    d = 'Invoice' if invoice else 'Proforma'
                    QMessageBox.critical(self, 'Update - Error', f'Warehouse reception for this {d} already exists')

    def proformas_purchases_ship_handler(self):
        proforma = self.get_proforma_purchase()
        if proforma:
            tracking, ok = getTracking(self, proforma)
            if ok:
                self.proformas_purchases_model.ship(proforma, tracking)
                QMessageBox.information(self, 'Information', 'Proforma updated')

    def get_proforma_purchase(self):
        rows = {index.row() for index in \
                self.proformas_purchases_view.selectedIndexes()}
        if len(rows) == 1:
            return self.proformas_purchases_model.proformas[rows.pop()]

    # PROFORMAS SALES HANDLERS
    def proformas_sales_newadv_handler(self):
        self.asp = advanced_sale_proforma_form.get_form(
            self,
            self.proformas_sales_view
        )
        self.asp.show()

    def proformas_sales_view_pdf_handler(self):
        view_document(self.proformas_sales_view,self.proformas_sales_model)

    def proformas_sales_export_excel_handler(self):
        proforma = self.get_sale_proforma()
        if not proforma:
            return

        from models import export_proformas_sales_excel
        from utils import get_file_path
        save_file_path = get_file_path(self)

        if not save_file_path:
            return

        try:
            export_proformas_sales_excel(proforma, save_file_path)
        except:
            QMessageBox.critical(self, 'Error', 'Error exporting data')
            raise
        else:
            QMessageBox.information(self, 'Success', 'Data exported successfully')

    def proformas_sales_ready_handler(self):
        indexes = self.proformas_sales_view.selectedIndexes()
        if not indexes:
            return
        self.proformas_sales_model.ready(indexes)

    def proformas_sales_advnorm_handler(self):
        indexes = self.proformas_sales_view.selectedIndexes()
        rows = {i.row() for i in indexes}
        if len(rows) != 1:
            return
        row = rows.pop()
        try:
            proforma = self.proformas_sales_model[row]
        except IndexError:
            return

        if not proforma.advanced_lines:
            return

        self.advnorm_form = advanced_definition_form.get_form(
            self,
            self.proformas_sales_view,
            proforma
        )

        self.advnorm_form.show()

    def proformas_sales_selection_changed(self):
        self.selection_changed_generic(self.proformas_sales_view)

    def proformas_sales_new_handler(self):

        self.sp = sale_proforma_form.get_form(
            self,
            self.proformas_sales_view
        )
        self.sp.show()

    def proformas_sales_edit_handler(self):
        proforma = self.get_sale_proforma()
        if not proforma:
            return
        if proforma.advanced_lines:
            self.sp = advanced_sale_proforma_form.get_form(
                self,
                self.proformas_sales_view,
                proforma
            )

        elif proforma.credit_note_lines:
            from credit_note_form import Form
            self.sp = Form(self, proforma)

        else:
            self.sp = sale_proforma_form.get_form(
                self,
                self.proformas_sales_view,
                proforma
            )


        self.sp.show()

    def proformas_sales_cancel_handler(self):
        indexes = self.proformas_sales_view.selectedIndexes()
        if not indexes:
            return
        try:
            if QMessageBox.question(
                    self,
                    'Proformas - Cancel',
                    'Cancel proforma/s ?',
                    QMessageBox.Yes | QMessageBox.No) == QMessageBox.No:
                return

            self.proformas_sales_model.cancel(indexes)
        except Exception:
            QMessageBox.critical(self, 'Update Error', 'Error updating proformas')
            raise

    def proformas_sales_whatsapp_handler(self):
        proforma = self.get_sale_proforma()
        if not proforma:
            return
        from partner_or_agent_form import Dialog
        Dialog(self, proforma).exec_()


    def proformas_sales_export_handler(self):
        # Don't need the object, but signals if one is   selected
        proforma = self.get_sale_proforma()
        if not proforma:
            return

        exported_path = self.export_documents(
            self.proformas_sales_view,
            self.proformas_sales_model
        )
        if exported_path:
            QMessageBox.information(self, 'Success', 'Document Exported successfully')

    def proformas_sales_mail_handler(self):
        proforma = self.get_sale_proforma()
        if not proforma:
            return
        from utils import get_email_recipient
        recipient = get_email_recipient(proforma)
        if not recipient:
            return

        from pdfbuilder import build_document
        temp_directory = os.path.abspath(os.path.join(os.curdir, 'temp'))

        pdf_document = build_document(proforma)
        abs_path = os.path.join(temp_directory, proforma.doc_repr + '.pdf')

        pdf_document.output(abs_path)

        import subprocess

        completed_subprocess = subprocess.run(['mailunch.exe', 'A', recipient, abs_path])


    def proformas_sales_payments_handler(self):
        proforma = self.get_sale_proforma()
        if proforma:
            payments_form.PaymentForm(self, proforma, sale=True).exec_()

    def proformas_sales_expenses_handler(self, invoice=None):
        proforma = self.get_sale_proforma()
        if proforma:
            expenses_form.ExpenseForm(self, proforma, sale=True).exec_()

    def proformas_sales_docs_handler(self, invoice=None):
        # documents_form.Form(self, self.get_sale_proforma()).exec_()
        pass

    def proformas_sales_toinv_handler(self):

        rows = {index.row() for index in self.proformas_sales_view.selectedIndexes()}
        if not rows:
            return
        model = self.proformas_sales_model
        if any(model[row].cancelled for row in rows):
            QMessageBox.critical(self, 'Error', 'Cancelled proforma/s')
            return

        if any(model[row].invoice is not None for row in rows):
            QMessageBox.critical(self, 'Error', 'Invoice already associated')
            return

        try:
            invoice = model.associateInvoice(rows)
        except ValueError as ex:
            QMessageBox.critical(self, 'Error', str(ex))
        else:
            QMessageBox.information(self, 'Success', f'{invoice.doc_repr} created')


    def proformas_sales_towh_handler(self, invoice=None):
        proforma = self.get_sale_proforma()
        if not proforma:
            return
        try:
            ok, note = getNote(self, proforma)
            if not ok:
                return
            self.proformas_sales_model.toWarehouse(proforma, note)
            QMessageBox.information(self, 'Information', 'Successfully created warehouse expedition')
            self.proformas_purchases_view.clearSelection()

        except IntegrityError as ex:
            if ex.orig.args[0] == 1048:
                d = 'Invoice' if invoice else 'Proforma'
                QMessageBox.critical(
                    self, 'Update - Error',
                    f'Warehouse expedition for this {d} already exists'
                )

        except ValueError as ex:
            QMessageBox.critical(self, 'Error', str(ex))

    def proformas_sales_ship_handler(self):
        proforma = self.get_sale_proforma()
        if not proforma:
            return
        tracking, ok = getTracking(self, proforma)
        if ok:
            self.proformas_sales_model.ship(proforma, tracking)
            QMessageBox.information(self, 'Information', 'Proforma Updated')

    def get_sale_proforma(self,):
        rows = {index.row() for index in self.proformas_sales_view.selectedIndexes()}
        if len(rows) == 1:
            return self.proformas_sales_model.proformas[rows.pop()]

    def launch_adv_sale_proforma_form(self, index=None):
        from contextlib import suppress
        proforma = None
        with suppress(AttributeError):
            proforma = self.proformas_sales_model.proformas[index.row()]

        self.advsp = advanced_sale_proforma_form.get_form(
            self,
            self.proformas_sales_view,
            proforma
        )

        self.advsp.show()

    def invoices_purchases_view_pdf_handler(self):
        try:
            view_document(
                self.invoices_purchases_view,
                self.invoices_purchases_model
            )
        except Exception:
            QMessageBox.critical(self, 'Error', 'Error viewing the file')
            raise


    def invoices_purchases_selection_changed(self):
        self.selection_changed_generic(self.invoices_purchases_view)

    def invoices_purchases_payments_handler(self):
        rows = {i.row() for i in self.invoices_purchases_view.selectedIndexes()}
        if len(rows) != 1:
            return
        row = rows.pop()
        invoice = self.invoices_purchases_model[row]
        from invoice_payment_form import Form

        Form(self, invoice).exec_()

    def invoices_purchases_ship_handler(self):
        invoice = self.get_purchases_invoice()
        tracking, ok = getTracking(self, invoice)
        if ok:
            self.proformas_purchases_model.ship_several(invoice.proformas, tracking)
            QMessageBox.information(self, 'Success', 'Dependant proformas updated')


    def invoices_purchases_towh_handler(self):
        invoice = self.get_purchases_invoice()
        self.proformas_purchases_towh_handler(invoice.proforma)

    def invoices_purchases_docs_handler(self):
        pass

    def invoices_purchases_expenses_handler(self):
        rows = {i.row() for i in self.invoices_purchases_view.selectedIndexes()}
        if len(rows) != 1:
            return
        row = rows.pop()
        invoice = self.invoices_purchases_model[row]
        from invoices_expenses_form import Form
        Form(self, invoice).exec_()

    def invoices_purchases_double_click_handler(self, index):
        invoice = self.invoices_purchases_model[index.row()]
        from invoice_form import Form
        self.f = Form(invoice)
        self.f.show()

    def invoices_purchases_export_handler(self):
        exported_path = self.export_documents(
            self.invoices_purchases_view,
            self.invoices_purchases_model,
        )
        if exported_path:
            QMessageBox.information(self, 'Success', 'Document Exported successfully')


    def invoices_purchases_print_handler(self):
        print('print ')

    def get_purchases_invoice(self):
        return self.get_invoice(
            self.invoices_purchases_model,
            self.invoices_purchases_view
        )

    # SALES INVOICE HANDLER:
    def invoices_sales_view_pdf_handler(self):
        view_document(self.invoices_sales_view, self.invoices_sales_model)

    def invoices_sales_selection_changed(self):
        self.selection_changed_generic(self.invoices_sales_view)

    def invoices_sales_payments_handler(self):
        invoice = self.get_sales_invoice()
        from invoice_payment_form import Form
        Form(self, invoice).exec_()

    def invoices_sales_ship_handler(self):
        invoice = self.get_sales_invoice()
        if not invoice:
            return
        tracking, ok = getTracking(self, invoice)
        if ok:
            self.proformas_sales_model.ship_several(invoice.proformas, tracking)
            QMessageBox.information(self, 'Error', 'Invoice updated')

    def invoices_sales_docs_handler(self):
        # invoice = self.get_sales_invoice()
        # documents_form.Form(self, invoice, is_invoice=True).exec_()
        pass

    def invoices_sales_expenses_handler(self):
        invoice = self.get_sales_invoice()
        if invoice:
            from invoices_expenses_form import Form
            Form(self, invoice).exec_()

    # def invoices_sales_double_click_handler(self, index):

    def invoices_sales_edit_handler(self):
        # TODO CREDIT NOTE LOGIC
        invoice = self.get_sales_invoice()

        if invoice:

            # Credit note logic
            if len(invoice.proformas) == 1 and invoice.proformas[0].credit_note_lines:
                from credit_note_form import Form
                proforma = invoice.proformas[0]
                self.sip = Form(self, proforma, is_invoice=True)

            # Everything else
            else:
                from invoice_form import Form
                self.sip = Form(invoice)

            self.sip.show()

    def invoices_sales_whatsapp_handler(self):
        invoice = self.get_sales_invoice()
        if not invoice:
            return
        from partner_or_agent_form import Dialog
        Dialog(self, invoice).exec_()

    def invoices_sales_ready_handler(self):
        invoice = self.get_sales_invoice()
        if invoice:
            self.proformas_sales_model.ready_several(invoice.proformas)

    def invoices_sales_export_handler(self):
        exported_path = self.export_documents(
            self.invoices_sales_view,
            self.invoices_sales_model,
        )
        if exported_path:

            QMessageBox.information(self, 'Success', 'Document Exported successfully')

    def invoices_sales_export_excel_handler(self):
        invoice = self.get_sales_invoice()
        from models import export_invoices_sales_excel
        from utils import get_file_path
        save_file_path = get_file_path(self)
        if not save_file_path:
            return

        # Code like this allows for future modification
        # The current flow is to raise the exception
        # and inspect what caused it on the users
        # powershell. If this were commercial ,
        # exception handling would dump to a logging file
        # Then you would inspect that file
        # in order to control this exceptional situation and
        # improve the application.
        try:
            export_invoices_sales_excel(invoice, save_file_path)
        except:
            QMessageBox.critical(self, 'Error', 'Error exporting data')
            raise
        else:
            QMessageBox.information(self, 'Success', 'Data exported successfully')

    def invoices_sales_import_handler(self):
        from utils import get_open_file_path
        file_path = get_open_file_path(self, csv_filter=True)
        if not file_path:
            return
        else:
            try:
                resolved, unresolved = models.resolve_dhl_expenses(file_path)
            except ValueError as ex:
                QMessageBox.critical(self, 'Error', str(ex))
            else:
                from dhl_form import Form
                Form(self, resolved, unresolved).exec_()



    def invoices_sales_mail_handler(self):
        invoice = self.get_sales_invoice()
        if not invoice:
            return
        from utils import get_email_recipient
        recipient = get_email_recipient(invoice)
        if not recipient:
            return

        from models import export_invoices_sales_excel
        from pdfbuilder import build_document
        doc_repr = invoice.doc_repr

        temp_dir = os.path.abspath(os.path.join(os.curdir, 'temp'))

        path1 = os.path.join(temp_dir, doc_repr + '.pdf')
        path2 = os.path.join(temp_dir, 'series' + doc_repr + '.xlsx')

        document = build_document(invoice)
        document.output(path1)

        try:
            export_invoices_sales_excel(invoice, path2)
            generated = True

        except AttributeError:
            generated = False

        import subprocess
        args = ['mailunch.exe', recipient, path1]

        if generated:
            args.append(path2)

        if invoice.proformas[0].credit_note_lines:
            args.insert(1, 'C')
        else:
            args.insert(1, 'B')

        completed_process = subprocess.run(args, shell=True)

    def get_sales_invoice(self):
        return self.get_invoice(
            self.invoices_sales_model,
            self.invoices_sales_view
        )

    def get_invoice(self, model, view):
        rows = {index.row() for index in view.selectedIndexes()}
        if len(rows) == 1:
            return model[rows.pop()]

    # WAREHOUSE RECEPTION HANDLERS:

    def warehouse_receptions_process_handler(self):
        reception = self.get_reception(
            self.warehouse_receptions_view,
            self.warehouse_receptions_model
        )
        if not reception:
            return
        else:
            try:
                reception_order.Form(self, reception).exec_()
            except IndexError:
                QMessageBox.information(
                    self,
                    'Information',
                    'This reception order is empty. I can not open'
                )

    def warehouse_receptions_double_click_handler(self, index):
        self.warehouse_receptions_process_handler()

    def warehouse_receptions_import_handler(self):
        from utils import get_open_file_path
        row = self.get_reception_selected_row()
        if row is None:
            return
        file_path = get_open_file_path(self)
        if not file_path:
            return
        try:
            self.warehouse_receptions_model.import_excel(file_path, row)
        except ValueError as ex:
            QMessageBox.information(self, 'Error', str(ex))
        else:
            QMessageBox.information(self, 'Success', 'Data imported successfully')

    def warehouse_receptions_template_handler(self):
        from utils import get_file_path
        row = self.get_reception_selected_row()
        if row is None:
            return

        save_file_path = get_file_path(self)

        if not save_file_path:
            return
        try:
            self.warehouse_receptions_model.generate_template(save_file_path, row)
        except ValueError as ex:
            QMessageBox.information(self, 'Error', str(ex))
        else:
            QMessageBox.information(self, 'Success', 'Template built successfully')

    def warehouse_receptions_export_handler(self):
        from utils import get_file_path
        row = self.get_reception_selected_row()
        if row is None:
            return

        save_file_path = get_file_path(self)
        if not save_file_path:
            return

        try:
            self.warehouse_receptions_model.export(save_file_path, row)
        except ValueError as ex:
            QMessageBox.information(self, 'Error', str(ex))
        else:
            QMessageBox.information(self, 'Success', 'Data exported successfully')

    def get_reception_selected_row(self):
        rows = {index.row() for index in self.warehouse_receptions_view.selectedIndexes()}
        if len(rows) == 1:
            return rows.pop()

    def get_reception(self, view, model):
        rows = {index.row() for index in view.selectedIndexes()}
        if len(rows) == 1:
            return model.receptions[rows.pop()]

    # WAREHOUSE EXPEDITION:

    def warehouse_expeditions_export_handler(self):
        expedition = self.get_expedition(
            self.warehouse_expeditions_view,
            self.warehouse_expeditions_model
        )
        from models import export_proformas_sales_excel
        from utils import get_file_path

        save_file_path = get_file_path(self)

        if not save_file_path:
            return

        try:
            export_proformas_sales_excel(expedition.proforma, save_file_path)
        except:
            QMessageBox.critical(self, 'Error', 'Error exporting data')
            raise
        else:
            QMessageBox.information(self, 'Success', 'Data exported successfully')


    def warehouse_expeditions_process_handler(self):
        expedition = self.get_expedition(
            self.warehouse_expeditions_view,
            self.warehouse_expeditions_model
        )
        if not expedition: return

        try:
            expedition_form.Form(self, expedition).exec_()
        except IndexError:
            QMessageBox.information(
                self,
                'Information',
                'This expedition order is empty. I can not open'
            )

    def warehouse_expeditions_double_click_handler(self, index):
        expedition = self.warehouse_expeditions_model.expeditions[index.row()]
        try:
            expedition_form.Form(self, expedition).exec_()
        except IndexError:
            QMessageBox.information(
                self,
                'Information',
                'This expedition order is empty. I can not open'
            )

    def warehouse_expeditions_dhl_handler(self):
        from dhl import Form
        expedition = self.get_expedition(self.warehouse_expeditions_view, self.warehouse_expeditions_model)
        if expedition is None:
            return
        Form(self, expedition).exec_()


    def warehouse_expeditions_delete_handler(self):
        expedition = self.get_expedition(
            self.warehouse_expeditions_view,
            self.warehouse_expeditions_model
        )
        if not expedition:
            return

    def get_expedition(self, view, model):
        rows = {index.row() for index in view.selectedIndexes()}
        if len(rows) == 1:
            return model.expeditions[rows.pop()]

    # RMAS HANDLERS:
    def rmas_incoming_new_handler(self):
        from rmas_incoming_form import Form
        self.f = Form(self)
        self.f.show()

    def rmas_incoming_towh_handler(self):
        rows = {i.row() for i in self.rmas_incoming_view.selectedIndexes()}
        if len(rows) != 1:
            return
        row = rows.pop()
        try:
            self.rmas_incoming_model.to_warehouse(row)
        except ValueError as ex:
            print('catched value error')
            print(str(ex))
            QMessageBox.critical(self, 'Error', str(ex))
        else:
            QMessageBox.information(self, 'Error', 'RMA WH order created')


    def warehouse_incoming_rmas_double_click_handler(self):
        from warehouse_rma_incoming_form import Form
        order = self.get_wh_incoming_rma_order()
        self.whirma = Form(self, order)
        self.whirma.show()

    def get_wh_incoming_rma_order(self):
        rows = {i.row() for i in self.warehouse_incoming_rmas_view.selectedIndexes()}
        if len(rows) != 1:
            return
        row = rows.pop()
        return self.warehouse_rma_incoming_model.orders[row]

    def warehouse_incoming_rmas_process_handler(self):
        from warehouse_rma_incoming_form import Form
        order = self.get_wh_incoming_rma_order()
        if not order:
            return
        self.whirma = Form(self, order)
        self.whirma.show()

    def warehouse_incoming_rmas_tocn_handler(self):

        wh_rma_order = self.get_wh_incoming_rma_order()

        if not wh_rma_order:
            return

        candidates = [
            line
            for line in wh_rma_order.lines
            if line.accepted == 'y' and models.rma_credit_difference(line.sn)
        ]

        if not candidates:
            QMessageBox.critical(self, 'Error', "I did not found candidates for a new Credit Note")
            return

        # from utils import get_open_file_path
        # file_path = get_open_file_path(self)
        # if not file_path:
        #     return

        partner_id = wh_rma_order.incoming_rma.lines[0].cust_id

        agent_id = wh_rma_order.incoming_rma.lines[0].agent_id

        from models import build_credit_note_and_commit

        proforma = build_credit_note_and_commit(partner_id, agent_id, wh_rma_order, candidates)

        # If None, no proforma was built because there was not candidates
        if not proforma:
            return

        invoice = self.proformas_sales_model.build_invoice_from_proforma(proforma, wh_rma_order)

        for line in candidates:
            imei = db.Imei()
            imei.imei = line.sn
            imei.item_id = line.item_id
            imei.condition = line.condition
            imei.spec = line.spec
            imei.warehouse_id = line.warehouse_id

            db.session.add(imei)

        try:
            db.session.commit()
        except IntegrityError:
            QMessageBox.critical(self, 'Error', 'Some imeis are already in Inventory')
            db.session.rollback()
        else:
            QMessageBox.information(
                self,
                'Success',
                f'Credit Note:{invoice.doc_repr} built successfully. Inventory Updated.'
            )

        # dump_rma(candidates, file_path)

    def rmas_incoming_double_click_handler(self, index):
        rma_order = self.rmas_incoming_model[index.row()]
        self.f = rmas_incoming_form.Form(self, order=rma_order)
        self.f.show()

    # TOOLS HANDLERS:
    def tools_trace_handler(self):
        from trace_form import Form
        f = Form(self)
        f.exec_()

    def tools_trace_handler(self):
        from trace_form import Form
        f = Form(self)
        f.exec_()

    def tools_available_stock_handler(self):
        import available_stock
        self.f = available_stock.Form(self)
        self.f.exec_()

    def tools_create_courier_handler(self):
        import courier
        d = courier.CourierForm(self)
        d.exec_()

    def tools_create_product_handler(self):
        d = product_form.ProductForm(self)
        d.exec_()

    def tools_check_inventory_handler(self):
        d = inventory_form.InventoryForm(self)
        d.exec_()

    def tools_change_spec_handler(self):
        from utils import getPassword
        password = getPassword(self)
        if password != PASSWORD:
            return

            # if models.stock_gap():
        #     QMessageBox.information(self, 'Information', 'Process all sales first.')
        #     return
            # d = spec_change_form.SpecChange(self)
        # d.exec_()

        from change_form import ChangeForm
        d = ChangeForm(parent=self, hint='spec')
        d.exec_()

    def tools_change_condition_handler(self):
        from utils import getPassword
        password = getPassword(self)
        if password != PASSWORD:
            return


        from change_form import ChangeForm
        d = ChangeForm(parent=self, hint='condition')
        d.exec_()

    def tools_change_warehouse_handler(self):
        from utils import getPassword
        password = getPassword(self)
        if password != PASSWORD:
            return

        from change_form import ChangeForm
        d = ChangeForm(parent=self, hint='warehouse')
        d.exec_()

    def tools_change_description_handler(self):
        from change_description_form import Form
        Form(self).exec_()

    def tools_harvest_handler(self):
        from harvest import Form
        Form(self).exec_()

    def tools_create_warehouse_handler(self):
        from warehouse import Form
        Form(self).exec_()

    def tools_create_condition_handler(self):
        from condition import Form
        Form(self).exec_()

    def tools_create_spec_handler(self):
        from spec import Form
        Form(self).exec_()

    def tools_sii_handler(self):
        from pysii import Form
        Form(self).exec_()

    def tools_facks_handler(self):
        from facks import Form
        Form(self).exec_()

    def tools_stock_valuation_handler(self):
        from stock_valuation import Form
        Form(self).exec_()

    def tools_issued_invoices_handler(self):
        from issued_invoices_form import Form
        Form(self).exec_()


    def tools_switch_handler(self):
        # from switch_form import Form
        # Form(self).exec_()

        QMessageBox.information(self, 'Info', 'Under Development')

    def tools_top_partners_handler(self):

        from top_partners_form import Form

        Form(self).exec_()

    def tab_changed(self, index):
        db.session.commit()

        if index == 1:
            prefix = 'agents_'
            self.set_mv(prefix, search_key=getattr(self, prefix + 'search').text())

        elif index == 2:
            prefix = 'partners_'
            self.set_mv(prefix, search_key=getattr(self, prefix + 'search').text())
        elif index == 3:
            for prefix in ['proformas_purchases_', 'proformas_sales_']:
                self.set_mv(prefix, search_key=getattr(self, prefix + 'search').text(),
                            filters=self.get_filters(prefix=prefix))

        elif index == 4:
            for prefix in ['invoices_purchases_', 'invoices_sales_']:
                self.set_mv(
                    prefix, search_key=getattr(self, prefix + 'search').text(),
                    filters=self.get_filters(prefix=prefix)
                )

        elif index == 5:

            for prefix in ['warehouse_expeditions_', 'warehouse_receptions_',
                           'warehouse_incoming_rmas_','warehouse_outgoing_rmas_']:
                self.set_mv(prefix, search_key=getattr(self, prefix + 'search').text(),
                            filters=self.get_filters(prefix=prefix))

        elif index == 6:
            prefix = 'rmas_incoming_'
            self.set_mv(prefix, search_key=getattr(self, prefix + 'search').text(),
                        filters=self.get_filters(prefix=prefix))

    def closeEvent(self, event):
        for w in self.opened_windows_instances:
            w.close()

        clean_up_directories()

        super().closeEvent(event)

