
import os 

from PyQt5.QtWidgets import QDialog, QMessageBox

import db
from ui_stock_valuation import Ui_Dialog

from models import StockValuationModelDocument, WarehouseValueModel
from models import StockValuationModelImei

import utils

from openpyxl import load_workbook
from pyvalidator import is_imei 


def match_doc_repr_and_year(doc_repr):
    import re
    pattern = r'^[1-6]-[0-9]{1,6}:20[2-9][2-9]$'
    return bool(re.match(pattern, doc_repr))


def parse_doc_repr_and_year(doc_repr):
    type_, number = doc_repr.split('-')
    number, year = number.split(':')
    return int(type_), int(number), int(year)


def prepare_filters(filters):

    if not filters['date']:
        raise ValueError('Date must be provided')
    else:
        try:
            filters['date'] = utils.parse_date(filters['date']).date()
        except ValueError:
            raise ValueError('Incorrect date format: ddmmyyyy')

    try:
        filters['warehouse_id'] = utils.warehouse_id_map[filters['warehouse']]
    except KeyError:
        filters['warehouse_id'] = None

    if not filters['all'] and not filters['warehouse_id']:
        raise ValueError('Warehouse must be entered or all wh. checked.')

    return filters

class Form(Ui_Dialog, QDialog):

    def __init__(self, parent):
        super(Form, self).__init__(parent=parent)
        self.setupUi(self)
        self.model = None
        self.calculate.clicked.connect(self.calculate_handler)
        self.by_document.toggled.connect(self.by_doc_toggled)
        self.by_serial.toggled.connect(self.by_serial_toggled)
        self.by_warehouse.toggled.connect(self.by_warehouse_toggled)
        self.export_.clicked.connect(self.export_handler)
        self.all.toggled.connect(self.all_toggled)
        self.batch.clicked.connect(self.batch_handler)

        utils.setCompleter(self.warehouse, utils.warehouse_id_map.keys())
        self.all_path = None

        self.date.setText(utils.today_date())
        self.warehouse.setText(db.session.query(db.Warehouse.description).where(
            db.Warehouse.id == 1
        ).scalar())

        # Data holder for the new batch feature. 
        # on this container we put the elements of the batch
        # Serves as a flag also. 
        self.imeis_batch = list() # Needs to be a falsy value. iter([]) doesn't work. Iter hides the truthiness/falsiness.


    def by_doc_toggled(self, check):
        self.by_serial.setChecked(False)
        self.by_warehouse.setChecked(False)

    def by_serial_toggled(self, check):
        self.by_document.setChecked(False)
        self.by_warehouse.setChecked(False)

    def by_warehouse_toggled(self, check):
        self.by_serial.setChecked(False)
        self.by_document.setChecked(False)

    def export_handler(self):
        if self.model is None:
            return
        filepath = utils.get_file_path(self)

        try:
            self.model.export(filepath)
        except AttributeError:
            raise
        else:
            QMessageBox.information(self, 'Info', 'Data exported')

    def calculate_handler(self, check):
        if self.by_document.isChecked():
            doc_repr = self.document.text()

            if match_doc_repr_and_year(doc_repr):
                type_, number, year = parse_doc_repr_and_year(doc_repr)

                try:
                    self.model = StockValuationModelDocument(type_, number, year, self.proforma.isChecked())
                    self.view.setModel(self.model)

                except ValueError as ex:
                    QMessageBox.critical(self, 'Error', str(ex))
                    return

            else:
                QMessageBox.critical(self, 'Error', 'Incorrect document representation. Correct format: d-nnnnnn:yyyy')
                return

        elif self.by_serial.isChecked():
            # Understand the short-circuit behavior
            # Batch data has the priority.
            # imeis field is the fallback.
            self.model = StockValuationModelImei(filter(bool, map(str.strip, self.serial.text().split(','))), self.imeis_batch)
            self.view.setModel(self.model)

        elif self.by_warehouse.isChecked():
            try:
                filters = self.build_filters()
            except ValueError as ex:
                QMessageBox.critical(self, 'Error', str(ex))
                return
            else:
                self.model = WarehouseValueModel(filters)
                self.view.setModel(self.model)

    def all_toggled(self, checked):
        if checked:
            directory = utils.get_directory(self)
            if not directory:
                self.all.setChecked(False)
            else:
                self.all_path = directory
                self.warehouse.setText('')
        else:
            self.all_path = None

    def build_filters(self):
        return prepare_filters(
            {
                'date': self.date.text(),
                'all': self.all_path,
                'book_value': self.book_value.isChecked(),
                'warehouse': self.warehouse.text()
            }
        )


    def batch_handler(self):
        filepath = utils.get_open_file_path(self) 
        if not filepath:
            return
        wb = load_workbook(filepath)
        ws = wb.active
        try:
            # use genexpression to optimize memory usage. 
            self.imeis_batch = (cell.value for cell in ws['A'] if cell.value and is_imei(cell.value))
            self.batch_info.setText(f'Imeis data read from {os.path.basename(filepath)}')
        except KeyError:
            QMessageBox.critical(self, 'Error', 'Invalid file. No data found in the file.')

       