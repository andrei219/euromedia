import datetime
import os
import time

import openpyxl
from collections import Counter, namedtuple

from sqlalchemy import func


import utils

from db import session, ReceptionSerie, ExpeditionSerie, \
	CreditNoteLine, Imei, ConditionChange, SpecChange, \
	WarehouseChange, PurchaseProforma, Reception, ReceptionLine, \
	WhIncomingRmaLine, Warehouse

warehouse_id_map = {k.lower(): v for k, v in utils.warehouse_id_map.items()}


Change = namedtuple('Change', 'created_on target')

# This structure is build in order to match
# the results from the query in models.ActualInventory
# and then models.no_serie_grouper can act on both results
# HistoricalInventory.inventory and ActualInventory.inventory
# because it sees the same properties .
# Aux suffix in order to distinguish from db.Imei.

ImeiAux = namedtuple('Imei', 'imei item_id condition spec warehouse_id')


class Changes:

	def __init__(self, cls):
		self.changes = {
			r.sn.lower(): Change(r.created_on, r.after.lower())
			for r in session.query(cls)
		}

	def __getitem__(self, serie):
		return self.changes[serie]

	def __contains__(self, serie_cutoff_date):
		serie, cutoff_date = serie_cutoff_date
		if serie in self.changes:
			change = self.changes[serie]
			if change.created_on.date() <= cutoff_date:
				return True
			else:
				return False
		return False


condition_changes = Changes(ConditionChange)
spec_changes = Changes(SpecChange)
warehouse_changes = Changes(WarehouseChange)


def get_warehouse_id_from_purchase(serie):
	return (
		session.query(PurchaseProforma.warehouse_id).
		join(Reception).join(ReceptionLine).join(ReceptionSerie).
		where(ReceptionSerie.serie == serie).
		order_by(ReceptionSerie.id.desc()).first().warehouse_id
	)

def get_warehouse_id_from_rma(serie):
	return session.query(
		WhIncomingRmaLine.warehouse_id
	).where(WhIncomingRmaLine.sn == serie).order_by(WhIncomingRmaLine.id.desc()).first().warehouse_id


def find_attributes_and_return_inventory_register(serie, cutoff_date):

	last_reception = (
		session.query(ReceptionSerie).
		where(ReceptionSerie.serie == serie).
		order_by(ReceptionSerie.id.desc()).
		first()
	)

	last_rma = (
		session.query(CreditNoteLine).
		where(CreditNoteLine.sn == serie).
		order_by(CreditNoteLine.id.desc()).
		first()
	)

	# Build base events:
	base_events = [last_reception]
	if last_rma is not None:
		base_events.insert(0, last_rma)

	# 1. GET ITEM_ID:
	item_id = max(base_events, key=lambda e: e.created_on.date()).item_id

	# 2. GET CONDITION:
	if (serie, cutoff_date) in condition_changes:
		events = [condition_changes[serie]] + base_events
	else:
		events = base_events
	last_cause = max(events, key=lambda e: e.created_on.date())

	try:
		condition = last_cause.target
	except AttributeError:
		condition = last_cause.condition

	# 3. GET SPEC:
	if (serie, cutoff_date) in spec_changes:
		events = base_events + [spec_changes[serie]]
	else:
		events = base_events
	last_cause = max(events, key=lambda e: e.created_on)

	try:
		spec = last_cause.target
	except AttributeError:
		spec = last_cause.spec

	# 3. GET WH_ID:
	if (serie, cutoff_date) in warehouse_changes:
		events = base_events + [warehouse_changes[serie]]
	else:
		events = base_events
	event = max(events, key=lambda e: e.created_on)

	try:
		warehouse_name = event.target
		warehouse_id = warehouse_id_map[warehouse_name]

	except AttributeError:
		if event == last_reception:
			warehouse_id = get_warehouse_id_from_purchase(serie)
		elif event == last_rma:
			warehouse_id = get_warehouse_id_from_rma(serie)

	return ImeiAux(serie.lower(), item_id, condition.lower(), spec.lower(), warehouse_id)

class Rmas:

	def __init__(self, cutoff_date):
		self.vectors = Counter(
			{
				r.serie: r.qnt for r in
				session.query(
					func.lower(CreditNoteLine.sn).label('serie'),
					func.count(CreditNoteLine.id).label('qnt')
				).where(func.date(CreditNoteLine.created_on) <= cutoff_date).group_by(CreditNoteLine.sn)
			}
		)

	def __iter__(self):
		return iter(self.vectors)

	def __getitem__(self, item):
		return self.vectors[item]

class Inputs:

	def __init__(self, cutoff_date):
		self.vectors = Counter(
			{
				r.serie: r.qnt for r in
				session.query(
					func.lower(ReceptionSerie.serie).label('serie'), func.count(ReceptionSerie.id).label('qnt')
				).where(func.date(ReceptionSerie.created_on) <= cutoff_date).group_by(
					ReceptionSerie.serie
				)
			}
		)

	def __isub__(self, other):
		self.vectors.subtract(other.vectors)
		return self

	def __iter__(self):
		return iter(self.vectors)

	def __getitem__(self, item):
		return self.vectors[item]

class Outputs:

	def __init__(self, cutoff_date):
		self.vectors = Counter(
			{
				r.serie: r.qnt for r in
				session.query(
					func.lower(ExpeditionSerie.serie).label('serie'), func.count(ExpeditionSerie.id).label('qnt')
				).where(func.date(ExpeditionSerie.created_on) <= cutoff_date).group_by(
					ExpeditionSerie.serie
				)
			}
		)

	def __isub__(self, rmas):
		self.vectors.subtract(rmas.vectors)
		return self

	def __iter__(self):
		return iter(self.vectors)

	def __getitem__(self, item):
		return self.vectors[item]


# TODO: FILTROS: Inventory Model, HistoricalInventory

class HistoricalInventory:

	def __init__(self, cutoff_date, serie_only=True):
		inputs, outputs, rmas = Inputs(cutoff_date), Outputs(cutoff_date), Rmas(cutoff_date)
		outputs -= rmas
		inputs -= outputs
		self.series = {serie for serie, qnt in inputs.vectors.items() if qnt == 1}

		if serie_only:
			self.series = set(filter(lambda s: not utils.valid_uuid(s), self.series))

		self.inventory = {find_attributes_and_return_inventory_register(s, cutoff_date) for s in self.series}

	def __len__(self):
		return len(self.inventory)

	def __iter__(self):
		return iter(self.inventory)


class Inventory30Dec:

	DESCRIPTION_COLUMN = 'A'
	IMEI_COLUMN = 'D'
	CONDITION_COLUMN = 'B'
	SPEC_COLUMN = 'C'

	files_wh_id = {r.description + '.xlsx': r.id for r in session.query(Warehouse)}

	def __len__(self):
		return len(self.inventory)

	def __init__(self):
		base_dir = r'C:\Users\Andrei\Desktop\Updated Prices'
		self.inventory = set()
		for file, wh_id in self.files_wh_id.items():
			filepath = os.path.join(base_dir, file)
			s = self.get_from_workbook(filepath, wh_id)
			self.inventory.update(s)

	def get_from_workbook(self, filepath, warehouse_id):
		try:
			workbook = openpyxl.load_workbook(filepath)
		except FileNotFoundError:
			return {}

		sheet = workbook.active
		s = set()
		ord_a = ord('A')

		first = True
		for row in sheet.iter_rows(values_only=True):
			if first:
				first = False
				continue

			serie = row[ord(self.IMEI_COLUMN) - ord_a]
			if not serie:
				continue
			else:
				s.add((
					serie.lower(),
					utils.description_id_map[row[ord(self.DESCRIPTION_COLUMN) - ord_a]],
					row[ord(self.CONDITION_COLUMN) - ord_a].lower(),
					row[ord(self.SPEC_COLUMN) - ord_a].lower(),
					warehouse_id
				))
		return s

def test():
	today = datetime.date.today()
	history = HistoricalInventory(today).series
	inventory = set(r.imei for r in session.query(func.lower(Imei.imei).label('imei')))

	print('Len(inventory)=', len(inventory))
	print('Len(HistoricalInventory)=', len(history))
	print('Len(Inventory - HistoricalInventory)=', len(inventory - history))
	print('Len(HistoricalInventory - Inventory)=', len(history - inventory))
	print('Len(History & Inventory)=', len(history & inventory))
	print('Historical inventory is equal to inventory = ', history == inventory)

def save_excel(data, filename):
	workbook = openpyxl.Workbook()
	worksheet = workbook.active
	for row in data:
		worksheet.append(row)
	workbook.save(filename)

def assert_complete_equality():
	from datetime import date

	actual_inventory = {
		(r.imei.lower().strip(), r.item_id, r.condition.lower().strip(), r.spec.lower().strip(), r.warehouse_id)
		for r in session.query(Imei)
	}

	print('Len(Actual Inventory)=', len(actual_inventory))

	actual_historical_inventory = HistoricalInventory(date.today(), serie_only=False)

	print('Len(Historical Inventory)=', len(actual_historical_inventory))

	actual_history = actual_inventory - actual_historical_inventory.inventory
	history_actual = actual_historical_inventory.inventory - actual_inventory

	print('Len(A - H) =', len(actual_history))
	print('Len(H - A) =', len(history_actual))

	save_excel(actual_history, r'C:\Users\Andrei\Desktop\Projects\euromedia\actual_minus_history.xlsx')
	save_excel(history_actual, r'C:\Users\Andrei\Desktop\Projects\euromedia\history_minus_actual.xlsx')

def assert_30_dic_imeis_equality():

	inventory_30_dec = Inventory30Dec().inventory
	print('Len(inventory_30_dec)=', len(inventory_30_dec))

	start = time.time()
	historic = HistoricalInventory(utils.parse_date('30122022').date(), serie_only=True).inventory
	print('Elapsed history computation=', time.time() - start)
	print('Len(historic)=', len(historic))

	dec_history = inventory_30_dec - historic
	history_dec = historic - inventory_30_dec

	print('Len(D - H)=', len(dec_history))
	print('Len(H - D)=', len(history_dec))

	save_excel(dec_history, r'C:\Users\Andrei\Desktop\Projects\euromedia\dec_history.xlsx')
	save_excel(history_dec, r'C:\Users\Andrei\Desktop\Projects\euromedia\history_dec.xlsx')


if __name__ == '__main__':

	history = HistoricalInventory(utils.parse_date('30042022').date())