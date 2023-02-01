import os
from openpyxl import load_workbook

from models import WarehouseSimpleValueEntry


from itertools import dropwhile


IMEI_COLUMN, COST_COLUMN = 4, 9


def process_workbook(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    for i, row in enumerate(dropwhile(lambda row: row[0] == 'Description', sheet.iter_rows(values_only=True)), 1):
        serie = row[IMEI_COLUMN-1]  # column numbering starts at 1, not 0

        m = WarehouseSimpleValueEntry(serie.strip())

        print(m)


    # wb.save(filename)

if __name__ == '__main__':
    directory = r'C:\Users\Andrei\Desktop\Updated Prices'
    for file in os.listdir(directory):
        if file.endswith('.xlsx'):
            process_workbook(os.path.join(directory, file))
