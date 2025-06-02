import csv
import decimal
import re
import typing
import math
import os

from pathlib import Path
from collections import namedtuple
from itertools import groupby, product, combinations
from operator import attrgetter
from datetime import datetime
from collections.abc import Iterable

import openpyxl
from openpyxl import Workbook

import sqlalchemy
from sqlalchemy import exists

from PyQt5 import QtCore
from PyQt5 import QtGui
from PyQt5.QtCore import QModelIndex
from PyQt5.QtCore import Qt

from sqlalchemy.exc import InvalidRequestError, NoResultFound
from sqlalchemy import func
from sqlalchemy import or_, and_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.sql import text

import pyVies.api as vies

import utils
import db
from utils import description_id_map
from exceptions import SeriePresentError, PartnerCloningError

from utils import get_next_num
from datetime import date

# COLORS:
# RED FOR CANCELLED
# GREEN FOR COMPLETED
# ORANGE FOR EMPTY OR PARTIAL
# YELLOW FOR OVERFLOW
RED, GREEN, YELLOW, ORANGE = '#FF7F7F', '#90EE90', '#FFFF66', '#FFD580'


class BaseTable:
	
	def headerData(self, section, orientation, role=Qt.DisplayRole):
		if role == Qt.TextAlignmentRole:
			if orientation == Qt.Horizontal:
				return Qt.AlignLeft | Qt.AlignVCenter
			return Qt.AlignRight | Qt.AlignVCenter
		if role != Qt.DisplayRole:
			return
		if orientation == Qt.Horizontal:
			return self._headerData[section]
	
	def columnCount(self, index=QModelIndex()):
		return len(self._headerData)
	
	def rowCount(self, index=QModelIndex()):
		return len(getattr(self, self.name))


def computeCreditAvailable(partner_id):
	max_credit = db.session.query(db.Partner.amount_credit_limit). \
		where(db.Partner.id == partner_id).scalar()
	
	total = db.session.query(func.sum(db.PurchaseProformaLine.quantity * db.PurchaseProformaLine.price)). \
		select_from(db.Partner, db.PurchaseProforma, db.PurchaseProformaLine). \
		where(db.PurchaseProformaLine.proforma_id == db.PurchaseProforma.id). \
		where(db.PurchaseProforma.partner_id == db.Partner.id). \
		where(db.Partner.id == partner_id).scalar()
	
	paid = db.session.query(func.sum(db.PurchasePayment.amount)).select_from(db.Partner,
	                                                                         db.PurchaseProforma,
	                                                                         db.PurchasePayment).where(
		db.PurchaseProforma.partner_id == db.Partner.id). \
		where(db.PurchasePayment.proforma_id == db.PurchaseProforma.id). \
		where(db.Partner.id == partner_id).scalar()
	
	credit_taken = db.session.query(func.sum(db.PurchaseProforma.credit_amount)). \
		where(db.PurchaseProforma.partner_id == partner_id).scalar()
	
	# Protect against None results from partners with no credits.
	if max_credit is None:
		max_credit = 0
	if total is None:
		total = 0
	if paid is None:
		paid = 0
	if credit_taken is None:
		credit_taken = 0
	
	return max_credit + paid - total - credit_taken


def stock_gap():
	return not all(
		proforma.completed
		for proforma in db.session.query(db.SaleProforma)
		.where(db.SaleProforma.cancelled == False)
		.order_by(db.SaleProforma.id.desc())
	)


class AgentModel(QtCore.QAbstractTableModel):
	ID, NAME, PHONE, EMAIL, COUNTRY, ACTIVE = 0, 1, 2, 3, 4, 5
	
	def __init__(self, search_key=None):
		super().__init__()
		self._headerData = ['Code', 'Agent', 'Phone Nº', 'E-mail', 'Country', 'Active']
		query = db.session.query(db.Agent)
		
		if search_key:
			query = query.filter(
				or_(
					db.Agent.fiscal_name.contains(search_key),
					db.Agent.fiscal_number.contains(search_key),
					db.Agent.email.contains(search_key),
					db.Agent.phone.contains(search_key),
					db.Agent.country.contains(search_key)
				)
			)
		
		self.agents = query.all()
	
	def data(self, index, role=Qt.DisplayRole):
		
		if not index.isValid() or \
				not (0 <= index.row() <= len(self.agents)):
			return
		agent = self.agents[index.row()]
		column = index.column()
		if role == Qt.DisplayRole:
			if column == AgentModel.ID:
				return agent.id
			elif column == AgentModel.NAME:
				return agent.fiscal_name
			elif column == AgentModel.PHONE:
				return agent.phone
			elif column == AgentModel.EMAIL:
				return agent.email
			elif column == AgentModel.COUNTRY:
				return agent.country
			elif column == AgentModel.ACTIVE:
				return 'Yes' if agent.active else 'No'
		elif role == Qt.DecorationRole:
			if column == AgentModel.ACTIVE:
				if agent.active:
					return QtGui.QIcon(':\greentick')
				else:
					return QtGui.QIcon(':\cross')
		elif role == Qt.TextAlignmentRole:
			if column == AgentModel.ID:
				return Qt.AlignVCenter | Qt.AlignHCenter
	
	def headerData(self, section, orientation, role=Qt.DisplayRole):
		if role == Qt.TextAlignmentRole:
			if orientation == Qt.Horizontal:
				return Qt.AlignLeft | Qt.AlignVCenter
			return Qt.AlignRight | Qt.AlignVCenter
		if role != Qt.DisplayRole:
			return
		if orientation == Qt.Horizontal:
			return self._headerData[section]
	
	def rowCount(self, index=QModelIndex()):
		return len(self.agents)
	
	def columnCount(self, index=QModelIndex()):
		return 6
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		return super().flags(index) | Qt.ItemIsEditable
	
	def add(self, agent):
		db.session.add(agent)
		try:
			db.session.commit()
			self.agents.append(agent)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
		
		try:
			self.layoutAboutToBeChanged.emit()
			db.session.commit()
			self.layoutChanged.emit()
		
		except:
			db.session.rollback()
			raise
	
	def delete(self, index):
		if not index.isValid():
			return
		row = index.row()
		candidate_agent = self.agents[
			row]
		db.session.delete(candidate_agent)
		try:
			db.session.commit()
			# Dont check ValueError, this code executes only if self.agents is populated
			self.agents.remove(candidate_agent)
			self.layoutChanged.emit()
		
		except IntegrityError:
			db.session.rollback()
			raise
	
	def sort(self, section, order):
		attr = {AgentModel.NAME: 'fiscal_name', AgentModel.EMAIL: 'email', \
		        AgentModel.COUNTRY: 'country'}.get(section)
		if attr:
			self.layoutAboutToBeChanged.emit()
			self.agents = sorted(self.agents, key=operator.attrgetter(attr), \
			                     reverse=True if order == Qt.DescendingOrder else False)
			self.layoutChanged.emit()


dropbox_base = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Dropbox', 'Programa')


def move(origin, dest, copy=False):
	origin = origin.replace('/', '\\')
	dest = dest.replace('/', '\\')
	with open(origin, "rb") as od, open(dest, "wb") as dd:
		dd.write(od.read())
	if not copy:
		os.remove(origin)


class DocumentModel(QtCore.QAbstractListModel):
	
	def __init__(self):
		super().__init__()
		self.document_names = list(filter(self.key, os.listdir(self.path)))
	
	def rowCount(self, parent: QModelIndex = ...) -> int:
		return len(self.document_names)
	
	def add(self, file_path):
		name = Path(file_path).name
		filename = self.prefix + name
		new_location = os.path.join(self.path, filename)
		self.layoutAboutToBeChanged.emit()
		move(file_path, new_location)
		self.document_names.append(filename)
		self.layoutChanged.emit()
	
	def delete(self, row):
		file = self.document_names[row]
		path = os.path.join(self.path, file)
		try:
			os.remove(path)
			self.document_names.pop(row)
			self.layoutChanged.emit()
		except FileNotFoundError:
			pass
	
	def export(self, directory, row):
		filename = self.document_names[row]
		origin = os.path.join(self.path, filename)
		dest = os.path.join(directory, filename.replace(self.prefix, ''))
		move(origin, dest, copy=True)
	
	def key(self, filename: str):
		return filename.startswith(self.prefix)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		filename = self.document_names[index.row()]
		if role == Qt.DisplayRole:
			return filename.replace(self.prefix, '')
		elif role == Qt.DecorationRole:
			return QtGui.QIcon(':\pdf')
	
	def view(self, row):
		import subprocess
		filename = self.document_names[row]
		filepath = os.path.join(self.path, filename)
		subprocess.Popen((filename,), shell=True)


class AgentsDocumentModel(DocumentModel):
	
	def __init__(self, agent):
		self.agent = agent
		super().__init__()
	
	@property
	def path(self):
		path = os.path.join(dropbox_base, db.company_name(), 'Agents')
		if not os.path.exists(path):
			os.makedirs(path)
		return path
	
	@property
	def prefix(self):
		prefix = str(self.agent.id).zfill(3) + '_'
		for word in self.agent.fiscal_name.split():
			prefix += word[0]
		prefix += '_'
		return prefix


class PartnersDocumentModel(DocumentModel):
	
	def __init__(self, partner):
		self.partner = partner
		super().__init__()
	
	@property
	def prefix(self):
		prefix = str(self.partner.id).zfill(6) + '_'
		for word in self.partner.fiscal_name.split():
			prefix += word[0]
		prefix += '_'
		return prefix
	
	@property
	def path(self):
		path = os.path.join(dropbox_base, db.company_name(), 'Partners')
		if not os.path.exists(path):
			os.makedirs(path)
		return path


number_month_dict = {
	1: 'Enero',
	2: 'Febrero',
	3: 'Marzo',
	4: 'Abril',
	5: 'Mayo',
	6: 'Junio',
	7: 'Julio',
	8: 'Agosto',
	9: 'Septiembre',
	10: 'Octubre',
	11: 'Noviembre',
	12: 'Diciembre'
}

end_path = {
	1: os.path.join('Interior', 'Reg. General'),
	2: os.path.join('Interior', 'ISP'),
	5: os.path.join('Interior', 'REBU'),
	3: os.path.join('Importacion'),
	4: os.path.join('Intracomunitaria', 'Regimen General'),
	6: os.path.join('Intracomunitaria', 'Marginal VAT')
}


class ProformasSalesDocumentModel(DocumentModel):
	
	def __init__(self, obj):
		self.proforma = obj
		super().__init__()
	
	@property
	def path(self):
		_type = self.proforma.type
		path = os.path.join(
			dropbox_base,
			db.company_name(),
			'Proformas Emitidas',
			db.year(),
			number_month_dict[self.proforma.date.month],
			'Exportacion' if _type == 3 else end_path[_type]
		)
		
		if not os.path.exists(path):
			os.makedirs(path)
		return path
	
	@property
	def prefix(self):
		return self.proforma.doc_repr + '_'


class ProformasPurchasesDocumentModel(DocumentModel):
	
	def __init__(self, obj):
		self.proforma = obj
		super().__init__()
	
	@property
	def path(self):
		path = os.path.join(
			dropbox_base,
			db.company_name(),
			'Proformas Recibidas',
			db.year(),
			number_month_dict[self.proforma.date.month],
			end_path[self.proforma.type]
		)
		
		if not os.path.exists(path):
			os.makedirs(path)
		return path
	
	@property
	def prefix(self):
		return self.proforma.doc_repr + '_'


class InvoicesPurchasesDocumentModel(DocumentModel):
	
	def __init__(self, obj):
		self.invoice = obj.invoice
		super().__init__()
	
	@property
	def path(self):
		path = os.path.join(
			dropbox_base,
			db.company_name(),
			'Facturas Recibidas',
			db.year(),
			number_month_dict[self.invoice.date.month],
			end_path[self.invoice.type]
		)
		if not os.path.exists(path):
			os.makedirs(path)
		return path
	
	@property
	def prefix(self):
		return self.invoice.doc_repr + '_'


class InvoicesSalesDocumentModel(DocumentModel):
	
	def __init__(self, obj):
		self.invoice = obj.invoice
		super().__init__()
	
	@property
	def path(self):
		_type = self.invoice.type
		path = os.path.join(
			dropbox_base,
			db.company_name(),
			'Facturas Emitidas',
			db.year(),
			number_month_dict[self.invoice.date.month],
			'Exportacion' if _type == 3 else end_path[_type]
		)
		
		if not os.path.exists(path):
			os.makedirs(path)
		return path
	
	@property
	def prefix(self):
		return self.invoice.doc_repr


class PartnerModel(QtCore.QAbstractTableModel):
	CODE, TRADING_NAME, FISCAL_NAME, FISCAL_NUMBER, COUNTRY, CONTACT, PHONE, EMAIL, ACTIVE = \
		0, 1, 2, 3, 4, 5, 6, 7, 8
	
	def __init__(self, search_key=None):
		super().__init__()
		
		self._headerData = ['Code', 'Trading Name', 'Fiscal Name', 'Fiscal Number', 'Country', 'Contact', \
		                    'Phone', 'E-mail', 'Active']
		query = db.session.query(db.Partner).outerjoin(db.PartnerContact)
		if search_key:
			query = query.filter(
				or_(
					db.Partner.fiscal_name.contains(search_key),
					db.Partner.fiscal_number.contains(search_key),
					db.Partner.trading_name.contains(search_key),
					db.Partner.billing_country.contains(search_key),
				)
			)
		
		self.partners = query.all()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		partner = self.partners[index.row()]
		partner: db.Partner
		column = index.column()
		if role == Qt.DisplayRole:
			return self.col_to_data_map(column, partner)
		elif role == Qt.DecorationRole:
			if column == PartnerModel.ACTIVE:
				return QtGui.QIcon(':\greentick') if partner.active else \
					QtGui.QIcon(':\cross')
	
	def columnCount(self, index=QModelIndex()):
		return len(self._headerData)
	
	def rowCount(self, index=QModelIndex()):
		return len(self.partners)
	
	def headerData(self, section, orientation, role=Qt.DisplayRole):
		if role == Qt.TextAlignmentRole:
			if orientation == Qt.Horizontal:
				return Qt.AlignLeft | Qt.AlignVCenter
			return Qt.AlignRight | Qt.AlignVCenter
		if role != Qt.DisplayRole:
			return
		if orientation == Qt.Horizontal:
			return self._headerData[section]
	
	def sort(self, section, order):
		attr = {
			PartnerModel.CODE: 'id',
			PartnerModel.FISCAL_NAME: 'fiscal_name',
			PartnerModel.FISCAL_NUMBER: 'fiscal_number',
			PartnerModel.TRADING_NAME: 'trading_name',
			PartnerModel.COUNTRY: 'billing_country'
		}.get(section)
		
		if attr:
			self.layoutAboutToBeChanged.emit()
			self.partners.sort(key=operator.attrgetter(attr),
			                   reverse=True if Qt.AscendingOrder else False)
			self.layoutChanged.emit()
	
	def add(self, partner):
		db.session.add(partner)
		try:
			db.session.commit()
			self.partners.append(partner)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def delete(self, index):
		if not index.isValid():
			return
		row = index.row()
		candidate_partner = self.partners[row]
		db.session.delete(candidate_partner)
		try:
			db.session.commit()
			self.partners.remove(candidate_partner)
			self.layoutChanged.emit()
		except Exception as e:
			db.session.rollback()
			raise 			
	

	def clone(self, index):
		if not index.isValid():
			raise ValueError('Invalid index')
		
		partner = self.partners[index.row()]
		
		try:
			# clone partner
			partner_clone = utils.duplicate_db_object(partner)
			db.session.add(partner_clone)
			db.session.commit() # because i need the partner_id for taggin the object
			# so it can appear in the search bars.

			partner_clone.trading_name = partner.trading_name + f" CLONE - {partner_clone.id}"
			
			# UPDATE POINTERS:
			# PartnerDocument -> partner.documents
			# PartnerContact  -> partner.contacts
			# PartnerAccount  -> partner.accounts
			# ShippingAddress -> partner.shipping_addresses

			# TODO: parameterize the above method to include a 
			# dependencies: dict and set it dinamically 

			for document in partner.documents:
				clone = db.PartnerDocument()
				clone.partner = partner_clone 
				clone.name = document.name 
				clone.document = document.document
				clone.created_on = document.created_on

				session.add(clone)

			for account in partner.accounts:
				clone = db.PartnerAccount()
				clone.partner = partner 
				clone.bank_name 	= account.bank_name
				clone.iban 			= account.iban
				clone.swift 		= account.swift
				clone.bank_address  = account.bank_address
				clone.bank_postcode = account.bank_postcode
				clone.bank_city 	= account.bank_city
				clone.bank_state 	= account.bank_state
				clone.bank_country  = account.bank_country
				clone.bank_routing  = account.bank_routing
				clone.currency 		= account.currency
				
			for contact in partner.contacts:
				clone = db.PartnerContact(contact.name, contact.position, contact.phone, contact.email, contact.note)
				clone.preferred = contact.preferred
				clone.partner = partner_clone
				db.session.add(clone)

			for address in partner.shipping_addresses:
				clone = db.ShippingAddress()
				clone.partner = partner_clone 
				clone.line1 = address.line1
				clone.line2 = address.line2
				clone.city = address.city
				clone.state = address.state
				clone.zipcode = address.zipcode
				clone.country = address.country 

				db.session.add(clone)
		
			db.session.commit()
		except Exception as ex:
			raise PartnerCloningError('Failed to clone partner') from ex 


	def col_to_data_map(self, col, partner: db.Partner):
		
		return {
			PartnerModel.CODE: partner.id,
			PartnerModel.TRADING_NAME: partner.trading_name,
			PartnerModel.FISCAL_NAME: partner.fiscal_name,
			PartnerModel.FISCAL_NUMBER: partner.fiscal_number,
			PartnerModel.COUNTRY: partner.billing_country,
			PartnerModel.CONTACT: partner.contacts[0].name,
			PartnerModel.PHONE: partner.contacts[0].phone,
			PartnerModel.EMAIL: partner.contacts[0].email,
			PartnerModel.ACTIVE: 'YES' if partner.active else 'NO'
		}.get(col)


from db import ShippingAddress
from typing import List


class ShippingAddressModel(BaseTable, QtCore.QAbstractTableModel):
	LINE1, LINE2, CITY, STATE, ZIP, COUNTRY = 0, 1, 2, 3, 4, 5
	
	def __init__(self, view, addresses):
		super().__init__()
		self.view = view
		self.addresses: List[ShippingAddress] = addresses
		self.name = 'addresses'
		self._headerData = ['Line 1', 'Line 2', 'City', 'State', 'Zip', 'Country']
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			return {
				self.LINE1: 	self.addresses[row].line1,
				self.LINE2: 	self.addresses[row].line2,
				self.CITY: 		self.addresses[row].city,
				self.STATE: 	self.addresses[row].state,
				self.ZIP: 		self.addresses[row].zipcode,
				self.COUNTRY: 	self.addresses[row].country
			}.get(col)
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return False
		
		row, col = index.row(), index.column()
		if role == Qt.EditRole:
			address = self.addresses[row]
			if col == self.LINE1:
				address.line1 = value
			elif col == self.LINE2:
				address.line2 = value
			elif col == self.CITY:
				address.city = value
			elif col == self.STATE:
				address.state = value
			elif col == self.COUNTRY:
				address.country = value
			elif col == self.ZIP:
				address.zipcode = value
			else:
				return False
			self.dataChanged.emit(index, index)
			return True
		return False
	
	def insertRow(self, position, rows=1, index=QModelIndex()) -> bool:
		self.beginInsertRows(QModelIndex(), position, position + rows - 1)
		address = db.ShippingAddress()
		self.addresses.insert(position, address)
		self.endInsertRows()
		return True
	
	def removeRows(self, position, rows=1, index=QModelIndex()) -> bool:
		address = self.addresses[position]
		if len(address.sale_proformas) > 0: 
			raise ValueError('This address cannot be removed. It has associated sales') 
		
		# proceed with removal 		
		self.beginRemoveRows(QModelIndex(), position, position + rows - 1)
		self.addresses.pop(position)
		self.endRemoveRows()
		return True
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
	
	@property
	def valid(self):
		return all(e.valid for e in self.addresses)

	@property
	def empty(self):
		return len(self.addresses) == 0


class PartnerContactModel(QtCore.QAbstractTableModel):
	NAME, POSITION, PHONE, EMAIL, NOTE = 0, 1, 2, 3, 4
	
	def __init__(self, view, contacts):
		super().__init__()
		self._headerData = ['Name', 'Position', 'Phone', 'Email', 'Note']
		self.contacts = contacts
		self.view = view
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid() and 0 <= index.row() < len(self.contacts):
			return
		if role != Qt.DisplayRole:
			return
		row, col = index.row(), index.column()
		contact = self.contacts[row]
		if col == PartnerContactModel.NAME:
			return contact.name
		elif col == PartnerContactModel.POSITION:
			return contact.position
		elif col == PartnerContactModel.PHONE:
			return contact.phone
		elif col == PartnerContactModel.EMAIL:
			return contact.email
		elif col == PartnerContactModel.NOTE:
			return contact.note
	
	def rowCount(self, index=Qt.DisplayRole):
		return len(self.contacts)
	
	def columnCount(self, index=Qt.DisplayRole):
		return 5
	
	def headerData(self, section, orientation, role=Qt.DisplayRole):
		if role == Qt.DisplayRole:
			if orientation == Qt.Horizontal:
				return self._headerData[section]
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
	
	def setData(self, index, value, role=Qt.EditRole):
		if index.isValid() and 0 <= index.row() < len(self.contacts):
			contact = self.contacts[index.row()]
			column = index.column()
			if column == PartnerContactModel.NAME:
				contact.name = value
			elif column == PartnerContactModel.POSITION:
				contact.position = value
			elif column == PartnerContactModel.PHONE:
				contact.phone = value
			elif column == PartnerContactModel.EMAIL:
				contact.email = value
			elif column == PartnerContactModel.NOTE:
				contact.note = value
			self.dataChanged.emit(index, index)
			self.view.resizeColumnToContents(column)
			return True
		return False
	
	def insertRows(self, position, rows=1, index=QModelIndex()):
		self.beginInsertRows(QModelIndex(), position, position + rows - 1)
		contact = db.PartnerContact('', '', '', '', '')
		self.contacts.insert(position, contact)
		self.endInsertRows()
		return True
	
	def removeRows(self, position, rows=1, index=QModelIndex()):
		self.beginRemoveRows(QModelIndex(), position, position + rows - 1)
		self.contacts.pop(position)
		self.endRemoveRows()
		return True


class SaleInvoiceModel(BaseTable, QtCore.QAbstractTableModel):
	TYPENUM, RELATIONSHIP, DATE, ETA, PARTNER, AGENT, FINANCIAL, LOGISTIC, SENT, CANCELLED, OWING, \
		TOTAL, EXT, INWH, READY, PROFORMA, WARNING, SOLUNION = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17
	
	def __init__(self, filters=None, search_key=None, last=10):
		super().__init__()
		self._headerData = [
			'Type & Num', 'Sales/Credit', 'Date', 'ETA', 'Partner', 'Agent',
			'Financial', 'Logistic', 'Sent', 'Cancelled', 'Owing',
			'Total', 'Ext. Doc.', 'In WH', 'Ready to go', 'Proforma', 'Warning', 'Solunion'
		]
		
		self.name = 'invoices'
		
		query = db.session.query(db.SaleInvoice).join(db.SaleProforma). \
			join(db.Agent, db.Agent.id == db.SaleProforma.agent_id). \
			join(db.Partner, db.Partner.id == db.SaleProforma.partner_id). \
			where(db.SaleInvoice.date > utils.get_last_date(last))
		
		# TODO: search key, last, and filters
		
		if search_key:
			predicates = []
			predicates.extend(
				[
					db.Agent.fiscal_name.contains(search_key),
					db.Partner.fiscal_name.contains(search_key)
				]
			)
			
			try:
				date = utils.parse_date(search_key)
			except ValueError:
				pass
			else:
				predicates.extend(
					[
						db.SaleInvoice.eta == date,
						db.SaleInvoice.date == date
					]
				)
			
			try:
				n = int(search_key)
			except ValueError:
				pass
			else:
				predicates.append(db.SaleInvoice.number == n)
			
			query = query.where(or_(*predicates))
		
		if filters and filters['types']:
			query = query.where(db.SaleInvoice.type.in_(filters['types']))
		
		self.invoices = query.all()
		
		if filters:
			if filters['financial']:
				if 'notpaid' in filters['financial']:
					self.invoices = filter(lambda i: i.not_paid, self.invoices)
				
				if 'fullypaid' in filters['financial']:
					self.invoices = filter(lambda i: i.fully_paid, self.invoices)
				
				if 'partiallypaid' in filters['financial']:
					self.invoices = filter(lambda i: i.partially_paid, self.invoices)
			
			if filters['logistic']:
				if 'overflowed' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Overflowed',
						self.invoices
					)
				
				if 'empty' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Empty',
						self.invoices
					)
				
				if 'partially_processed' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Partially Prepared',
						self.invoices
					)
				
				if 'completed' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Completed',
						self.invoices
					)
			
			if filters['shipment']:
				
				if 'sent' in filters['shipment']:
					self.invoices = filter(
						lambda i: i.sent == 'Yes',
						self.invoices
					)
				
				if 'notsent' in filters['shipment']:
					self.invoices = filter(
						lambda i: i.sent == 'No',
						self.invoices
					)
			
			if filters['cancelled']:
				
				if 'cancelled' in filters:
					self.invoices = filter(
						lambda i: i.cancelled == 'Yes',
						self.invoices
					)
				
				if 'notcancelled' in filters:
					self.invoices = filter(
						lambda i: i.cancelled == 'No',
						self.invoices
					)
			
			if isinstance(self.invoices, filter):
				self.invoices = list(self.invoices)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:

		from datetime import date # I know. Bad, but i don't have the pattience to do it properly. 
		# This thing is cached, so no performance issues (more than already exist) 

		if not index.isValid():
			return
		row, col = index.row(), index.column()
		invoice = self.invoices[row]
		
		# Avoid double computations: display and decoration branches
		logistic_status_string = invoice.logistic_status_string
		financial_status_string: str = invoice.financial_status_string
		ready_status_string = invoice.ready

		if financial_status_string in ('Not Paid', 'Partially Paid') and int(invoice.solunion) > 0:
			solunion = invoice.solunion - ((date.today()-invoice.date)).days
		else:
			solunion = ''  

		if role == Qt.DisplayRole:
			return [
				invoice.doc_repr,
				invoice.self_referential_relationship,
				invoice.date.strftime('%d/%m/%Y'),
				invoice.eta.strftime('%d/%m/%Y'),
				invoice.partner_name,
				invoice.agent,
				financial_status_string,
				logistic_status_string,
				invoice.sent,
				invoice.cancelled,
				invoice.owing_string,
				invoice.total,
				invoice.external,
				invoice.inwh,
				ready_status_string,
				invoice.origin_proformas,
				invoice.warning,
				solunion
			][col]
		
		elif role == Qt.DecorationRole:
			
			if col == self.DATE or col == self.ETA:
				return QtGui.QIcon(':\calendar')
			
			elif col == self.FINANCIAL:
				if financial_status_string in ('Not Paid', 'Not Returned'):
					color = YELLOW
				elif financial_status_string in ('Returned', 'Paid'):
					color = GREEN
				elif financial_status_string.find('Partially') == 0:
					color = ORANGE
				elif financial_status_string.find('Over') == 0:
					color = RED
				return QtGui.QColor(color)
			
			elif col == self.LOGISTIC:
				if logistic_status_string == 'Empty':
					return QtGui.QColor(YELLOW)
				elif logistic_status_string == 'Overflowed':
					return QtGui.QColor(RED)
				elif logistic_status_string == 'Partially Prepared':
					return QtGui.QColor(ORANGE)
				elif logistic_status_string == 'Completed':
					return QtGui.QColor(GREEN)
			
			elif col == self.READY:
				return QtGui.QColor(GREEN if invoice.ready == 'Yes' else RED)
			
			elif col == self.AGENT:
				return QtGui.QIcon(':\\agents')
			
			elif col == self.PARTNER:
				return QtGui.QIcon(':\partners')
			
			elif col == self.SENT:
				return QtGui.QColor(GREEN if invoice.sent == 'Yes' else RED)
			
			elif col == self.CANCELLED:
				return QtGui.QColor(GREEN if not invoice.cancelled == 'Yes' else RED)
			
			elif col == self.SOLUNION:
				if solunion != '':
					if solunion > 0: 
						return QtGui.QColor(GREEN)
					elif solunion == 0:
						return QtGui.QColor(ORANGE)
					elif solunion < 0:
						return QtGui.QColor(RED) 
				

	# TODO: move logic from main gui here for towh button
	def to_warehouse(self, invoice):
		pass
	
	def sort(self, column: int, order: Qt.SortOrder = ...) -> None:
		reverse = True if order == Qt.AscendingOrder else False
		attrs = {
			self.TYPENUM: ('type', 'number'),
			self.DATE: ('date',),
			self.ETA: ('eta',),
			self.PARTNER: ('partner_name',),
			self.AGENT: ('agent',),
		}.get(column)
		
		if attrs:
			self.layoutAboutToBeChanged.emit()
			self.invoices.sort(key=operator.attrgetter(*attrs), reverse=reverse)
			self.layoutChanged.emit()
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return Qt.ItemIsEnabled
		
		if index.column() == self.WARNING:
			return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable
		else:
			return super().flags(index)
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return
		if role == Qt.EditRole:
			row, col = index.row(), index.column()
			if col == self.WARNING:
				for proforma in self.invoices[row].proformas:
					proforma.warning = value
				db.session.commit()
				self.layoutChanged.emit()
				return True
			return False
		return False
	
	def __getitem__(self, item):
		return self.invoices[item]


class SaleInvoiceLineModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SPEC, PUBLIC_CONDITION, QUANTITY, \
		PRICE, SUBTOTAL, TAX, TOTAL = 0, 1, 2, 3, 4, 5, 6, 7, 8
	
	def __init__(self, invoice):
		super().__init__()
		self.invoice = invoice
		self.name = 'lines'
		self._headerData = [
			'Description',
			'Condition',
			'Spec',
			'Public Condition',
			'Qty.',
			'Price ',
			'Subtotal',
			'Tax ',
			'Total'
		]
		
		# Credit notes are excluded
		self.lines = [line for proforma in invoice.proformas for line in proforma.advanced_lines or proforma.lines]
	
	def get_data(self, line, col):
		if col == self.CONDITION:
			return line.condition
		elif col == self.PUBLIC_CONDITION:
			return line.showing_condition
		elif col == self.SPEC:
			return line.spec
		elif col == self.QUANTITY:
			return line.quantity
		elif col == self.PRICE:
			return line.price
		elif col == self.TAX:
			return line.tax
		elif col == self.SUBTOTAL:
			return str(line.subtotal)
		elif col == self.TOTAL:
			return str(line.total)
		elif col == self.DESCRIPTION:
			if isinstance(line, db.AdvancedLine):
				return line.free_description or line.mixed_description or utils.description_id_map.inverse[line.item_id]
			elif isinstance(line, db.SaleProformaLine):
				return line.description or utils.description_id_map.inverse[line.item_id]
	
	def __getitem__(self, item):
		return self.lines[item]
	
	def delete(self, rows):
		for row in rows:
			line = self.lines[row]
			db.session.delete(line)
		
		db.session.flush()
		
		self.layoutAboutToBeChanged.emit()
		for row in reversed(list(rows)):
			del self.lines[row]
		self.layoutChanged.emit()
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		line = self.lines[row]
		if role == Qt.DisplayRole:
			return self.get_data(line, col)
	
	@property
	def quantity(self):
		return sum(line.quantity for line in self.lines)


class PurchaseInvoiceModel(BaseTable, QtCore.QAbstractTableModel):
	TYPENUM, DATE, ETA, PARTNER, AGENT, FINANCIAL, LOGISTIC, SENT, CANCELLED, OWING, \
		TOTAL, EXT, INWH, PROFORMA = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13
	
	def __init__(self, filters=None, search_key=None, last=10):
		super().__init__()
		self._headerData = [
			'Type & Num', 'Date', 'ETA', 'Partner', 'Agent',
			'Financial', 'Logistic', 'Sent', 'Cancelled', 'Owing',
			'Total', 'Ext. Doc.', 'In WH', 'Proforma'
		]
		
		self.name = 'invoices'
		query = db.session.query(db.PurchaseInvoice).join(db.PurchaseProforma). \
			join(db.Agent, db.Agent.id == db.PurchaseProforma.agent_id). \
			join(db.Partner, db.Partner.id == db.PurchaseProforma.partner_id). \
			where(db.PurchaseInvoice.date > utils.get_last_date(last))
		
		if search_key:
			predicates = []
			predicates.extend(
				[
					db.Agent.fiscal_name.contains(search_key),
					db.Partner.fiscal_name.contains(search_key)
				]
			)
			try:
				date = utils.parse_date(search_key)
			except ValueError:
				pass
			else:
				predicates.extend(
					[
						db.PurchaseInvoice.eta == date,
						db.PurchaseInvoice.date == date
					]
				)
			try:
				n = int(search_key)
			except ValueError:
				pass
			else:
				predicates.append(db.PurchaseInvoice.number == n)
			
			query = query.where(or_(*predicates))
		
		if filters and filters['types']:
			query = query.where(db.PurchaseInvoice.type.in_(filters['types']))
		
		self.invoices = query.all()
		
		if filters:
			if filters['financial']:
				if 'notpaid' in filters['financial']:
					self.invoices = filter(lambda i: i.not_paid, self.invoices)
				
				if 'fullypaid' in filters['financial']:
					self.invoices = filter(lambda i: i.fully_paid, self.invoices)
				
				if 'partiallypaid' in filters['financial']:
					self.invoices = filter(lambda i: i.partially_paid, self.invoices)
			
			if filters['logistic']:
				if 'overflowed' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Overflowed',
						self.invoices
					)
				
				if 'empty' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Empty',
						self.invoices
					)
				
				if 'partially_processed' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Partially received',
						self.invoices
					)
				
				if 'completed' in filters['logistic']:
					self.invoices = filter(
						lambda i: i.logistic_status_string == 'Completed',
						self.invoices
					)
			
			if filters['shipment']:
				
				if 'sent' in filters['shipment']:
					self.invoices = filter(
						lambda i: i.sent == 'Yes',
						self.invoices
					)
				
				if 'notsent' in filters['shipment']:
					self.invoices = filter(
						lambda i: i.sent == 'No',
						self.invoices
					)
			
			if filters['cancelled']:
				
				if 'cancelled' in filters:
					self.invoices = filter(
						lambda i: i.cancelled == 'Yes',
						self.invoices
					)
				
				if 'notcancelled' in filters:
					self.invoices = filter(
						lambda i: i.cancelled == 'No',
						self.invoices
					)
			
			if isinstance(self.invoices, filter):
				self.invoices = list(self.invoices)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		invoice = self.invoices[row]
		
		# Avoid double computations: display and decoration branches
		logistic_status_string = invoice.logistic_status_string
		financial_status_string = invoice.financial_status_string
		
		if role == Qt.DisplayRole:
			return [
				invoice.doc_repr,
				invoice.date.strftime('%d/%m/%Y'),
				invoice.eta.strftime('%d/%m/%Y'),
				invoice.partner_name,
				invoice.agent,
				invoice.financial_status_string,
				logistic_status_string,
				invoice.sent,
				invoice.cancelled,
				invoice.owing_string,
				invoice.total,
				invoice.external,
				invoice.inwh,
				invoice.origin_proformas
			][col]
		
		elif role == Qt.DecorationRole:
			if col == self.DATE or col == self.ETA:
				return QtGui.QIcon(':\calendar')
			
			elif col == self.FINANCIAL:
				if financial_status_string == 'Not Paid':
					return QtGui.QColor(YELLOW)
				elif financial_status_string == 'Paid':
					return QtGui.QColor(GREEN)
				elif financial_status_string == 'Partially Paid':
					return QtGui.QColor(ORANGE)
				elif financial_status_string == 'They Owe':
					return QtGui.QColor(RED)
			
			elif col == self.PARTNER:
				return QtGui.QIcon(':\partners')
			elif col == self.AGENT:
				return QtGui.QIcon(':\\agents')
			
			elif col == self.LOGISTIC:
				if logistic_status_string == 'Empty':
					return QtGui.QColor(YELLOW)
				elif logistic_status_string == 'Overflowed':
					return QtGui.QColor(RED)
				elif logistic_status_string == 'Partially received':
					return QtGui.QColor(ORANGE)
				elif logistic_status_string == 'Completed':
					return QtGui.QColor(GREEN)
			
			elif col == self.SENT:
				sent = invoice.sent
				if sent == 'Yes':
					return QtGui.QColor(GREEN)
				elif sent == 'Partially':
					return QtGui.QColor(ORANGE)
				elif sent == 'No':
					return QtGui.QColor(RED)
			
			elif col == self.CANCELLED:
				if invoice.cancelled == 'Yes':
					return QtGui.QColor(RED)
				elif invoice.cancelled == 'Partially':
					return QtGui.QColor(ORANGE)
				elif invoice.cancelled == 'No':
					return QtGui.QColor(GREEN)
	
	def sort(self, section: int, order: Qt.SortOrder = ...) -> None:
		reverse = True if order == Qt.AscendingOrder else False
		if section == self.TYPENUM:
			self.layoutAboutToBeChanged.emit()
			self.invoices.sort(key=lambda i: (i.type, i.number), reverse=reverse)
			self.layoutChanged.emit()
		else:
			attr = {
				self.DATE: 'date',
				self.PARTNER: 'partner_name',
				self.AGENT: 'agent',
				self.ETA: 'eta'
			}.get(section)
			
			if attr:
				self.layoutAboutToBeChanged.emit()
				self.invoices.sort(key=operator.attrgetter(attr), reverse=reverse)
				self.layoutChanged.emit()
	
	def __getitem__(self, item):
		return self.invoices[item]


class PurchaseInvoiceLineModel(BaseTable, QtCore.QAbstractTableModel):
	
	def __init__(self, invoice):
		super().__init__()
		self.invoice = invoice
		self.name = 'lines'
		self._headerData = [
			'Description',
			'Condition',
			'Spec',
			'Qty.',
			'Price ',
			'Subtotal',
			'Tax ',
			'Total'
		]
		self.lines = [line for proforma in invoice.proformas for line in proforma.lines]
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		line = self.lines[row]
		
		if role == Qt.DisplayRole:
			return [
				line.description or utils.description_id_map.inverse[line.item_id],
				line.condition,
				line.spec,
				line.quantity,
				str(line.price),
				str(line.subtotal),
				str(line.tax),
				str(line.total)
			][col]
	
	def delete(self, rows):
		for row in rows:
			line = self.lines[row]
			db.session.delete(line)
		
		db.session.flush()
		
		self.layoutAboutToBeChanged.emit()
		for row in reversed(list(rows)):
			del self.lines[row]
		self.layoutChanged.emit()
	
	def __getitem__(self, item):
		return self.lines[item]
	
	@property
	def quantity(self):
		return sum(line.quantity for line in self.lines)


def buildReceptionLine(line, reception):
	reception_line = db.ReceptionLine()
	
	reception_line.condition = line.condition
	reception_line.spec = line.spec
	reception_line.quantity = line.quantity
	
	reception_line.item_id = line.item_id
	reception_line.description = line.description
	
	if reception_line.item_id or reception_line.description in \
			utils.descriptions:
		reception_line.reception = reception


# reception is attached to session, will cascade-commit the lines.


def update_purchase_warehouse(proforma):
	if proforma.reception is None:
		return False
	
	warehouse_lines = set(proforma.reception.lines)
	proforma_lines = set(proforma.lines)
	
	for line in warehouse_lines.difference(proforma_lines):
		line.quantity = 0
		if len(line.series) == 0:
			db.session.delete(line)
	
	for line in proforma_lines.difference(warehouse_lines):
		if line.item_id or line.description in utils.descriptions:
			buildReceptionLine(line, proforma.reception)
	
	for proforma_line in proforma_lines:
		for reception_line in proforma.reception.lines:
			if reception_line == proforma_line:
				reception_line.quantity = proforma_line.quantity
	try:
		db.session.commit()
	except Exception as ex:
		db.session.rollback()
		raise


def ship_several(proformas, tracking=None):
	for proforma in proformas:
		proforma.tracking = tracking
		proforma.sent = True
	
	db.session.commit()


class PurchaseProformaModel(BaseTable, QtCore.QAbstractTableModel):
	TYPE_NUM, DATE, ETA, PARTNER, AGENT, FINANCIAL, LOGISTIC, SENT, CANCELLED, \
		OWING, TOTAL, EXTERNAL, INVOICED, IN_WAREHOUSE = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13
	
	def __init__(self, filters=None, search_key=None, last=10):
		super().__init__()
		self._headerData = ['Type & Num', 'Date', 'ETA', 'Partner', 'Agent', \
		                    'Financial', 'Logistic', 'Sent', 'Cancelled', 'Owing', 'Total',
		                    'External', 'Invoiced', 'In Warehouse']
		self.name = 'proformas'
		
		query = db.session.query(db.PurchaseProforma).select_from(db.Agent, db.Partner). \
			where(
			db.Agent.id == db.PurchaseProforma.agent_id,
			db.Partner.id == db.PurchaseProforma.partner_id
		)
		
		query = db.session.query(db.PurchaseProforma). \
			join(db.Partner, db.Partner.id == db.PurchaseProforma.partner_id). \
			join(db.Agent, db.Agent.id == db.PurchaseProforma.agent_id)
		
		query = query.where(db.PurchaseProforma.date > utils.get_last_date(last))
		
		if search_key:
			
			predicates = []
			
			predicates.extend(
				[
					db.Agent.fiscal_name.contains(search_key),
					db.Partner.fiscal_name.contains(search_key)
				]
			)
			
			try:
				date = utils.parse_date(search_key)
			except ValueError:
				pass
			else:
				predicates.extend(
					[
						db.PurchaseProforma.eta == date,
						db.PurchaseProforma.date == date
					]
				)
			try:
				n = int(search_key)
			except ValueError:
				pass
			else:
				predicates.append(db.PurchaseProforma.number == n)
			
			query = query.where(or_(*predicates))
		
		if filters:
			self.proformas = query.all()
			
			if filters['types']:
				self.proformas = filter(lambda p: p.type in filters['types'], self.proformas)
			
			if filters['financial']:
				if 'notpaid' in filters['financial']:
					self.proformas = filter(lambda p: p.not_paid, self.proformas)
				if 'fullypaid' in filters['financial']:
					self.proformas = filter(lambda p: p.fully_paid, self.proformas)
				if 'partiallypaid' in filters['financial']:
					self.proformas = filter(lambda p: p.partially_paid, self.proformas)
			
			if filters['logistic']:
				if 'overflowed' in filters['logistic']:
					self.proformas = filter(lambda p: p.overflowed, self.proformas)
				if 'empty' in filters['logistic']:
					self.proformas = filter(lambda p: p.empty, self.proformas)
				if 'partially_processed' in filters['logistic']:
					self.proformas = filter(lambda p: p.partially_processed, self.proformas)
				if 'completed' in filters['logistic']:
					self.proformas = filter(lambda p: p.completed, self.proformas)
			
			if filters['shipment']:
				if 'sent' in filters['shipment']:
					self.proformas = filter(lambda p: p.sent, self.proformas)
				if 'notsent' in filters['shipment']:
					self.proformas = filter(lambda p: not p.sent, self.proformas)
			
			if filters['cancelled']:
				if 'cancelled' in filters['cancelled']:
					self.proformas = filter(lambda p: p.cancelled, self.proformas)
				if 'notcancelled' in filters['cancelled']:
					self.proformas = filter(lambda p: not p.cancelled, self.proformas)
			
			if isinstance(self.proformas, filter):
				self.proformas = list(self.proformas)
		else:
			self.proformas = query.all()
	
	def __getitem__(self, index):
		return self.proformas[index]
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		proforma = self.proformas[row]
		
		if role == Qt.DisplayRole:
			if col == PurchaseProformaModel.TYPE_NUM:
				s = str(proforma.type) + '-' + str(proforma.number).zfill(6)
				return s
			elif col == PurchaseProformaModel.DATE:
				return proforma.date.strftime('%d/%m/%Y')
			elif col == PurchaseProformaModel.ETA:
				return proforma.eta.strftime('%d/%m/%Y')
			
			elif col == self.INVOICED:
				try:
					return proforma.invoice.doc_repr
				except AttributeError:
					return 'No'
			
			elif col == self.IN_WAREHOUSE:
				return 'Yes' if proforma.reception is not None else 'No'
			
			elif col == PurchaseProformaModel.PARTNER:
				return proforma.partner_name
			elif col == PurchaseProformaModel.AGENT:
				return proforma.agent.fiscal_name
			elif col == PurchaseProformaModel.FINANCIAL:
				return proforma.financial_status_string
			
			elif col == PurchaseProformaModel.LOGISTIC:
				return proforma.logistic_status_string
			
			elif col == PurchaseProformaModel.SENT:
				return "Yes" if proforma.sent else "No"
			
			elif col == PurchaseProformaModel.CANCELLED:
				return 'Yes' if proforma.cancelled else 'No'
			
			elif col == PurchaseProformaModel.OWING:
				sign = ' -€' if proforma.eur_currency else ' $'
				owing = round(proforma.total_debt - proforma.total_paid, 2)
				return str(owing) + sign
			
			elif col == PurchaseProformaModel.TOTAL:
				sign = ' -€' if proforma.eur_currency else ' $'
				return str(proforma.total_debt) + sign
			elif col == self.EXTERNAL:
				return proforma.external
		
		elif role == Qt.DecorationRole:
			if col == PurchaseProformaModel.FINANCIAL:
				if proforma.not_paid:
					return QtGui.QColor(YELLOW)
				elif proforma.fully_paid:
					return QtGui.QColor(GREEN)
				elif proforma.partially_paid:
					return QtGui.QColor(ORANGE)
				elif proforma.overpaid:
					return QtGui.QColor(RED)
			
			elif col == PurchaseProformaModel.DATE or col == PurchaseProformaModel.ETA:
				return QtGui.QIcon(':\calendar')
			elif col == PurchaseProformaModel.PARTNER:
				return QtGui.QIcon(':\partners')
			elif col == PurchaseProformaModel.AGENT:
				return QtGui.QIcon(':\\agents')
			
			elif col == PurchaseProformaModel.LOGISTIC:
				if proforma.empty:
					return QtGui.QColor(YELLOW)
				elif proforma.overflowed:
					return QtGui.QColor(RED)
				elif proforma.partially_processed:
					return QtGui.QColor(ORANGE)
				elif proforma.completed:
					return QtGui.QColor(GREEN)
			
			elif col == PurchaseProformaModel.CANCELLED:
				return QtGui.QColor(RED) if proforma.cancelled else \
					QtGui.QColor(GREEN)
			
			elif col == PurchaseProformaModel.SENT:
				return QtGui.QColor(GREEN) if proforma.sent else \
					QtGui.QColor(RED)
	
	def sort(self, section, order):
		reverse = True if order == Qt.AscendingOrder else False
		if section == PurchaseProformaModel.TYPE_NUM:
			self.layoutAboutToBeChanged.emit()
			self.proformas.sort(key=lambda p: (p.type, p.number), reverse=reverse)
			self.layoutChanged.emit()
		
		else:
			attr = {
				PurchaseProformaModel.DATE: 'date',
				PurchaseProformaModel.PARTNER: 'partner.trading_name',
				PurchaseProformaModel.AGENT: 'agent.fiscal_name',
				PurchaseProformaModel.SENT: 'sent',
				PurchaseProformaModel.ETA: 'eta'
			}.get(section)
			
			if attr:
				self.layoutAboutToBeChanged.emit()
				self.proformas.sort(key=operator.attrgetter(attr), reverse=reverse)
				self.layoutChanged.emit()
	
	def add(self, proforma):
		db.session.add(proforma)
		db.session.commit()
		self.layoutAboutToBeChanged.emit()
		self.proformas.append(proforma)
		self.layoutChanged.emit()
	
	def cancel(self, indexes):
		rows = {index.row() for index in indexes}
		for row in rows:
			p = self.proformas[row]
			if not p.cancelled:
				p.cancelled = True
		try:
			# Update advanced lines depending on this purchase
			ids = [line.id for line in p.lines]
			db.session.query(db.AdvancedLine). \
				where(db.AdvancedLine.origin_id.in_(ids)). \
				update({db.AdvancedLine.quantity: 0})
			
			db.session.commit()
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def associateInvoice(self, rows: set):
		for r1, r2 in combinations(rows, r=2):
			p1, p2 = self.proformas[r1], self.proformas[r2]
			if not all((
					p1.type == p2.type, p1.agent_id == p2.agent_id, p1.partner_id == p2.partner_id,
					p1.eur_currency == p2.eur_currency
			)):
				raise ValueError('Incompatible proformas')
		
		for row in rows:
			break  # Peek an element (set is not subscriptable)
		
		any_proforma = self.proformas[row]
		
		next_num = get_next_num(db.PurchaseInvoice, any_proforma.type)
		invoice = db.PurchaseInvoice(any_proforma.type, next_num)
		
		invoice.date = any_proforma.date
		invoice.eta = any_proforma.eta
		
		for row in rows:
			proforma = self.proformas[row]
			proforma.invoice = invoice
		
		try:
			db.session.commit()
			return proforma.invoice
		except:
			db.session.rollback()
			raise
	
	def ship(self, proforma, tracking=None):
		proforma.tracking = tracking
		proforma.sent = True
		try:
			db.session.commit()
		except:
			db.session.rollback()
			raise
	
	def toWarehouse(self, proforma, note):
		reception = db.Reception(proforma, note)
		db.session.add(reception)
		
		for line in proforma.lines:
			buildReceptionLine(line, reception)
		try:
			db.session.commit()
		except:
			db.session.rollback()
			raise


def item_key(line):
	return line.item_id in utils.description_id_map.inverse


def line_with_stock_key(line):
	if line.item_id:
		return True
	else:
		if isinstance(line, db.AdvancedLine):
			if line.free_description:
				return False
			else:
				return True
	return False


def build_associated_reception(sale_proforma):
	try:
		origin_id = line = sale_proforma.advanced_lines[0].origin_id
	except IndexError:
		return
	else:
		proforma = db.session.query(
			db.PurchaseProforma
		).join(db.PurchaseProformaLine).where(
			db.PurchaseProformaLine.id == origin_id
		).first()
	
	reception = db.Reception(proforma, note='Created automatically', auto=True)
	db.session.add(reception)
	
	for line in proforma.lines:
		buildReceptionLine(line, reception)
	
	try:
		db.session.commit()
	except IntegrityError:
		db.session.rollback()


def build_expedition_line(line, expedition):
	exp_line = db.ExpeditionLine()
	exp_line.condition = line.condition
	exp_line.spec = line.spec
	exp_line.item_id = line.item_id
	exp_line.quantity = line.quantity
	
	exp_line.expedition = expedition


def update_sale_warehouse(proforma):
	# Restore the original __eq__ method
	from db import SaleProformaLine, original_sale_proforma_line_eq
	SaleProformaLine.__eq__ = original_sale_proforma_line_eq
	
	if proforma.expedition is None:
		return
	
	wh_lines = set(proforma.expedition.lines)
	pr_lines = set(
		[d for line in proforma.advanced_lines for d in line.definitions]
		or proforma.lines
		or proforma.advanced_lines
	)
	
	for line in wh_lines.difference(pr_lines):
		line.quantity = 0
		if len(line.series) == 0:
			db.session.delete(line)
	
	for line in pr_lines.difference(wh_lines):
		if item_key(line):
			build_expedition_line(line, proforma.expedition)
	
	for pline in pr_lines:
		for whline in wh_lines:
			if pline == whline:
				whline.quantity = pline.quantity
	
	try:
		db.session.commit()
	except Exception as ex:
		db.session.rollback()
		raise


class SaleProformaModel(BaseTable, QtCore.QAbstractTableModel):
	TYPE_NUM, DATE, PARTNER, AGENT, FINANCIAL, LOGISTIC, SENT, \
		CANCELLED, OWING, TOTAL, ADVANCED, DEFINED, READY, IN_WAREHOUSE, \
		WARNING, INVOICED = \
		0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15
	
	def __init__(self, search_key=None, filters=None, last=10):
		super().__init__()
		self._headerData = ['Type & Num', 'Date', 'Partner', 'Agent',
		                    'Financial', 'Logistic', 'Sent', 'Cancelled',
		                    'Owes', 'Total', 'Presale', 'Defined', 'Ready To Go',
		                    'In WH', 'Warning', 'Inv.']
		
		self.proformas = []
		self.name = 'proformas'
		
		query = db.session.query(db.SaleProforma). \
			join(db.Agent, db.Agent.id == db.SaleProforma.agent_id). \
			join(db.Partner, db.Partner.id == db.SaleProforma.partner_id). \
			join(db.Warehouse, db.Warehouse.id == db.SaleProforma.warehouse_id, isouter=True). \
			join(db.SaleInvoice, db.SaleInvoice.id == db.SaleProforma.sale_invoice_id, isouter=True). \
			where(db.SaleProforma.date >= utils.get_last_date(last))
		
		if search_key:
			predicates = []
			predicates.extend(
				[
					db.Agent.fiscal_name.contains(search_key),
					db.Partner.fiscal_name.contains(search_key),
					db.Warehouse.description.contains(search_key)
				]
			)
			
			try:
				date = utils.parse_date(search_key)
			except ValueError:
				pass
			else:
				predicates.append(db.SaleProforma.date == date)
			
			try:
				n = int(search_key)
			except ValueError:
				pass
			else:
				predicates.append(db.SaleProforma.number == n)
			
			query = query.where(or_(*predicates))
		
		if filters and filters['types']:
			query = query.where(db.SaleProforma.type.in_(filters['types']))
		
		if filters:
			self.proformas = query.all()
			
			if filters['financial']:
				if 'notpaid' in filters['financial']:
					self.proformas = filter(lambda p: p.not_paid, self.proformas)
				if 'fullypaid' in filters['financial']:
					self.proformas = filter(lambda p: p.fully_paid, self.proformas)
				if 'partiallypaid' in filters['financial']:
					self.proformas = filter(lambda p: p.partially_paid, self.proformas)
			
			if filters['logistic']:
				if 'overflowed' in filters['logistic']:
					self.proformas = filter(lambda p: p.overflowed, self.proformas)
				if 'empty' in filters['logistic']:
					self.proformas = filter(lambda p: p.empty, self.proformas)
				if 'partially_processed' in filters['logistic']:
					self.proformas = filter(lambda p: p.partially_processed, self.proformas)
				if 'completed' in filters['logistic']:
					self.proformas = filter(lambda p: p.completed, self.proformas)
			
			if filters['shipment']:
				if 'sent' in filters['shipment']:
					self.proformas = filter(lambda p: p.sent, self.proformas)
				if 'notsent' in filters['shipment']:
					self.proformas = filter(lambda p: not p.sent, self.proformas)
			
			if filters['cancelled']:
				if 'cancelled' in filters['cancelled']:
					self.proformas = filter(lambda p: p.cancelled, self.proformas)
				if 'notcancelled' in filters['cancelled']:
					self.proformas = filter(lambda p: not p.cancelled, self.proformas)
			
			if isinstance(self.proformas, filter):
				self.proformas = list(self.proformas)
		
		else:
			self.proformas = query.all()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		proforma = self.proformas[row]
		
		financial_status_string = proforma.financial_status_string
		sign = ' EUR' if proforma.eur_currency else ' USD'
		if role == Qt.DisplayRole:
			if col == self.TYPE_NUM:
				return proforma.doc_repr
			elif col == self.DATE:
				return proforma.date.strftime('%d/%m/%Y')
			elif col == self.PARTNER:
				return proforma.partner_name
			elif col == self.AGENT:
				return proforma.agent.fiscal_name.split()[0]
			
			elif col == self.WARNING:
				return proforma.warning
			
			elif col == self.INVOICED:
				try:
					return proforma.invoice.doc_repr
				except AttributeError:
					return 'No'
			
			elif col == self.IN_WAREHOUSE:
				return 'Yes' if proforma.expedition is not None else 'No'
			
			elif col == self.FINANCIAL:
				return financial_status_string
			
			elif col == self.LOGISTIC:
				
				if proforma.is_credit_note:
					return 'Completed'
				
				if proforma.empty:
					return 'Empty'
				elif proforma.overflowed:
					return 'Overflowed'
				elif proforma.partially_processed:
					return 'Partially Prepared'
				elif proforma.completed:
					return 'Completed'
			
			elif col == self.SENT:
				return 'Yes' if proforma.sent else 'No'
			elif col == self.CANCELLED:
				return 'Yes' if proforma.cancelled else 'No'
			elif col == self.OWING:
				return f"{proforma.owing} {sign}"
			
			elif col == self.TOTAL:
				return f"{proforma.total_debt} {sign}"
			
			elif col == self.ADVANCED:
				return 'Yes' if proforma.advanced_lines else 'No'
			
			elif col == self.DEFINED:
				line_iter = filter(line_with_stock_key, iter(proforma.lines or proforma.advanced_lines))
				return 'Yes' if all(line.defined for line in line_iter) else 'No'
			
			elif col == self.READY:
				return 'Yes' if proforma.ready else 'No'
		
		elif role == Qt.DecorationRole:
			if col == self.FINANCIAL:
				if financial_status_string in ('Not Paid', 'Not Returned'):
					color = YELLOW
				elif financial_status_string in ('Returned', 'Paid'):
					color = GREEN
				elif financial_status_string.find('Partially') == 0:
					color = ORANGE
				elif financial_status_string.find('Over') == 0:
					color = RED
				return QtGui.QColor(color)
			
			elif col == self.DATE:
				return QtGui.QIcon(':\calendar')
			elif col == self.PARTNER:
				return QtGui.QIcon(':\partners')
			elif col == self.AGENT:
				return QtGui.QIcon(':\\agents')
			elif col == self.LOGISTIC:
				
				if proforma.is_credit_note:
					return QtGui.QColor(GREEN)
				
				if proforma.empty:
					return QtGui.QColor(YELLOW)
				elif proforma.overflowed:
					return QtGui.QColor(RED)
				elif proforma.partially_processed:
					return QtGui.QColor(ORANGE)
				elif proforma.completed:
					return QtGui.QColor(GREEN)
			
			elif col == self.SENT:
				return QtGui.QColor(GREEN if proforma.sent else YELLOW)
			elif col == self.CANCELLED:
				return QtGui.QColor(RED if proforma.cancelled else GREEN)
			elif col == self.READY:
				return QtGui.QColor(GREEN if proforma.ready else YELLOW)
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		if index.column() == self.WARNING:
			# return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
			return Qt.ItemIsEnabled | Qt.ItemIsEditable | Qt.ItemIsSelectable
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return
		if role == Qt.EditRole:
			proforma = self.proformas[index.row()]
			proforma.warning = value
			db.session.commit()
			return True
		return False
	
	def __getitem__(self, index):
		return self.proformas[index]
	
	def sort(self, section, order):
		reverse = True if order == Qt.AscendingOrder else False
		if section == self.TYPE_NUM:
			self.layoutAboutToBeChanged.emit()
			self.proformas.sort(key=lambda p: (p.type, p.number), reverse=reverse)
			self.layoutChanged.emit()
		
		else:
			attr = {
				self.DATE: 'date',
				self.PARTNER: 'partner.trading_name',
				self.AGENT: 'agent.fiscal_name',
				self.SENT: 'sent',
				self.READY: 'ready'
			}.get(section)
			
			if attr:
				self.layoutAboutToBeChanged.emit()
				self.proformas.sort(key=operator.attrgetter(attr), reverse=reverse)
				self.layoutChanged.emit()
	
	def add(self, proforma):
		db.session.add(proforma)
		db.session.commit()
		self.layoutAboutToBeChanged.emit()
		self.proformas.append(proforma)
		self.layoutChanged.emit()
	
	def cancel(self, indexes):
		rows = {index.row() for index in indexes}
		for row in rows:
			p = self.proformas[row]
			if not p.cancelled:
				p.cancelled = True
		try:
			db.session.commit()
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def build_invoice_from_proforma(self, proforma, wh_order):
		_next = get_next_num(db.SaleInvoice, proforma.type)
		proforma.invoice = db.SaleInvoice(proforma.type, _next)
		proforma.invoice.wh_incoming_rma_id = wh_order.id
		db.session.commit()
		
		return proforma.invoice
	
	def associateInvoice(self, rows: set, bypass_vies=False):
		if any(bool(self.proformas[row].credit_note_lines) for row in rows):
			raise ValueError('Credit Note already exists')
		
		for r1, r2 in combinations(rows, r=2):
			p1, p2 = self.proformas[r1], self.proformas[r2]
			if not all((
				p1.type == p2.type, p1.agent_id == p2.agent_id, p1.partner_id == p2.partner_id,
				p1.eur_currency == p2.eur_currency, 
				p1.shipping_address_id == p2.shipping_address_id
			)):
				raise ValueError('Incompatible proformas')
		
		for row in rows:
			break  # Peek an element (set is not subscriptable)
		
		any_proforma = self.proformas[row]
		
		if not bypass_vies:
			number, country = any_proforma.partner.fiscal_number, any_proforma.partner.billing_country
			
			from utils import get_country_code
			code = get_country_code(country)
			print('cpuntry_code=', code)
			print('fiscal_number=', number)
			if number.startswith(code):
				try:
					request = vies.Vies().request(number)
					print(repr(request))
					print(request) 

					if request.valid:
						register = db.ViesRequest(request.requestDate, request.valid, number)
						db.session.add(register)
				
				except (vies.ViesValidationError, vies.ViesError, vies.ViesHTTPError):
					raise
		
		next_num = get_next_num(db.SaleInvoice, any_proforma.type)
		
		invoice = db.SaleInvoice(any_proforma.type, next_num)
		
		for row in rows:
			proforma = self.proformas[row]
			proforma.invoice = invoice
		
		try:
			db.session.commit()
			return proforma.invoice
		except:
			db.session.rollback()
			raise
	
	def ready(self, indexes):
		rows = {index.row() for index in indexes}
		for row in rows:
			p = self.proformas[row]
			p.ready = not p.ready  # Invert sense
		
		self.layoutAboutToBeChanged.emit()
		db.session.commit()
		self.layoutChanged.emit()
	
	def ready_several(self, proformas):
		for p in proformas:
			p.ready = not p.ready
		
		self.layoutAboutToBeChanged.emit()
		db.session.commit()
		self.layoutChanged.emit()
	
	def ship_several(self, proformas, tracking):
		
		for proforma in proformas:
			proforma.tracking = tracking
			proforma.sent = True
		
		self.layoutAboutToBeChanged.emit()
		db.session.commit()
		self.layoutChanged.emit()
	
	def ship(self, proforma, tracking):
		proforma.tracking = tracking
		proforma.sent = True
		self.layoutAboutToBeChanged.emit()
		db.session.commit()
		self.layoutChanged.emit()
	
	def to_warehouse_several(self, proformas, note):
		pass
	
	def toWarehouse(self, proforma, note):
		fast = True
		expedition = db.Expedition(proforma)
		proforma.warning = note
		
		if proforma.advanced_lines:
			for line in proforma.advanced_lines:
				if any((
						line.mixed_description,
						line.spec == 'Mix',
						line.condition == 'Mix'
				)) and not line.definitions:
					raise ValueError("You have to define the presale first")
			
			# Defined
			else:
				lines = []
				for line in proforma.advanced_lines:
					if line.definitions:
						lines.extend(list(definition for definition in line.definitions))
						if fast:
							fast = False
					
					else:
						lines.append(line)
			
			if fast:
				expedition.from_sale_type = db.FAST
				build_associated_reception(proforma)
			
			else:
				expedition.from_sale_type = db.DEFINED
		
		else:
			expedition.from_sale_type = db.NORMAL
			lines = proforma.lines
		
		if not lines:
			return
		
		db.session.add(expedition)
		
		for line in lines:
			exp_line = db.ExpeditionLine()
			exp_line.item_id = line.item_id
			exp_line.condition = line.condition
			exp_line.spec = line.spec
			exp_line.quantity = line.quantity
			try:
				exp_line.showing_condition = line.showing_condition
			except AttributeError:
				# Arreglar definition showing condition
				exp_line.showing_condition = ''
			
			if line.item_id:
				exp_line.expedition = expedition
		try:
			db.session.commit()
		except Exception:
			db.session.rollback()
			raise
	
	def stock_available(self, wh_id, lines):
		query = db.session.query(
			func.count(db.Imei.imei).label('quantity'),
			db.Imei.item_id, db.Imei.condition,
			db.Imei.spec
		).join(db.Warehouse).where(
			db.Warehouse.id == wh_id
		).group_by(
			db.Imei.item_id, db.Imei.spec, db.Imei.condition
		)
		return any((
			line.quantity > stock.quantity and line == stock
			for line in lines
			for stock in query
		))
	
	def pline_eline_equal(self, pline, eline):
		return all((
			pline.item_id == eline.item_id,
			pline.condition == eline.condition,
			pline.spec == eline.spec
		))
	
	def difference(self, proforma, direction='proforma_expedition'):
		
		if direction == 'proforma_expedition':
			iter_a = filter(item_key, proforma.advanced_lines or proforma.lines)
			iter_b = iter(proforma.expedition.lines)
		
		elif direction == 'expedition_proforma':
			iter_a = iter(proforma.expedition.lines)
			iter_b = filter(item_key, proforma.advanced_lines or proforma.lines)
		
		for pline in iter_a:
			for eline in iter_b:
				if self.pline_eline_equal(pline, eline):
					break
			else:
				yield pline


def __my__eq__(self, other):
	if self.id is not None and other.id is not None:
		return self.id == other.id
	elif self.description is not None and other.description is not None:
		return self.description == other.description
	else:
		return all((
			self.item_id == other.item_id,
			self.condition == other.condition,
			self.spec == other.spec
		))


class OrganizedLines:
	
	def __init__(self, lines):
		self.instrumented_lines = lines
		self.organized_lines = self.organize_lines(lines)
		
		from db import SaleProformaLine
		
		SaleProformaLine.__eq__ = __my__eq__
	
	@property
	def lines(self):
		return self.instrumented_lines
	
	@staticmethod
	def get_next_mix():
		import uuid
		return str(uuid.uuid4())
	
	def delete(self, i, j=None):
		update_stock_view = True
		if j is None:
			lines = self.organized_lines.pop(i)
			
			for line in lines:
				if line.item_id is None:  # exploit the run
					update_stock_view = False
				
				self.instrumented_lines.remove(line)
		
		else:
			line = self.organized_lines[i].pop(j)
			self.instrumented_lines.remove(line)
		
		self.organized_lines = [e for e in self.organized_lines if len(e) > 0]
		
		db.session.flush()
		
		return update_stock_view
	
	def delete_all(self):
		for i in range(len(self.organized_lines)):
			try:
				self.delete(i)
			except IndexError:
				continue
	
	def append(self, price, ignore_spec, tax, showing, *stocks, row=None):
		if len(stocks) == 0:
			raise ValueError("Provide stocks")
		
		# Update quantity
		hit = False
		for stock in stocks:
			for line in self.lines:
				if stock == line:
					line.quantity += stock.request
					hit = True
					break
		if hit:
			db.session.flush()
			return
		
		# Check if stocks are compatible for mixing between them:
		for a, b in product(stocks, stocks):
			if not utils.mixing_compatible(a, b):
				raise ValueError('Incompatible Mixing One to One.')
		
		# Adding
		if row is None:
			new_lines = self.build_lines_from_stocks(
				price, ignore_spec, tax,
				showing, *stocks
			)
			
			mix_id = self.get_next_mix()
			for line in new_lines:
				line.mix_id = mix_id
			
			self.organized_lines.append(new_lines)
		# Updating:
		else:
			if self.backward_compatible(stocks, row):
				new_lines = self.update_organized_lines(
					price, ignore_spec,
					tax, showing, *stocks,
					row=row)
			else:
				raise ValueError(f"Incompatible mixing with line {row + 1}")
		
		self.instrumented_lines.extend(new_lines)
		
		db.session.flush()
	
	def build_lines_from_stocks(self, price, ignore_spec, tax, showing, *stocks, mix_id=None):
		return [
			self.build_line(price, ignore_spec, tax, showing, stock, mix_id=mix_id) for stock in stocks
		]
	
	def update_organized_lines(self, price, ignore_spec, tax, showing, *stocks, row=-1):
		previous_lines = self.organized_lines[row]
		previous_mix_id = previous_lines[0].mix_id
		
		new_lines = self.build_lines_from_stocks(price, ignore_spec, tax, showing, *stocks, mix_id=previous_mix_id)
		previous_lines.extend(new_lines)
		
		return new_lines
	
	def backward_compatible(self, stocks, row):
		previous_lines = self.organized_lines[row]
		try:
			result = all((
				utils.mixing_compatible(stock, line)
				for stock, line in product(previous_lines, stocks)
			
			))
			return result
		
		except TypeError:
			line = previous_lines
			result = all((
				utils.mixing_compatible(stock, line)
				for stock in stocks
			))
			return result
	
	def insert_free(self, description, quantity, price, tax):
		
		line = db.SaleProformaLine()
		line.description = description
		line.quantity = quantity
		line.price = price
		line.tax = tax
		line.mix_id = self.get_next_mix()
		
		lines = [line, ]  # Homogeneo
		
		self.organized_lines.append(lines)
		
		self.instrumented_lines.extend(lines)
		
		db.session.flush()
	
	@staticmethod
	def organize_lines(lines):
		mix_ids = list(dict.fromkeys([line.mix_id for line in lines]))
		matrix = []
		for mix_id in mix_ids:
			aux = []
			for line in lines:
				if line.mix_id == mix_id:
					aux.append(line)
			matrix.append(aux)
		return matrix
	
	def repr(self, row, col):
		lines = self.organized_lines[row]
		
		diff_items = {line.item_id for line in lines}
		
		# For free lines, ugly, but works and uses previous code
		if diff_items == {None}:
			line = lines[0]
			return self.simple_line_repr(line, col)
		
		diff_conditions = {line.condition for line in lines}
		diff_specs = {line.spec for line in lines}
		
		if len(diff_items) == 1:
			description = utils.description_id_map.inverse[diff_items.pop()]
		else:
			description = utils.build_description(lines)
		
		if len(diff_conditions) == 1:
			condition = diff_conditions.pop()
		else:
			condition = 'Mix'
		showing_condition = lines[0].showing_condition
		
		if len(diff_specs) == 1:
			spec = diff_specs.pop()
		else:
			spec = 'Mix'
		
		ignore = 'Yes' if lines[0].ignore_spec else 'No'
		price = lines[0].price
		quantity = sum([line.quantity for line in lines])
		subtotal = round(price * quantity, 2)
		tax = lines[0].tax
		total = round(subtotal * (1 + tax / 100), 2)
		return {
			SaleProformaLineModel.DESCRIPTION: description,
			SaleProformaLineModel.CONDITION: condition,
			SaleProformaLineModel.SHOWING_CONDITION: showing_condition,
			SaleProformaLineModel.SPEC: spec,
			SaleProformaLineModel.IGNORING_SPEC: ignore,
			SaleProformaLineModel.SUBTOTAL: str(subtotal),
			SaleProformaLineModel.PRICE: str(price),
			SaleProformaLineModel.QUANTITY: str(quantity),
			SaleProformaLineModel.TAX: str(tax),
			SaleProformaLineModel.TOTAL: str(total)
		}.get(col)
	
	@staticmethod
	def simple_line_repr(line, col):
		total = round(line.quantity * line.price * (1 + line.tax / 100), 1)
		subtotal = round(line.quantity * line.price, 2)
		ignore_spec = 'Yes' if line.ignore_spec else 'No'
		showing_condition = line.showing_condition or line.condition
		return {
			SaleProformaLineModel.DESCRIPTION: line.description \
			                                   or utils.description_id_map.inverse.get(line.item_id, ''), 
			SaleProformaLineModel.CONDITION: line.condition,
			SaleProformaLineModel.SHOWING_CONDITION: showing_condition,
			SaleProformaLineModel.SPEC: line.spec,
			SaleProformaLineModel.IGNORING_SPEC: ignore_spec,
			SaleProformaLineModel.SUBTOTAL: str(subtotal),
			SaleProformaLineModel.PRICE: str(line.price),
			SaleProformaLineModel.QUANTITY: str(line.quantity),
			SaleProformaLineModel.TAX: str(line.tax),
			SaleProformaLineModel.TOTAL: str(total)
		}.get(col)
	
	@staticmethod
	def build_line(price, ignore, tax, showing, stock, mix_id=None):
		line = db.SaleProformaLine()
		line.item_id = stock.item_id
		line.condition = stock.condition
		line.spec = stock.spec
		line.ignore_spec = ignore
		line.showing_condition = showing
		line.quantity = stock.request
		line.price = price
		line.tax = tax
		line.mix_id = mix_id
		return line
	
	def update_condition(self, row, condition):
		lines = self.organized_lines[row]
		if isinstance(lines, Iterable):
			for line in lines:
				line.showing_condition = condition
		else:
			lines.showing_condition = condition
		return True
	
	def update_spec(self, row, spec):
		spec = spec.lower()
		flag = spec != 'no'
		lines = self.organized_lines[row]
		if isinstance(lines, Iterable):
			for line in lines:
				line.ignore_spec = flag
		else:
			lines.ignore_spec = flag
		return True
	
	def update_price(self, row, price):
		lines = self.organized_lines[row]
		if isinstance(lines, Iterable):
			for line in lines:
				line.price = price
		else:
			lines.price = price
		return True

	def update_description(self, row, description):
		lines = self.organized_lines[row]
		if isinstance(lines, Iterable):
			for line in lines:
				line.description = description
		else:
			lines.description = description
		return True

	
	def update_quantity(self, row, quantity):
		lines = self.organized_lines[row]
		if isinstance(lines, Iterable):
			for line in lines:
				line.quantity = quantity
		else:
			lines.quantity = quantity
		return True


	def update_tax(self, row, tax):
		lines = self.organized_lines[row]
		if isinstance(lines, Iterable):
			for line in lines:
				line.tax = tax
		else:
			lines.tax = tax
		return True
	
	def __getitem__(self, index):
		try:
			return self.organized_lines[index]
		except IndexError:
			return None
	
	def __len__(self):
		return len(self.organized_lines)
	
	def __bool__(self):
		return bool(self.instrumented_lines)


class SaleProformaLineModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SHOWING_CONDITION, SPEC, IGNORING_SPEC, \
		QUANTITY, PRICE, SUBTOTAL, TAX, TOTAL = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9
	
	def __init__(self, proforma, form):
		super().__init__()
		self.form = form
		self._headerData = ['Description', 'Condition', 'Public Condt.(Editable)', 'Spec', \
		                    'Ignoring Spec?(Editable)', 'Qty.', 'Price(Editable)', 'Subtotal', 'Tax(Editable)', 'Total']
		
		self.proforma = proforma
		self.organized_lines = OrganizedLines(proforma.lines)
		self.name = 'organized_lines'
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		if role == Qt.DisplayRole:
			return self.organized_lines.repr(row, column)
	
	def __iter__(self):
		return iter(self.organized_lines.lines)
	
	def delete_all(self):
		self.layoutAboutToBeChanged.emit()
		self.organized_lines.delete_all()
		self.layoutChanged.emit()
	
	def get_price(self, row):
		# Remember : Matrix Structure : [[...], [...], ...]
		return self.organized_lines[row][0].price
	
	@property
	def quantity(self):
		return sum(line.quantity for line in self.lines)
	
	@property
	def lines(self):
		return self.organized_lines.lines
	
	@property
	def tax(self):
		return round(sum(line.quantity * line.price * line.tax / 100 for line in self.lines), 2)
	
	@property
	def subtotal(self):
		return round(sum(line.quantity * line.price for line in self.lines), 2)
	
	@property
	def total(self):
		return round(self.tax + self.subtotal, 2)
	
	def add(self, price, ignore_spec, tax, showing, *stocks, row=None):
		self.layoutAboutToBeChanged.emit()
		self.organized_lines.append(
			price,
			ignore_spec,
			tax, showing,
			*stocks,
			row=row
		)
		self.layoutChanged.emit()
	
	def insert_free(self, description, quantity, price, tax):
		self.layoutAboutToBeChanged.emit()
		self.organized_lines.insert_free(
			description,
			quantity,
			price,
			tax
		)
		self.layoutChanged.emit()
	
	def delete(self, i, j=None):
		self.layoutAboutToBeChanged.emit()
		update = self.organized_lines.delete(i, j)
		self.layoutChanged.emit()
		return update
	
	def actual_lines_from_mixed(self, row):
		try:
			lines = self.organized_lines[row]
			if lines[0].item_id:
				return lines
		except TypeError:
			pass
	
	def __bool__(self):
		return bool(self.organized_lines)
	
	def reset(self):
		self.layoutAboutToBeChanged.emit()
		self.organized_lines = OrganizedLines([])
		self.layoutChanged.emit()
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid():
			return False

		row, column = index.row(), index.column()
		updated = False

		if role==Qt.EditRole and self.is_free_line(index) and self.free_line_editable_column(column):
			
			if column == self.DESCRIPTION:
				updated = self.organized_lines.update_description(row, value)
			elif column == self.QUANTITY:
				try:
					quantity = int(value)
					if quantity < 0:
						raise ValueError
				except ValueError:
					return False
				else:
					updated = self.organized_lines.update_quantity(row, quantity)
			elif column == self.PRICE:
				try:
					price = float(value)
					if price < 0:
						raise ValueError
				except ValueError:
					return False
				else:
					updated = self.organized_lines.update_price(row, price)
			elif column == self.TAX:
				try:
					tax = int(value)
					if tax not in (0, 4, 10, 21):
						return False
				except ValueError:
					return False
				else:
					updated = self.organized_lines.update_tax(row, tax)
			else:
				return False 

		if role == Qt.EditRole:
			if column == self.__class__.SHOWING_CONDITION:
				return self.organized_lines.update_condition(row, value)
			elif column == self.__class__.IGNORING_SPEC:
				if value.lower() not in ('yes', 'no'):
					return False
				else:
					updated = self.organized_lines.update_spec(row, value)
			elif column == self.__class__.TAX:
				try:
					tax = int(value)
					if tax not in (0, 4, 10, 21):
						return False
				except ValueError:
					return False
				else:
					updated = self.organized_lines.update_tax(row, tax)
			elif column == self.__class__.PRICE:
				try:
					price = float(value)
					if price < 0:
						raise ValueError
				except ValueError:
					return False
				else:
					updated = self.organized_lines.update_price(row, price)
			else:
				return False 

		db.session.commit()
		if updated:
			self.form.update_totals()
			return True
			
		return False
	
	def editable_column(self, column):
		return column in (
			self.SHOWING_CONDITION,
			self.IGNORING_SPEC,
			self.PRICE,
			self.TAX
		)
	
	def free_line_editable_column(self, column):
		return column in (self.DESCRIPTION, self.QUANTITY, self.PRICE, self.TAX)

	def is_free_line(self, index):
		if not index.isValid():
			return False 
		row = index.row() 
		try:
			lines = self.organized_lines[row]
		except IndexError:
			return False 

		return (line:=next(iter(lines), False)) and getattr(line, 'item_id', object()) == None

	def flags(self, index):
		if not index.isValid():
			return Qt.ItemFlags(~Qt.ItemIsEnabled)
		
		if not self.is_free_line(index) and self.editable_column(index.column()): 
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)

		if self.is_free_line(index) and self.free_line_editable_column(index.column()):
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)

		return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def sync_with_warehouse(self):
		import itertools as it 
		for pline, eline in it.product(self.proforma.lines, self.proforma.expedition.lines):
			if all((
				pline.item_id == eline.item_id,
				pline.condition == eline.condition,
				pline.spec == eline.spec
			)):
				pline.quantity = eline.processed_series
				eline.quantity = eline.processed_series
				break
		
		db.session.commit()


class ActualLinesFromMixedModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SPEC, REQUEST = 0, 1, 2, 3
	
	def __init__(self, lines=None):
		super().__init__()
		self._headerData = ['Description', 'Condition', 'spec', 'Requested Quantity']
		self.name = 'lines'
		self.lines = lines or []
	
	def sort(self, section, order):
		attr = {
			self.DESCRIPTION: 'item_id',
			self.CONDITION: 'condition',
			self.SPEC: 'spec',
		}.get(section)
		
		if attr:
			reverse = True if order == Qt.AscendingOrder else False
			self.layoutAboutToBeChanged.emit()
			self.lines.sort(key=operator.attrgetter(attr), reverse=reverse)
			self.layoutChanged.emit()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		line = self.lines[row]
		if role == Qt.DisplayRole:
			if column == self.DESCRIPTION:
				return utils.description_id_map.inverse[line.item_id]
			elif column == self.CONDITION:
				return line.condition
			elif column == self.SPEC:
				return line.spec
			elif column == self.REQUEST:
				return str(line.quantity)
	
	def reset(self):
		self.layoutAboutToBeChanged.emit()
		self.lines = []
		self.layoutChanged.emit()


class ProductModel(BaseTable, QtCore.QAbstractTableModel):
	MPN, MANUFACTURER, CATEGORY, MODEL, CAPACITY, COLOR, WEIGHT, BATTERY_WEIGHT, HAS_SERIE = \
		0, 1, 2, 3, 4, 5, 6, 7, 8
	
	def __init__(self):
		
		super().__init__()
		self._headerData = ['MPN', 'Manufacturer', 'Category', 'Model',
		                    'Capacity', 'Color', 'Weight', 'Battery Weight', 'Has Serie']
		self.name = 'items'
		self.items = db.session.query(db.Item).all()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		item = self.items[index.row()]
		col = index.column()
		if role != Qt.DisplayRole:
			return
		else:
			return {
				self.MPN: item.mpn,
				self.MANUFACTURER: item.manufacturer,
				self.CATEGORY: item.category,
				self.MODEL: item.model,
				self.CAPACITY: item.capacity,
				self.COLOR: item.color,
				self.HAS_SERIE: 'Yes' if item.has_serie else 'No',
				self.WEIGHT: str(item.weight),
				self.BATTERY_WEIGHT: str(item.battery_weight)
			}.get(col)
	
	def addItem(self, mpn, manufacturer, category, model, capacity, color, weight, battery_weight, has_serie):
		item = db.Item(mpn, manufacturer, category, model, capacity, color, weight, battery_weight, has_serie)
		db.session.add(item)
		try:
			db.session.commit()
			self.items.append(item)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def removeItem(self, index):
		if not index.isValid():
			return
		row = index.row()
		candidate = self.items[row]
		db.session.delete(candidate)
		try:
			db.session.commit()
			del self.items[row]
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid():
			return False
		if '|' in value or '?' in value or 'Mix' in value:
			return False
		if role == Qt.EditRole:
			row, column = index.row(), index.column()
			item = self.items[row]
			if column == self.MPN:
				item.mpn = value
			elif column == self.MANUFACTURER:
				if not value.strip():
					return False  # These specific fields cannot be empty
				item.manufacturer = value
			elif column == self.CATEGORY:
				if not value.strip():
					return False
				item.category = value
			elif column == self.MODEL:
				if not value.strip():
					return False
				item.model = value
			elif column == self.CAPACITY:
				item.capacity = value
			elif column == self.COLOR:
				item.color = value
			elif column == self.HAS_SERIE:
				if value.lower() not in ('yes', 'no'):
					return False
				item.has_serie = True if value.lower() == 'yes' else False
			elif column == self.WEIGHT:
				try:
					decimal.Decimal(value)
				except decimal.InvalidOperation:
					return False
				else:
					item.weight = value
					return True
			
			elif column == self.BATTERY_WEIGHT:
				try:
					decimal.Decimal(value)
				except decimal.InvalidOperation:
					return False
				else:
					item.battery_weight = value
					return True
			try:
				db.session.commit()
			except IntegrityError:  # UNIQUE VIOLATION
				db.session.rollback()
				return False
			return True
		return False
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
	
	def sort(self, section, order):
		attr = {
			self.MPN: 'mpn',
			self.MANUFACTURER: 'manufacturer',
			self.CATEGORY: 'category',
			self.MODEL: 'model',
			self.CAPACITY: 'capacity',
			self.COLOR: 'color',
			self.HAS_SERIE: 'has_serie',
			self.WEIGHT: 'weight',
			self.BATTERY_WEIGHT: 'battery_weight'
		}.get(section)
		if attr:
			self.layoutAboutToBeChanged.emit()
			self.items = sorted(
				self.items,
				key=operator.attrgetter(attr),
				reverse=True if order == Qt.DescendingOrder else False
			)
			self.layoutChanged.emit()
	
	def __iter__(self):
		return iter(self.items)


def update_advanced_line_after_purchase_line_update(newquantity, origin_id):
	# Not yet purchase line persisted and given an id
	# remember make a flush not commit
	# if the changes will be rolled back or committed
	
	if not origin_id: return
	
	lines = db.session.query(db.AdvancedLine). \
		where(db.AdvancedLine.origin_id == origin_id). \
		order_by(db.AdvancedLine.id).all()
	
	total, first = 0, True
	for i, line in enumerate(lines):
		if total <= newquantity:
			total += line.quantity
			if total > newquantity:
				lines[i].quantity = line.quantity + newquantity - total
				continue
		if total > newquantity:
			lines[i].quantity = 0
	
	db.session.flush()


class PurchaseProformaLineModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SPEC, QUANTITY, PRICE, SUBTOTAL, \
		TAX, TOTAL = 0, 1, 2, 3, 4, 5, 6, 7
	
	def __init__(self, lines=None, form=None):
		super().__init__()
		self._headerData = ['Description', 'Condition', 'Spec', 'Qty.(Editable)',
		                    'Price (Editable)', 'Subtotal', 'Tax (Editable)', 'Total']
		self.name = 'lines'
		self.lines = lines
		self.form = form
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row = index.row()
		try:
			self.lines[row]
		except IndexError:
			return
		line = self.lines[row]
		col = index.column()
		if role == Qt.DisplayRole:
			total = round(line.quantity * line.price * (1 + line.tax / 100), 2)
			subtotal = round(line.quantity * line.price, 2)
			# Allowing mixed and precised lines:
			if col == 0:
				if line.description is not None:
					return line.description
				else:
					return utils.description_id_map.inverse[line.item_id]
			return {
				self.__class__.CONDITION: line.condition,
				self.__class__.SPEC: line.spec,
				self.__class__.QUANTITY: str(line.quantity),
				self.__class__.PRICE: str(line.price),
				self.__class__.SUBTOTAL: str(subtotal),
				self.__class__.TAX: str(line.tax),
				self.__class__.TOTAL: str(total)
			}.get(col)
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid():
			return False
		row, column = index.row(), index.column()
		if not self.editable_column(column):
			return False
		try:
			line = self.lines[row]
		except IndexError:
			return False
		if role == Qt.EditRole:
			if column == self.__class__.PRICE:
				try:
					value = float(value)
				except:
					return False
				else:
					line.price = value
					self.form._updateTotals()
					return True
			elif column == self.__class__.QUANTITY:
				try:
					value = int(value)
				except:
					return False
				else:
					
					if line.quantity > value:
						update_advanced_line_after_purchase_line_update(value, line.id)
					
					line.quantity = value
					self.form._updateTotals()
					return True
			elif column == self.__class__.TAX:
				try:
					value = int(value)
					if value not in (0, 4, 10, 21):
						return False
				except:
					return False
				else:
					line.tax = value
					self.form._updateTotals()
					return True
			else:
				return False
	
	def editable_column(self, col):
		return col in (
			self.__class__.PRICE,
			self.__class__.QUANTITY,
			self.__class__.TAX
		)
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		if self.editable_column(index.column()):
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def is_stock_relevant(self, line):
		return line.item_id or line.description in utils.descriptions
	
	@property
	def tax(self):
		return round(sum(line.quantity * line.price * line.tax / 100 for line in self.lines), 2)
	
	@property
	def subtotal(self):
		return round(sum(line.quantity * line.price for line in self.lines), 2)
	
	@property
	def total(self):
		return round(self.tax + self.subtotal, 2)
	
	@property
	def quantity(self):
		return sum(line.quantity for line in self.lines)
	
	def add(self, description, condition, spec, quantity, price, tax):
		line = db.PurchaseProformaLine()
		try:
			line.item_id = utils.description_id_map[description]
		except KeyError:
			line.description = description
		line.condition = condition or None
		line.spec = spec or None
		line.quantity = quantity
		line.tax = tax
		line.price = price
		
		# Dont apply mean for non stock relevant lines:
		if line.item_id is None:
			self.lines.append(line)
			self.layoutChanged.emit()
		else:
			updated = self.update_if_pre_exists(line)
			if not updated:
				self.lines.append(line)
			self.layoutChanged.emit()
	
	def update_if_pre_exists(self, new_line):
		
		# PQn = p1*q1 + p2*q2 + ··· + pn*qn Recursively
		# Qn = q1 + q2 + ··· + qn
		# pn = PQn / Qn
		# Can be determined recursively, no need to rememeber terms
		# PQn = pn-1 * qn-1 + qn*pn
		# Qn = qn-1 + qn
		
		for line in self.lines:
			if all((
					new_line.condition == line.condition,
					new_line.item_id == line.item_id,
					new_line.description == line.description,
					new_line.spec == line.spec
			)):
				new_quantity = line.quantity + new_line.quantity
				new_price = (line.quantity * line.price + new_line.quantity * new_line.price) \
				            / new_quantity
				line.quantity = new_quantity
				line.price = new_price
				return True
		
		return False
	
	def delete(self, indexes):
		rows = {index.row() for index in indexes}
		for row in sorted(rows, reverse=True):
			try:
				del self.lines[row]
			except:
				pass
		self.layoutChanged.emit()
	
	def delete_if_not_stock_relevant(self, rows):
		deleted = False
		for row in sorted(rows, reverse=True):
			line = self.lines[row]
			if self.is_stock_relevant(line):
				continue
			else:
				del self.lines[row]  # Should not rise Index Error
				deleted = True
		if deleted:
			self.layoutChanged.emit()
	
	def save(self, proforma):
		if not self.lines: return
		for line in self.lines:
			line.proforma = proforma
		try:
			db.session.commit()
		except:
			db.session.rollback()
			raise
	
	def __bool__(self):
		return bool(len(self.lines))


class PaymentModel(BaseTable, QtCore.QAbstractTableModel):
	DATE, AMOUNT, RATE, NOTE, PROFORMA = 0, 1, 2, 3, 4
	
	def __init__(self, proforma, sale, form):
		super().__init__()
		self.proforma = proforma
		self._headerData = ['Date', 'Amount', 'Rate', 'Info', 'Proforma']
		self.name = 'payments'
		self.form = form
		if sale:
			self.Payment = db.SalePayment
		else:
			self.Payment = db.PurchasePayment
		
		self.payments = db.session.query(self.Payment). \
			where(self.Payment.proforma.has(id=proforma.id)).all()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		payment = self.payments[index.row()]
		column = index.column()
		
		if role == Qt.DisplayRole:
			if column == self.DATE:
				return payment.date.strftime('%d/%m/%Y')
			elif column == self.AMOUNT:
				return str(payment.amount)
			elif column == self.NOTE:
				return payment.note
			elif column == self.RATE:
				return payment.rate
			elif column == self.PROFORMA:
				return payment.proforma.doc_repr
		
		elif role == Qt.DecorationRole:
			if column == self.__class__.DATE:
				return QtGui.QIcon(':\calendar')
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid(): return False
		row, column = index.row(), index.column()
		payment = self.payments[row]
		if role == Qt.EditRole:
			if column == self.DATE:
				try:
					d = utils.parse_date(value)
					payment.date = d
					return True
				except ValueError:
					return False
			elif column == self.AMOUNT:
				try:
					v = float(value.replace(',', '.'))
					payment.amount = v
					self.form.updateOwing()
					return True
				except ValueError:
					return False
			elif column == self.NOTE:
				payment.note = value[0:255]
				return True
			elif column == self.RATE:
				try:
					v = float(value.replace(',', '.'))
					payment.rate = v
					self.form.updateOwing()
					return True
				except ValueError:
					return False
		return False
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
	
	def add(self, date, amount, rate, note):
		payment = self.Payment(date, amount, rate, note, self.proforma)
		db.session.add(payment)
		try:
			db.session.commit()
			self.payments.append(payment)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def delete(self, indexes):
		rows = {index.row() for index in indexes}
		for row in rows:
			payment = self.payments[row]
			if payment.note.startswith('CN'):
				raise ValueError("You can't delete this Payment. Remove from applied credit notes")
			db.session.delete(payment)
		try:
			db.session.commit()
		except:
			db.session.rollback()
			raise
		else:
			for row in sorted(rows, reverse=True):
				del self.payments[row]
			self.layoutChanged.emit()
	
	def sort(self, section, order):
		reverse = True if order == Qt.AscendingOrder else False
		attr = {
			self.__class__.DATE: 'date',
			self.__class__.AMOUNT: 'amount',
			self.__class__.NOTE: 'note'
		}.get(section)
		if attr:
			self.layoutAboutToBeChanged.emit()
			self.payments.sort(key=operator.attrgetter(attr), reverse=reverse)
			self.layoutChanged.emit()
	
	@property
	def paid(self):
		return sum([p.amount for p in self.payments])


class ExpenseModel(BaseTable, QtCore.QAbstractTableModel):
	DATE, AMOUNT, NOTE = 0, 1, 2
	
	def __init__(self, proforma, sale):
		super().__init__()
		self.proforma = proforma
		self._headerData = ['Date', 'Amount', 'Info']
		self.name = 'expenses'
		if sale:
			self.Expense = db.SaleExpense
		else:
			self.Expense = db.PurchaseExpense
		
		self.expenses = db.session.query(self.Expense). \
			where(self.Expense.proforma.has(id=proforma.id)).all()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		expense = self.expenses[index.row()]
		colum = index.column()
		if role == Qt.DisplayRole:
			if colum == 0:
				return expense.date.strftime('%d/%m/%Y')
			elif colum == 1:
				return str(expense.amount)
			elif colum == 2:
				return expense.note
		elif role == Qt.DecorationRole:
			if colum == 0:
				return QtGui.QIcon(':\calendar')
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid(): return False
		row, column = index.row(), index.column()
		expense = self.expenses[row]
		if role == Qt.EditRole:
			if column == self.__class__.DATE:
				try:
					d = utils.parse_date(value)
					expense.date = d
					return True
				except ValueError:
					return False
			elif column == self.__class__.AMOUNT:
				try:
					v = float(value.replace(',', '.'))
					expense.amount = v
					self.form.updateOwing()
					return True
				except ValueError:
					return False
			elif column == self.__class__.NOTE:
				expense.note = value[0:255]
				return True
		return False
	
	def sort(self, section, order):
		reverse = True if order == Qt.AscendingOrder else False
		attr = {
			self.__class__.DATE: 'date',
			self.__class__.AMOUNT: 'amount',
			self.__class__.NOTE: 'note'
		}.get(section)
		if attr:
			self.layoutAboutToBeChanged.emit()
			self.expenses.sort(key=operator.attrgetter(attr), reverse=reverse)
			self.layoutChanged.emit()
	
	def add(self, date, amount, info):
		expense = self.Expense(date, amount, info, self.proforma)
		db.session.add(expense)
		try:
			db.session.commit()
			self.expenses.append(expense)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def delete(self, indexes):
		rows = {index.row() for index in indexes}
		for row in rows:
			expense = self.expenses[row]
			db.session.delete(expense)
		try:
			db.session.commit()
		except:
			db.session.rollback()
			raise
		else:
			for row in sorted(rows, reverse=True):
				del self.expenses[row]
			self.layoutChanged.emit()
	
	@property
	def spent(self):
		return round(sum([expense.amount for expense in self.expenses]), 2)


from functools import wraps


def change_layout_and_commit(func):
	wraps(func)
	
	def wrapper(self, *args, **kwargs):
		self.layoutAboutToBeChanged.emit()
		r = func(self, *args, **kwargs)
		self.commit()
		self.layoutChanged.emit()
		return r
	
	return wrapper


class SerieModel(QtCore.QAbstractListModel):
	## OOOOOO branch dhl
	def __init__(self, line, expedition):
		super().__init__()
		self.expedition = expedition
		self.line = line
		
		if not utils.has_serie(line):
			self.series = []
		else:
			self.series = db.session.query(db.ExpeditionSerie).join(db.ExpeditionLine). \
				where(db.ExpeditionSerie.line_id == line.id).all()
			
			self.series_at_expedition_level = self.get_series_at_expedition_level()
	
	def get_series_at_expedition_level(self):
		return {
			r[0] for r in db.session.query(db.ExpeditionSerie.serie).
			join(db.ExpeditionLine).join(db.Expedition).
			where(db.Expedition.id == self.expedition.id)
		}
	
	def serie_present_in_dependant_purchase(self, serie):
		return db.session.execute(
			db.purchase_level_query,
			params={'proforma_id': self.expedition.proforma_id, 'serie': serie}
		).scalar()
	
	def add(self, line, _serie):
		if _serie in self:
			raise SeriePresentError
		
		if self.serie_present_in_dependant_purchase(_serie):
			from exceptions import SeriePresentAtPurchaseSpace
			raise SeriePresentAtPurchaseSpace
		
		serie = db.ExpeditionSerie()
		serie.serie = _serie
		serie.line = line
		db.session.add(serie)
		try:
			db.session.commit()
			self.series.append(serie)
			self.series_at_expedition_level = self.get_series_at_expedition_level()
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def delete_all(self):
		for expedition_serie in self.series:
			db.session.delete(expedition_serie)
		try:
			db.session.commit()
			self.series = []
			self.series_at_expedition_level = self.get_series_at_expedition_level()
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def select_invented_from_imeis(self, limit=0):
		
		query = db.session.query(db.Imei.imei).where(
			db.Imei.warehouse_id == self.expedition.proforma.warehouse_id,
			db.Imei.spec == self.line.spec,
			db.Imei.condition == self.line.condition,
			db.Imei.item_id == self.line.item_id
		).limit(limit)
		
		return [r.imei for r in query]
	
	def select_from_expedition_series(self, *, limit=0):
		query = db.session.query(db.ExpeditionSerie).where(
			db.ExpeditionSerie.line_id == self.line.id
		).limit(limit)
		
		return [e for e in query]
	
	def handle_invented(self, difference):
		if difference == 0:
			return
		
		elif difference < 0:  # user wants to correct, invent new series
			difference = abs(difference)
			for expedition_serie in self.select_from_expedition_series(limit=difference):
				db.session.delete(expedition_serie)
			
			try:
				db.session.commit()
			except:
				db.session.rollback()
				raise
		
		elif difference > 0:  # User wants to process, select and
			
			fictitious_series = []
			for invented in self.select_invented_from_imeis(limit=difference):
				serie = db.ExpeditionSerie()
				serie.serie = invented
				serie.line = self.line
				fictitious_series.append(serie)
			
			db.session.add_all(fictitious_series)
			
			try:
				db.session.commit()
			except:
				db.session.rollback()
				raise
	
	def delete(self, series):
		for serie in series:
			db.session.delete(serie)
		try:
			db.session.commit()
			for serie in series:
				self.series.remove(serie)
			self.layoutChanged.emit()
			self.series_at_expedition_level = self.get_series_at_expedition_level()
		except:
			db.session.rollback()
			raise
	
	def rowCount(self, index):
		return len(self.series)
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		if role == Qt.DisplayRole:
			return self.series[index.row()].serie
	
	def __contains__(self, serie):
		return serie in self.series_at_expedition_level
	
	def index_of(self, key):
		for index, serie in enumerate(self.series):
			if key == serie.serie:
				return index


class ExpeditionModel(BaseTable, QtCore.QAbstractTableModel):
	ID, WAREHOUSE, DATE, TOTAL, PROCESSED, LOGISTIC, CANCELLED, PARTNER, \
		AGENT, WARNING, FROM_PROFORMA, READY, PRESALE = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12
	
	def __init__(self, search_key=None, filters=None, last=10):
		super().__init__()
		self._headerData = ['Expedition ID', 'Warehouse', 'Date', 'Total', 'Processed',
		                    'Logistic', 'Cancelled', 'Partner', 'Agent', 'Warning', 'From Proforma', 'Ready To Go',
		                    'Presale']
		
		self.name = 'expeditions'
		
		query = db.session.query(db.Expedition). \
			select_from(
			db.SaleProforma,
			db.Partner, db.Agent,
			db.Warehouse). \
			where(
			db.Agent.id == db.SaleProforma.agent_id,
			db.Partner.id == db.SaleProforma.partner_id,
			db.Expedition.proforma_id == db.SaleProforma.id,
			db.Warehouse.id == db.SaleProforma.warehouse_id
		)
		
		query = query.where(db.SaleProforma.date > utils.get_last_date(last))
		
		if search_key:
			clause = or_(
				db.Agent.fiscal_name.contains(search_key), db.Partner.fiscal_name.contains(search_key), \
				db.Warehouse.description.contains(search_key))
			query = query.where(clause)
		
		if filters:
			self.expeditions = query.all()
			
			if filters['logistic']:
				if 'empty' in filters['logistic']:
					self.expeditions = filter(lambda e: e.proforma.empty, self.expeditions)
				if 'overflowed' in filters['logistic']:
					self.expeditions = filter(lambda e: e.proforma.overflowed, self.expeditions)
				if 'partially_processed' in filters['logistic']:
					self.expeditions = filter(lambda e: e.proforma.partially_processed, self.expeditions)
				if 'completed' in filters['logistic']:
					self.expeditions = filter(lambda e: e.proforma.completed, self.expeditions)
			if filters['cancelled']:
				if 'cancelled' in filters['cancelled']:
					self.expeditions = filter(lambda e: e.proforma.cancelled, self.expeditions)
				if 'notcancelled' in filters['cancelled']:
					self.expeditions = filter(lambda e: not e.proforma.cancelled, self.expeditions)
			
			if isinstance(self.expeditions, filter):
				self.expeditions = list(self.expeditions)
		else:
			self.expeditions = query.all()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		expedition = self.expeditions[index.row()]
		column = index.column()
		
		logisitic_status_string = expedition.logistic_status_string
		
		if role == Qt.DisplayRole:
			if column == self.ID:
				return str(expedition.id).zfill(6)
			elif column == self.WAREHOUSE:
				return expedition.proforma.warehouse.description
			elif column == self.TOTAL:
				return expedition.total_quantity
			elif column == self.PARTNER:
				return expedition.proforma.partner_name
			elif column == self.PROCESSED:
				return expedition.total_processed
			elif column == self.LOGISTIC:
				return logisitic_status_string
			
			elif column == self.PRESALE:
				return 'Yes' if expedition.proforma.advanced_lines else 'No'
			
			elif column == self.DATE:
				return expedition.proforma.date.strftime('%d/%m/%Y')
			
			elif column == self.CANCELLED:
				return 'Yes' if expedition.proforma.cancelled else 'No'
			elif column == self.AGENT:
				return expedition.proforma.agent.fiscal_name
			elif column == self.WARNING:
				return expedition.proforma.warning
			elif column == self.FROM_PROFORMA:
				return str(expedition.proforma.type) + '-' + str(expedition.proforma.number).zfill(6)
			elif column == self.READY:
				return 'Yes' if expedition.proforma.ready else 'No'
		
		elif role == Qt.DecorationRole:
			
			if column == self.AGENT:
				return QtGui.QIcon(':\\agents')
			
			elif column == self.DATE:
				return QtGui.QIcon(':\calendar')
			
			elif column == self.PARTNER:
				return QtGui.QIcon(':\partners')
			elif column == self.LOGISTIC:
				if logisitic_status_string == 'Empty':
					return QtGui.QColor(YELLOW)
				elif logisitic_status_string == 'Overflowed':
					return QtGui.QColor(RED)
				elif logisitic_status_string == 'Partially Processed':
					return QtGui.QColor(ORANGE)
				elif logisitic_status_string == 'Completed':
					return QtGui.QColor(GREEN)
			elif column == self.CANCELLED:
				return QtGui.QColor(RED if expedition.proforma.cancelled else GREEN)
			
			elif column == self.READY:
				return QtGui.QColor(GREEN if expedition.proforma.ready else RED)
	
	def sort(self, section, order):
		reverse = True if order == Qt.AscendingOrder else False
		if section == self.FROM_PROFORMA:
			self.layoutAboutToBeChanged.emit()
			self.expeditions = sorted(
				self.expeditions, key=lambda r: (r.proforma.type, r.proforma.number),
				reverse=reverse
			)
			self.layoutChanged.emit()
		else:
			attr = {
				self.ID: 'id',
				self.WAREHOUSE: 'proforma.warehouse_id',
				self.CANCELLED: 'proforma.cancelled',
				self.PARTNER: 'proforma.partner_name',
				self.AGENT: 'proforma.agent.fiscal_name',
				self.READY: 'proforma.ready',
				self.DATE: 'proforma.date'
			}.get(section)
			
			if attr:
				self.layoutAboutToBeChanged.emit()
				self.expeditions = sorted(self.expeditions, key=operator.attrgetter(attr), \
				                          reverse=True if order == Qt.DescendingOrder else False)
				self.layoutChanged.emit()


class ReceptionModel(BaseTable, QtCore.QAbstractTableModel):
	ID, WAREHOUSE, TOTAL, PROCESSED, LOGISTIC, CANCELLED, PARTNER, \
		AGENT, WARNING, FROM_PROFORMA = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9
	
	def __init__(self, search_key=None, filters=None, last=10):
		super().__init__()
		self._headerData = ['Reception ID', 'Warehouse', 'Total', 'Processed', 'Logistic', 'Cancelled',
		                    'Partner', 'Agent', 'Warning', 'From Proforma']
		self.name = 'receptions'
		
		query = db.session.query(db.Reception). \
			select_from(
			db.PurchaseProforma,
			db.Partner, db.Agent,
			db.Warehouse). \
			where(
			db.Agent.id == db.PurchaseProforma.agent_id,
			db.Partner.id == db.PurchaseProforma.partner_id,
			db.Reception.proforma_id == db.PurchaseProforma.id,
			db.Warehouse.id == db.PurchaseProforma.warehouse_id
		)
		
		query = query.where(db.PurchaseProforma.date > utils.get_last_date(last))
		
		if search_key:
			clause = or_(
				db.Agent.fiscal_name.contains(search_key), db.Partner.fiscal_name.contains(search_key), \
				db.Warehouse.description.contains(search_key))
			query = query.where(clause)
		
		if filters:
			self.receptions = query.all()
			if filters['logistic']:
				if 'empty' in filters['logistic']:
					self.receptions = filter(lambda r: r.empty, self.receptions)
				if 'overflowed' in filters['logistic']:
					self.receptions = filter(lambda r: r.overflowed, self.receptions)
				if 'partially_processed' in filters['logistic']:
					self.receptions = filter(lambda r: r.partially_processed, self.receptions)
				if "completed" in filters['logistic']:
					self.receptions = filter(lambda r: r.completed, self.receptions)
			
			if filters['cancelled']:
				if 'cancelled' in filters['cancelled']:
					self.receptions = filter(lambda r: r.proforma.cancelled, self.receptions)
				if 'notcancelled' in filters['cancelled']:
					self.receptions = filter(lambda r: not r.proforma.cancelled, self.receptions)
			
			if isinstance(self.receptions, filter):
				self.receptions = list(self.receptions)
		else:
			self.receptions = query.all()
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		reception = self.receptions[index.row()]
		column = index.column()
		logistic_status_string = reception.logistic_status_string
		if role == Qt.DisplayRole:
			if column == ReceptionModel.ID:
				return str(reception.id).zfill(6)
			elif column == ReceptionModel.WAREHOUSE:
				return reception.proforma.warehouse.description
			elif column == ReceptionModel.TOTAL:
				return str(reception.total_quantity)
			elif column == ReceptionModel.PARTNER:
				return reception.proforma.partner_name
			elif column == ReceptionModel.PROCESSED:
				return str(reception.total_processed)
			elif column == ReceptionModel.LOGISTIC:
				return logistic_status_string
			elif column == ReceptionModel.CANCELLED:
				return "Yes" if reception.proforma.cancelled else "No"
			elif column == ReceptionModel.AGENT:
				return reception.proforma.agent.fiscal_name
			elif column == ReceptionModel.WARNING:
				return reception.note
			elif column == ReceptionModel.FROM_PROFORMA:
				return str(reception.proforma.type) + '-' + str(reception.proforma.number).zfill(6)
		
		elif role == Qt.DecorationRole:
			if column == ReceptionModel.AGENT:
				return QtGui.QIcon(':\\agents')
			elif column == ReceptionModel.PARTNER:
				return QtGui.QIcon(':\partners')
			elif column == ReceptionModel.LOGISTIC:
				if logistic_status_string == 'Empty':
					return QtGui.QColor(YELLOW)
				elif logistic_status_string == 'Overflowed':
					return QtGui.QColor(RED)
				elif logistic_status_string == 'Partially Prepared':
					return QtGui.QColor(ORANGE)
				elif logistic_status_string == 'Completed':
					return QtGui.QColor(GREEN)
			elif column == ReceptionModel.CANCELLED:
				return QtGui.QColor(RED) if reception.proforma.cancelled \
					else QtGui.QColor(GREEN)
	
	def sort(self, section, order):
		reverse = True if order == Qt.AscendingOrder else False
		if section == ReceptionModel.FROM_PROFORMA:
			self.layoutAboutToBeChanged.emit()
			self.receptions = sorted(
				self.receptions, key=lambda r: (r.proforma.type, r.proforma.number),
				reverse=reverse
			)
			self.layoutChanged.emit()
		else:
			attr = {
				ReceptionModel.ID: 'id',
				ReceptionModel.WAREHOUSE: 'proforma.warehouse_id',
				ReceptionModel.CANCELLED: 'proforma.cancelled',
				ReceptionModel.PARTNER: 'proforma.partner_name',
				ReceptionModel.AGENT: 'proforma.agent.fiscal_name',
			}.get(section)
			
			if attr:
				self.layoutAboutToBeChanged.emit()
				self.receptions = sorted(self.receptions, key=operator.attrgetter(attr), \
				                         reverse=True if order == Qt.DescendingOrder else False)
				self.layoutChanged.emit()
	
	def generate_template(self, file_path, row):
		reception = self.receptions[row]
		try:
			from openpyxl import Workbook
			book = Workbook()
			sheet = book.active
			
			header = ['Imei/SN', 'Line Nº', 'Description', 'Condition', 'Spec']
			
			sheet.append([
				'Reception Id = ' + str(reception.id),
				'Partner = ' + reception.proforma.partner_name,
				'Agent = ' + reception.proforma.agent.fiscal_name,
				'Date = ' + str(reception.proforma.date)
			])
			
			sheet.append(header)
			
			sheet.append([])
			
			for row in self.generate_excel_rows(reception):
				sheet.append(row)
			
			book.save(file_path)
		
		except Exception as ex:
			raise ValueError(str(ex))
	
	def import_excel(self, file_path, row):
		reception = self.receptions[row]
		
		# First check that template was not altered
		from openpyxl import load_workbook
		book = load_workbook(file_path)
		sheet = book.active
		
		if sheet['A1'].value != 'Reception Id = ' + str(reception.id):
			raise ValueError("Reception doesn't match with Template")
		
		excel_rows = list(sheet.iter_rows(min_row=4, min_col=0, max_col=6, values_only=True))
		
		for excel_row, rec_row in zip(
				excel_rows,
				list(self.generate_excel_rows(reception))
		):
			if excel_row[1:] != rec_row[1:]:
				raise ValueError("Reception doesn't match with Template")
		
		from db import ReceptionSerie
		for row in excel_rows:
			# No me queda otra, pa no tocar todas las ocurrencias
			# de Reception serie. Debi haber dado flexibilidad a ese
			# Constructor antes. Ahora tenemos prisa jajaja
			line = self.get_line_from_line_id(reception, row[-1])
			rs = ReceptionSerie.from_excel(*row, line)
			db.session.add(rs)
		
		try:
			db.session.commit()
		except Exception as ex:
			db.session.rollback()
			# raise ValueError(str(ex))
			raise
	
	def export(self, file_path, row):
		reception = self.receptions[row]
		
		try:
			from openpyxl import Workbook
			book = Workbook()
			sheet = book.active
			
			for row in self.generate_export_rows(reception):
				sheet.append(row)
			
			book.save(file_path)
		except Exception as ex:
			raise ValueError(str(ex))
	
	def get_line_from_line_id(self, reception, line_id):
		for line in reception.lines:
			if line.id == line_id:
				return line
	
	def generate_export_rows(self, reception):
		for line in reception.lines:
			for serie in line.series:
				yield (serie.serie, serie.item.clean_repr)
	
	def generate_excel_rows(self, reception):
		for line_no, line in enumerate(reception.lines, start=1):
			if all((
					line.item_id,
					line.condition != 'Mix',
					line.spec != 'Mix',
					utils.has_serie(line)
			)):
				for i in range(line.quantity):
					yield (None, line_no, line.item.clean_repr, \
					       line.condition, line.spec, line.id)


import operator, functools


class StockEntry:
	
	def __init__(self, item_id, condition, spec, quantity):
		self._item_id = item_id
		self._spec = spec
		self._condition = condition
		self._quantity = int(quantity)
		self._request = 0
		# Make self.__eq__ with saleproformaline.__eq__ compatible
		self.description = None
	
	@property
	def excel_row(self):
		from utils import description_id_map
		return description_id_map.inverse.get(self.item_id), self.condition, self.spec, self.quantity
	
	@classmethod
	def fake(cls, item_id):
		return cls(item_id, '', '', 10)
	
	@property
	def item_id(self):
		return self._item_id
	
	@property
	def spec(self):
		return self._spec
	
	@property
	def condition(self):
		return self._condition
	
	@property
	def quantity(self):
		return self._quantity
	
	@quantity.setter
	def quantity(self, quantity):
		self._quantity = quantity
	
	@property
	def request(self):
		return self._request
	
	@request.setter
	def request(self, request):
		try:
			request = int(request)
		except ValueError:
			raise
		if self.quantity < request:
			raise ValueError('self.quantity < request')
		self._request = request
	
	def __iter__(self):
		return (i for i in (self.item_id, self.condition, \
		                    self.spec, self.quantity, self.request))
	
	def __repr__(self):
		class_name = self.__class__.__name__
		return '{}({!r}, {!r}, {!r}, {!r}, {!r})'.format(class_name, *self)
	
	def __eq__(self, other):
		if id(self) == id(other):
			return True
		return all((
			self.item_id == other.item_id,
			self.spec == other.spec,
			self.condition == other.condition
		))
	
	def __iadd__(self, other):
		# if self != other:
		# 	raise ValueError('Unsuported add for non-eqivalent vectors')
		self.quantity += other.quantity
		return self
	
	def __isub__(self, other):
		self.quantity -= other.quantity
		return self
	
	def __hash__(self):
		# None at the end of the list in able to make this hash
		# compatible with that of SaleProformaLine and enable
		# set operations. This is shit. But it works.
		hashes = (hash(x) for x in (self._item_id, self._spec, self._condition, self.description))
		return functools.reduce(operator.xor, hashes, 0)


class StockModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SPEC, QUANTITY, REQUEST = \
		0, 1, 2, 3, 4
	
	def __init__(self, warehouse_id, description, condition, spec, check=False):
		super().__init__()
		self._headerData = ['Description', 'Condition', 'Spec', 'Quantity avail. ', 'Requested quant.']
		
		if check:  # For checking only
			self._headerData.remove('Requested quant.')
		
		self.name = 'stocks'
		self.warehouse_id = warehouse_id
		
		self.stocks = self.computeStock(
			warehouse_id,
			description,
			condition,
			spec
		)
	
	@classmethod
	def stocks(cls, warehouse_id, description=None, condition=None, spec=None):
		
		return cls(warehouse_id, description=description, condition=condition, \
		           spec=spec).stocks
	
	def computeStock(self, warehouse_id, description, condition, spec, session=db.session):
		
		# session = session
		item_id = utils.description_id_map.get(description)
		
		query = session.query(
			db.Imei.item_id, db.Imei.condition,
			db.Imei.spec, func.count(db.Imei.imei).label('quantity')
		).where(
			db.Imei.warehouse_id == warehouse_id
		).group_by(
			db.Imei.item_id, db.Imei.condition, db.Imei.spec
		)
		
		if item_id:
			query = query.where(db.Imei.item_id == item_id)
		else:
			item_ids = utils.get_itemids_from_mixed_description(description)
			if item_ids:
				query = query.where(db.Imei.item_id.in_(item_ids))
		
		if condition:
			query = query.where(db.Imei.condition == condition)
		
		if spec:
			query = query.where(db.Imei.spec == spec)
		
		imeis = {StockEntry(r.item_id, r.condition, r.spec, r.quantity) for r in query}
		
		query = session.query(
			db.ImeiMask.item_id, db.ImeiMask.condition,
			db.ImeiMask.spec, func.count(db.ImeiMask.imei).label('quantity')
		).where(
			db.Warehouse.id == warehouse_id,
			db.Warehouse.id == db.ImeiMask.warehouse_id,
		).group_by(
			db.ImeiMask.item_id, db.ImeiMask.condition, db.ImeiMask.spec
		)
		
		if condition:
			query = query.where(db.ImeiMask.condition == condition)
		
		if spec:
			query = query.where(db.ImeiMask.spec == spec)
		
		imeis_mask = {StockEntry(r.item_id, r.condition, r.spec, r.quantity) for r in query}
		
		query = session.query(
			db.SaleProformaLine.item_id,
			db.SaleProformaLine.condition,
			db.SaleProformaLine.spec,
			func.sum(db.SaleProformaLine.quantity).label('quantity')
		).join(
			db.SaleProforma
		).group_by(
			db.SaleProformaLine.item_id,
			db.SaleProformaLine.condition,
			db.SaleProformaLine.spec
		).where(
			db.SaleProforma.warehouse_id == warehouse_id,
			db.SaleProforma.cancelled == False
		)
		
		sales = {StockEntry(r.item_id, r.condition, r.spec, r.quantity) for r in query}
		
		query = session.query(
			db.AdvancedLine.item_id,
			db.AdvancedLine.condition,
			db.AdvancedLine.spec,
			func.sum(db.AdvancedLine.quantity).label('quantity')
		).join(
			db.SaleProforma
		).group_by(
			db.AdvancedLine.item_id,
			db.AdvancedLine.condition,
			db.AdvancedLine.spec
		).where(
			db.SaleProforma.warehouse_id == warehouse_id,
			db.SaleProforma.cancelled == False,
			db.AdvancedLine.item_id != None
		)
		
		advanced_sales = {StockEntry(r.item_id, r.condition, r.spec, r.quantity) for r in query}
		
		query = session.query(
			db.ExpeditionLine.item_id,
			db.ExpeditionLine.condition,
			db.ExpeditionLine.spec,
			func.count(db.ExpeditionSerie.serie).label('quantity')
		).select_from(
			db.ExpeditionLine, db.ExpeditionSerie,
			db.SaleProforma, db.Warehouse, db.Expedition,
		).where(
			db.ExpeditionLine.id == db.ExpeditionSerie.line_id,
			db.ExpeditionLine.expedition_id == db.Expedition.id,
			db.SaleProforma.warehouse_id == db.Warehouse.id,
			db.Expedition.proforma_id == db.SaleProforma.id,
			db.Warehouse.id == warehouse_id,
		).group_by(
			db.ExpeditionLine.item_id,
			db.ExpeditionLine.condition,
			db.ExpeditionLine.spec
		)
		
		if item_id:
			query = query.where(db.ExpeditionLine.item_id == item_id)
		
		if condition:
			query = query.where(
				db.ExpeditionLine.condition == condition
			)
		if spec:
			query = query.where(
				db.ExpeditionLine.spec == spec
			)
		
		outputs = {StockEntry(r.item_id, r.condition, r.spec, r.quantity) for r in query}
		
		query = session.query(
			db.AdvancedLineDefinition.item_id,
			db.AdvancedLineDefinition.condition,
			db.AdvancedLineDefinition.spec,
			func.sum(db.AdvancedLineDefinition.quantity).label('quantity')
		).join(
			db.AdvancedLine
		).join(
			db.SaleProforma
		).where(
			db.SaleProforma.warehouse_id == warehouse_id,
			db.SaleProforma.cancelled == False
		).group_by(
			db.AdvancedLineDefinition.item_id,
			db.AdvancedLineDefinition.condition,
			db.AdvancedLineDefinition.spec,
		)
		
		defined_sales = {
			StockEntry(r.item_id, r.condition, r.spec, r.quantity) for r in query
		}
		
		# combine sales and definitions:
		r = sales.symmetric_difference(defined_sales)
		for sale in sales:
			for defined_sale in defined_sales:
				if sale == defined_sale:
					sale += defined_sale
					r.add(sale)
		sales = r
		
		# Combine sales and advanced sales:
		r = sales.symmetric_difference(advanced_sales)
		for sale in sales:
			for advanced_sale in advanced_sales:
				if sale == advanced_sale:
					sale += advanced_sale
					r.add(sale)
		sales = r
		
		stocks = self.resolve(imeis, imeis_mask, sales, outputs)
		
		return list(filter(lambda stock: stock.quantity != 0, stocks))
	
	def lines_against_stock(self, warehouse_id, lines):
		
		Session = db.sessionmaker(bind=db.engine)
		with Session.begin() as session:
			
			lines = set([line for line in lines if line.item_id is not None])
			stocks = set(self.computeStock(warehouse_id, None, None, None, session=session))
			
			if lines.difference(stocks):
				return True
			
			if any((
					line == stock and line.quantity > stock.quantity
					for line in lines
					for stock in stocks
			)):
				return True
			return False
	
	def resolve(self, imeis, imeis_mask, sales, outputs):
		
		for imei_mask in imeis_mask:
			for imei in imeis:
				if imei == imei_mask:
					imei -= imei_mask
		
		for output in outputs:
			for sale in sales:
				if sale == output:
					sale -= output
		
		for sale in sales:
			for imei in imeis:
				if sale == imei:
					imei -= sale
		
		return imeis
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		stock = self.stocks[row]
		
		if role == Qt.DisplayRole:
			if column == self.DESCRIPTION:
				return utils.description_id_map.inverse[stock.item_id]
			elif column == self.CONDITION:
				return stock.condition
			elif column == self.SPEC:
				return stock.spec
			elif column == self.QUANTITY:
				return str(stock.quantity) + ' pcs'
			elif column == self.REQUEST:
				return str(stock.request)
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid():
			return
		stock = self.stocks[index.row()]
		try:
			stock.request = value
		except ValueError as ex:
			return False
		
		self.dataChanged.emit(index, index)
		return True
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		if index.column() == 4:
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def reset(self):
		self.layoutAboutToBeChanged.emit()
		self.stocks = []
		self.layoutChanged.emit()
	
	def sort(self, section, order):
		attr = {
			self.DESCRIPTION: 'item_id',
			self.CONDITION: 'condition',
			self.SPEC: 'spec',
			self.QUANTITY: 'quantity'
		}.get(section)
		
		if attr:
			reverse = True if order == Qt.AscendingOrder else False
			self.layoutAboutToBeChanged.emit()
			self.stocks.sort(key=operator.attrgetter(attr), reverse=reverse)
			self.layoutChanged.emit()
	
	@property
	def requested_stocks(self):
		return list(filter(lambda s: s.request > 0, self.stocks))
	
	def excel_export(self, file_path):
		
		if not file_path:
			return
		from utils import warehouse_id_map
		from utils import description_id_map
		from openpyxl import Workbook
		
		wb = Workbook()
		ws = wb.active
		header = ['WH', 'Description', 'Condition', 'Spec', 'Quantity']
		ws.append(header)
		warehouse = warehouse_id_map.inverse.get(self.warehouse_id)
		for stock in self.stocks:
			ws.append((warehouse,) + stock.excel_row)
		
		wb.save(file_path)
	
	def whatsapp_export(self, file_path):
		print('whatsapp_export')


class IncomingVector:
	
	def __init__(self, line, session=db.session):
		
		self.origin_id = line.id
		self.origin = line
		self.type = line.proforma.type
		self.number = line.proforma.number
		self.eta = line.proforma.eta.strftime('%d/%m/%Y')
		self.item_id = line.item_id
		self.spec = line.spec
		self.condition = line.condition
		self.description = line.description
		
		quantity = line.quantity
		
		asked = sum(
			line.quantity for line in session.query(db.AdvancedLine.quantity).
			join(db.SaleProforma).where(db.SaleProforma.cancelled == False).
			where(db.AdvancedLine.origin_id == line.id)
		)
		
		processed = self.compute_processed(line)
		
		if quantity == processed:
			self.available = 0
		else:
			self.available = quantity - asked
	
	@property
	def excel_row(self):
		description = self.description or utils.description_id_map.inverse.get(self.item_id)
		return self.document, self.eta, description, self.condition, self.spec, self.available
	
	@property
	def document(self):
		return str(self.type) + '-' + str(self.number).zfill(6)
	
	@staticmethod
	def compute_processed(line):
		line_alias = line
		try:
			for line in line.proforma.reception.lines:
				if line_alias == line:
					return len(line.series)
			else:  # Reaches end, does not find it's brother in warehouse
				return 0
		
		except AttributeError as ex:
			return 0
	
	def __iter__(self):
		return iter(self.__dict__.values())
	
	def __repr__(self):
		clsname = self.__class__.__name__
		s = f"{clsname}(origin_id={self.origin_id}, description={self.description}, "
		s += f"item_id={self.item_id}, condition={self.condition}, spec={self.spec}, "
		s += f"available={self.available})"
		return s


class IncomingStockModel(BaseTable, QtCore.QAbstractTableModel):
	DOCUMENT, ETA, DESC, CONDITION, SPEC, AVAILABLE = 0, 1, 2, 3, 4, 5
	
	# check arg , make compatible with stock model init call
	def __init__(self, warehouse_id, *, description, condition, spec, type=None, number=None, check=False):
		super().__init__()
		self._headerData = ['Document', 'ETA', 'Description', 'Condition', 'Spec', 'Available']
		self.name = 'stocks'
		self.warehouse_id = warehouse_id
		self.stocks = self.computeIncomingStock(
			warehouse_id,
			description=description,
			condition=condition,
			spec=spec,
			type=type,
			number=number
		)
	
	def computeIncomingStock(self, warehouse_id, *, description, condition, spec,
	                         type=None, number=None, session=db.session):
		
		query = session.query(db.PurchaseProformaLine). \
			join(db.PurchaseProforma).join(db.Partner).outerjoin(db.Reception). \
			join(db.Warehouse).where(db.Warehouse.id == warehouse_id). \
			where(db.PurchaseProforma.cancelled == False)
		
		item_id = utils.description_id_map.get(description)
		
		if type:
			query = query.where(db.PurchaseProforma.type == type)
		
		if number:
			query = query.where(db.PurchaseProforma.number == number)
		
		if item_id:
			query = query.where(db.PurchaseProformaLine.item_id == item_id)
		else:
			item_ids = utils.get_itemids_from_mixed_description(description)
			if item_ids:
				predicates = (
					db.PurchaseProformaLine.item_id.in_(item_ids),
					db.PurchaseProformaLine.description == description  # ??
				)
				query = query.where(or_(*predicates))
		
		if condition:
			query = query.where(db.PurchaseProformaLine.condition == condition)
		
		if spec:
			query = query.where(db.PurchaseProformaLine.spec == spec)
		
		stocks = [
			IncomingVector(line, session=session) for line in query
			if self.line_contains_stock(line)
		]
		
		r = list(filter(lambda s: s.available != 0, stocks))
		
		return r
	
	@staticmethod
	def line_contains_stock(line):
		return line.description in utils.descriptions or line.item_id in utils.description_id_map.inverse
	
	def lines_against_stock(self, warehouse_id, lines):
		
		Session = db.sessionmaker(bind=db.engine)
		session = Session()
		
		for stock in self.computeIncomingStock(
				warehouse_id, description=None,
				condition=None, spec=None, session=session
		):
			for line in lines:
				if line == stock and line.quantity > stock.available:
					return True
		return False
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		
		row, col = index.row(), index.column()
		
		if role == Qt.DisplayRole:
			vector = self.stocks[row]
			if col == self.__class__.DOCUMENT:
				return vector.document
			elif col == self.__class__.ETA:
				return vector.eta
			elif col == self.__class__.DESC:
				return vector.description or \
					utils.description_id_map.inverse.get(vector.item_id)
				return vector.description
			elif col == self.__class__.CONDITION:
				return vector.condition
			elif col == self.__class__.SPEC:
				return vector.spec
			elif col == self.__class__.AVAILABLE:
				return str(vector.available)
	
	def reset(self):
		self.layoutAboutToBeChanged.emit()
		self.stocks = []
		self.layoutChanged.emit()
	
	def __getitem__(self, index):
		return self.stocks[index]
	
	def whatsapp_export(self):
		print('whatsapp export')
	
	def excel_export(self, file_path):
		
		header = ['WH', 'Document', 'ETA', 'Description', 'Condition', 'Spec', 'Available']
		from openpyxl import Workbook
		from utils import warehouse_id_map
		
		warehouse = warehouse_id_map.inverse.get(self.warehouse_id)
		
		wb = Workbook()
		ws = wb.active
		
		ws.append(header)
		
		for vector in self.stocks:
			ws.append((warehouse,) + vector.excel_row)
		
		wb.save(file_path)


class AdvancedLinesModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SHOWING_CONDITION, SPEC, IGNORING_SPEC, \
		QUANTITY, PRICE, SUBTOTAL, TAX, TOTAL = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9
	
	def __init__(self, proforma, form=None, show_free=True):
		super().__init__()
		self.form = form
		self._headerData = ['Description', 'Condition', 'Public Condt.(Editable)', 'Spec',
		                    'Ignoring Spec?(Editable)', 'Qty.', 'Price(Editable)', 'Subtotal',
		                    'Tax (Editable)', 'Total']
		self.name = 'lines'
		self.proforma = proforma 
		self._lines = proforma.advanced_lines
		if not show_free:
			self._lines = [line for line in self._lines if not line.free_description]
		
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			line = self._lines[row]
			if col == self.DESCRIPTION:
				return line.free_description or line.mixed_description \
					or utils.description_id_map.inverse.get(line.item_id)
			
			elif col == self.CONDITION:
				try:
					return line.condition
				except AttributeError:
					return ''
			elif col == self.SHOWING_CONDITION:
				return line.showing_condition or ''
			elif col == self.SPEC:
				try:
					return line.spec
				except AttributeError:
					return ''
			elif col == self.IGNORING_SPEC:
				return 'Yes' if line.ignore_spec else 'No'
			elif col == self.QUANTITY:
				return str(line.quantity)
			elif col == self.PRICE:
				return str(line.price)
			elif col == self.TAX:
				return str(line.tax)
			elif col == self.SUBTOTAL:
				return str(round(line.price * line.quantity, 2))
			elif col == self.TOTAL:
				total = round(line.quantity * line.price * (1 + line.tax / 100), 2)
				return str(total)
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return False
		row, column = index.row(), index.column()
	
		updated = False
		line = self.lines[row]

		if role==Qt.EditRole and self.is_free_line(index) and self.free_line_editable_column(column):
			if column == self.DESCRIPTION:
				line.free_description = value
				updated = True
			elif column == self.QUANTITY:
				try:
					v = int(value)
					line.quantity = v
					updated = True
				except ValueError:
					return False
			elif column == self.PRICE:
				try:
					v = float(value)
					line.price = v
					updated = True
				except ValueError: 
					return False 
			elif column == self.TAX:
				try:
					v = int(value) 
					if v not in (0, 4, 10, 21):
						return False 
					line.tax = v
					updated = True
				except ValueError:
					return False 
			else:
				return False 

		if role == Qt.EditRole:
			if column == self.SHOWING_CONDITION:
				line.showing_condition = value
				updated = True
			elif column == self.PRICE:
				try:
					v = float(value)
				except ValueError:
					pass
				else:
					line.price = v
					updated = True
			elif column == self.IGNORING_SPEC:
				value_lower = value.lower()
				if value_lower in ('yes', 'no'):
					line.ignore_spec = True if value_lower == 'yes' else False
					updated = True
			elif column == self.TAX:
				try:
					tax = int(value)
					if tax not in (0, 4, 10, 21):
						raise ValueError
				except ValueError:
					pass
				else:
					line.tax = tax
					updated = True

		if updated:
			db.session.flush() 
			if self.form is not None:  # In definition form is not necessary
				self.form.update_totals()
		
		return updated
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return Qt.ItemIsEnabled
		if self.editable_column(index.column()):
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return Qt.ItemFlags(~Qt.ItemIsEnabled)	
		
		if not self.is_free_line(index) and self.editable_column(index.column()):
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)

		if self.is_free_line(index) and self.free_line_editable_column(index.column()):
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)

		return Qt.ItemFlags(~Qt.ItemIsEditable)

	def is_free_line(self, index) -> bool:
		if not index.isValid():
			return False
		try:
			line = self.lines[index.row()]
		except IndexError:
			return False 
		
		return line.free_description is not None

	def is_free_line_by_row(self, row:int) -> bool:
		try:
			line = self._lines[row]
		except IndexError:
			return False 
		else:
			return line.free_description is not None

	def free_line_editable_column(self, column) -> bool:
		return column in (self.DESCRIPTION, self.QUANTITY, self.PRICE, self.TAX)

	def editable_column(self, column):
		return column in (self.PRICE, self.IGNORING_SPEC, self.SHOWING_CONDITION, self.PRICE,self.TAX)
	
	def add(self, quantity, price, ignore, tax, showing, vector):
		line = db.AdvancedLine()
		line.origin_id = vector.origin_id
		line.quantity = quantity
		line.price = price
		line.ignore_spec = ignore
		line.tax = tax
		line.showing_condition = showing
		line.condition = vector.condition
		line.spec = vector.spec
		line.item_id = vector.item_id
		line.mixed_description = vector.description
		
		if not self.update_if_preexists(vector, quantity):
			self._lines.append(line)
		
		db.session.flush()
		self.layoutChanged.emit()
	
	def delete(self, row):
		del self._lines[row]
		db.session.flush()
		self.layoutChanged.emit()
	
	def update_if_preexists(self, vector: IncomingVector, quantity):
		for line in self.lines:
			if all((
					vector.item_id == line.item_id,
					vector.description == line.mixed_description,
					vector.condition == line.condition,
					vector.spec == line.spec
			)):
				line.quantity += quantity
				return True
		return False
	
	def insert_free(self, description, quantity, price, tax):
		
		line = db.AdvancedLine()
		line.free_description = description
		line.quantity = quantity
		line.price = price
		line.tax = tax
		
		self._lines.append(line)

		db.session.flush()
		self.layoutChanged.emit()
	
	def reset(self):
		self.layoutAboutToBeChanged.emit()
		self._lines = []
		self.layoutChanged.emit()
	
	@property
	def quantity(self):
		return sum(line.quantity for line in self._lines)
	
	@property
	def lines(self):
		return self._lines
	
	@property
	def tax(self):
		return round(sum(
			line.quantity * line.price * line.tax / 100
			for line in self._lines
		), 2)
	
	@property
	def total(self):
		return round(self.subtotal + self.tax, 2)
	
	@property
	def subtotal(self):
		return round(sum(
			line.quantity * line.price for line in self._lines
		), 2)
	
	def update_count_relevant(self):
		for line in self._lines:
			for line_def in line.definitions:
				line_def.local_count_relevant = False
				line_def.global_count_relevant = True
	
	def __getitem__(self, index):
		return self._lines[index]
	
	def __bool__(self):
		return bool(self._lines)


from historical_inventory import HistoricalInventory

from utils import description_id_map
from utils import warehouse_id_map


class InventoryEntry:
	SERIE, ITEM_ID, CONDITION, SPEC, WAREHOUSE_ID, QUANTITY = 0, 1, 2, 3, 4, 5
	
	def __init__(self, **kwargs):
		
		kwargs['description'] = description_id_map.inverse[kwargs['item_id']]
		del kwargs['item_id']
		
		kwargs['warehouse'] = warehouse_id_map.inverse[kwargs['warehouse_id']]
		del kwargs['warehouse_id']
		
		self.__dict__.update(kwargs)
	
	@classmethod
	def from_tuple(cls, t):
		return cls(**{
			'serie': t[cls.SERIE],
			'item_id': t[cls.ITEM_ID],
			'condition': t[cls.CONDITION],
			'spec': t[cls.SPEC],
			'warehouse_id': t[cls.WAREHOUSE_ID],
			'quantity': 1
		})
	
	def __getitem__(self, item):
		if item == self.SERIE:
			return self.serie
		elif item == self.ITEM_ID:
			return self.description
		elif item == self.CONDITION:
			return self.condition
		elif item == self.SPEC:
			return self.spec
		elif item == self.WAREHOUSE_ID:
			return self.warehouse
		elif item == self.QUANTITY:
			return self.quantity
	
	@property
	def as_tuple(self):
		return tuple(self.__dict__.values())


def no_serie_grouper(registers):
	uuid_group = list(filter(lambda o: utils.valid_uuid(o.imei), registers))
	key = operator.attrgetter('item_id', 'condition', 'spec', 'warehouse_id')
	uuid_group = sorted(uuid_group, key=key)
	
	for key, group in groupby(uuid_group, key=key):
		item_id, condition, spec, warehouse_id = key
		yield InventoryEntry(
			serie="",
			item_id=item_id,
			condition=condition,
			spec=spec,
			warehouse_id=warehouse_id,
			quantity=len(list(group))
		)


class ActualInventory:
	
	def __init__(self, filters):
		q = db.session.query(
			db.Imei.imei, db.Imei.item_id, db.Imei.condition, db.Imei.spec, db.Imei.warehouse_id
		)
		
		if 'warehouse_id' in filters:
			q = q.where(db.Imei.warehouse_id == filters['warehouse_id'])
		
		if 'serie' in filters:
			q = q.where(db.Imei.imei.contains(filters['serie']))
		
		if 'item_ids' in filters:
			q = q.where(db.Imei.item_id.in_(filters['item_ids']))
		
		if 'condition' in filters:
			q = q.where(db.Imei.condition == filters['condition'])
		
		if 'spec' in filters:
			q = q.where(db.Imei.condition == filters['spec'])
		
		# Same Logic here. We want public access to self.registers
		# before any further processing is applied.
		self.registers = registers = [r for r in q]
		
		has_serie = lambda o: not utils.valid_uuid(o.imei)
		
		sn_group = list(filter(has_serie, registers))
		self.inventory = [InventoryEntry.from_tuple(e) for e in sn_group]
		
		# Check contains, because prepare_filters_dict cleans
		# removing entries with falsy value. Then
		# contains is equivalent to filter['include_no_serie'] being True.
		
		if 'include_no_serie' in filters:
			self.inventory.extend(list(no_serie_grouper(registers)))
	
	@classmethod
	def with_no_filters(cls):
		return cls({})
	
	@classmethod
	def from_warehouse(cls, warehouse_id):
		return cls({'warehouse_id': warehouse_id})
	
	def __getitem__(self, item):
		return self.inventory[item]
	
	@property
	def total(self):
		return sum(e.quantity for e in self.inventory)
	
	def __len__(self):
		return len(self.inventory)


from datetime import date


class InventoryModel(BaseTable, QtCore.QAbstractTableModel):
	
	def __init__(self, filters):
		super().__init__()
		self._headerData = ['Serie', 'Description', 'Condition', 'Spec', 'Warehouse', 'Quantity']
		self.name = 'inventory'
		
		if date.today() == filters['date']:
			self.inventory = ActualInventory(filters)
		else:
			self.inventory = HistoricalInventoryModel(filters)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		if role == Qt.DisplayRole:
			return self.inventory[index.row()][index.column()]
	
	@property
	def total(self):
		return self.inventory.total
	
	def __getitem__(self, item):
		return self.inventory[item]
	
	def export(self, file_path):
		wb = openpyxl.Workbook()
		sheet = wb.active
		for entry in self.inventory:
			sheet.append(entry.as_tuple)
		
		wb.save(file_path)


from historical_inventory import HistoricalInventory


class HistoricalInventoryModel:
	
	def __init__(self, filters):
		cutoff_date = filters['date']
		
		# Containing equivalent to filters['include_no_serie'] being true
		# because of the cleaning inventory_form.prepare_dict
		# performs.
		series_only = 'include_no_series' not in filters
		
		history = HistoricalInventory(
			cutoff_date,
			serie_only=series_only
		)
		
		# Make these registers public
		self.registers = registers = history.inventory
		
		# In case from_warehouse is used to instantiate
		# We don't want all filters to be applied
		# to self.registers, just warehouse.
		if 'warehouse_id' in filters:
			self.registers = registers = list(
				filter(lambda o: o.warehouse_id == filters['warehouse_id'], registers)
			)
		
		if 'item_ids' in filters:
			registers = list(
				filter(lambda o: o.item_id in filters['item_ids'], registers)
			)
		
		if 'condition' in filters:
			registers = list(
				filter(lambda o: o.condition == filters['condition'], registers)
			)
		
		if 'spec' in filters:
			registers = list(
				filter(lambda o: o.spec == filters['spec'], registers)
			)
		
		if 'serie' in filters:
			registers = list(
				filter(lambda o: filters['serie'] in o.imei, registers)
			)
		
		sn_group = list(filter(lambda o: not utils.valid_uuid(o.imei), registers))
		self.inventory = [InventoryEntry.from_tuple(e) for e in sn_group]
		
		if not series_only:
			self.inventory.extend(list(no_serie_grouper(registers)))
	
	@classmethod
	def with_no_filters(cls, cutoff_date):
		return cls({'date': cutoff_date, 'include_no_series': True})
	
	@classmethod
	def from_warehouse(cls, cutoff_date, warehouse_id):
		return cls({'warehouse_id': warehouse_id, 'date': cutoff_date, 'include_no_series': True})
	
	@property
	def total(self):
		return sum(e.quantity for e in self.inventory)
	
	def __getitem__(self, item):
		return self.inventory[item]
	
	def __len__(self):
		return len(self.inventory)


# Inventory Value Stuff
cr_named_field_names = [
	'description',
	'condition',
	'spec',
	'imei',
	'date',
	'partner',
	'doc',
	'currency',
	'cost',
	'quantity'
]


def create_valid_filename(name: str):
	# Windows file name restrictions
	invalid_chars = r'[\\/:*?"<>|]'
	return re.sub(invalid_chars, '', name)


class WarehouseValueModel(BaseTable, QtCore.QAbstractTableModel):
	def __init__(self, filters):
		super().__init__()
		self._headerData = [
			'Description',
			'Condition',
			'Spec',
			'Serie',
			'Date',
			'Partner',
			'Doc',
			'Currency',
			'Cost',
			'Quantity'
		]
		
		self.name = '_data'
		
		self._data = []
		
		_date = filters.get('date')
		all_path = filters.get('all')
		warehouse_id = filters.get('warehouse_id')
		book_value = filters.get('book_value')
		
		from datetime import date
		if _date == date.today():
			if warehouse_id:
				self.registers = ActualInventory.from_warehouse(warehouse_id).registers
			else:
				self.registers = ActualInventory.with_no_filters().registers
		
		else:
			if warehouse_id:
				self.registers = HistoricalInventoryModel.from_warehouse(
					cutoff_date=_date,
					warehouse_id=warehouse_id
				).registers
			
			else:
				self.registers = HistoricalInventoryModel.with_no_filters(cutoff_date=_date).registers
		
		promoted = [CostRegister(r, book_value=book_value) for r in self.registers]
		
		key = operator.attrgetter('description', 'condition', 'spec', 'purchase_date',
		                          'partner', 'doc_repr', 'currency', 'cost')
		
		if all_path:
			for name, warehouse_id in warehouse_id_map.items():
				filepath = os.path.join(all_path, create_valid_filename(name + '.xlsx'))
				wb = openpyxl.Workbook()
				sheet = wb.active
				sheet.append(self._headerData)
				warehouse_group = list(filter(lambda x: x.warehouse_id == warehouse_id, promoted))
				
				sn_group = filter(lambda o: not utils.valid_uuid(o.imei), warehouse_group)
				for elm in sn_group:
					sheet.append(elm.as_tuple)
				
				uuid_group = sorted(list(filter(lambda o: utils.valid_uuid(o.imei), warehouse_group)), key=key)
				
				for key, group in groupby(uuid_group, key=key):
					description, condition, spec, date, partner, doc_repr, currency, cost = key
					
					sheet.append((
						description, condition, spec, '', date, partner, doc_repr, currency,
						cost, len(list(group))
					))
				
				wb.save(filepath)
		
		else:
			self._data.extend(elm.as_tuple for elm in filter(lambda o: not utils.valid_uuid(o.imei), promoted))
			uuid_group = sorted(list(filter(lambda o: utils.valid_uuid(o.imei), promoted)), key=key)
			for key, group in groupby(uuid_group, key=key):
				description, condition, spec, date, partner, doc_repr, currency, cost = key
				self._data.append((
					description, condition, spec, '', date, partner, doc_repr, currency,
					cost, len(list(group))
				))
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		if role == QtCore.Qt.DisplayRole:
			row, col = index.row(), index.column()
			return self._data[row][col]
	
	def export(self, path):
		wb = openpyxl.Workbook()
		sheet = wb.active
		sheet.append(self._headerData)
		for t in self._data:
			sheet.append(t)
		wb.save(path)


class CostRegister:
	
	def __init__(self, inventory_entry, book_value=False):
		
		self.warehouse_id = inventory_entry.warehouse_id  # For later grouping
		
		self.description = description_id_map.inverse[inventory_entry.item_id]
		self.condition = inventory_entry.condition
		self.spec = inventory_entry.spec
		self.imei = inventory_entry.imei
		
		proforma = (
			db.session.query(db.PurchaseProforma).join(db.Reception).
			join(db.ReceptionLine).join(db.ReceptionSerie).
			where(db.ReceptionSerie.serie == inventory_entry.imei).
			order_by(db.ReceptionSerie.id.desc()).first()
		)
		
		reception_line = (
			db.session.query(db.ReceptionLine).join(db.ReceptionSerie).
			where(db.ReceptionSerie.serie == inventory_entry.imei).
			order_by(db.ReceptionSerie.id.desc()).first()
		)
		
		self.purchase_date = proforma.date.strftime('%d/%m/%Y')
		self.partner = proforma.partner_name
		
		try:
			self.doc_repr = proforma.invoice.doc_repr
		except AttributeError:
			self.doc_repr = proforma.doc_repr
		
		if proforma.eur_currency:
			self.currency = 'EUR'
			if book_value:
				self.cost = self.find_book_cost(proforma, reception_line)
			else:
				self.cost = do_cost_price(inventory_entry.imei).peuro
		else:
			if proforma.fully_paid:
				self.currency = 'EUR'
				if book_value:
					rate = get_avg_rate(proforma)
					self.cost = self.find_book_cost(proforma, reception_line) / rate
				else:
					self.cost = do_cost_price(inventory_entry.imei).peuro
			else:
				self.currency = 'USD'
				self.cost = self.find_book_cost(proforma, reception_line)
		
		self.qnt = 1
	
	@property
	def as_tuple(self):
		return (
			self.description, self.condition,
			self.spec, self.imei,
			self.purchase_date, self.partner,
			self.doc_repr, self.currency,
			self.cost, self.qnt
		)
	
	@staticmethod
	def find_book_cost(proforma, reception_line):
		for line in proforma.lines:
			if line == reception_line:
				return line.price
		raise ValueError('Could not match line - ReceptionLine ')
	
	def __repr__(self):
		return f'{self.__class__.__name__}{repr(self.__dict__.values())}'


class WarehouseListModel(QtCore.QAbstractListModel):
	
	def __init__(self):
		super().__init__()
		self.warehouses = [w for w in db.session.query(db.Warehouse)]
	
	def rowCount(self, index):
		return len(self.warehouses)
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		if role == Qt.DisplayRole:
			return self.warehouses[index.row()].description
	
	def delete(self, row):
		try:
			warehouse = self.warehouses[row]
		except IndexError:
			return
		else:
			db.session.delete(warehouse)
			try:
				db.session.commit()
				del self.warehouses[row]
				self.layoutChanged.emit()
			except:
				db.session.rollback()
				raise
	
	def add(self, warehouse_name):
		if warehouse_name in self.warehouses:
			raise ValueError
		
		warehouse = db.Warehouse(warehouse_name)
		db.session.add(warehouse)
		try:
			db.session.commit()
			self.warehouses.append(warehouse)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise


class ConditionListModel(QtCore.QAbstractListModel):
	
	def __init__(self):
		super().__init__()
		self.conditions = [c for c in db.session.query(db.Condition)]
	
	def rowCount(self, index=QModelIndex()):
		return len(self.conditions)
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		if role == Qt.DisplayRole:
			return self.conditions[index.row()].description
	
	def delete(self, row):
		try:
			condition = self.conditions[row]
		except:
			return
		else:
			if condition == 'Mix': return
			db.session.delete(condition)
			try:
				db.session.commit()
				del self.conditions[row]
				self.layoutChanged.emit()
			except:
				db.session.rollback()
				raise
	
	def add(self, condition_name):
		if condition_name in self.conditions:
			raise ValueError
		condition = db.Condition(condition_name)
		db.session.add(condition)
		try:
			db.session.commit()
			self.conditions.append(condition)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise


class CourierListModel(QtCore.QAbstractListModel):
	
	def __init__(self):
		super().__init__()
		self.couriers = db.session.query(db.Courier).all()
	
	def rowCount(self, index=QModelIndex()):
		return len(self.couriers)
	
	def data(self, index, role=QModelIndex()):
		if not index.isValid(): return
		if role == Qt.DisplayRole:
			return self.couriers[index.row()].description
	
	def delete(self, row):
		try:
			courier = self.couriers[row]
		except IndexError:
			return
		else:
			db.session.delete(courier)
			try:
				db.session.commit()
				del self.couriers[row]
				self.layoutChanged.emit()
			except:
				db.session.rollback()
				raise
	
	def add(self, name):
		
		if name in self:
			raise ValueError
		
		courier = db.Courier(name)
		db.session.add(courier)
		try:
			db.session.commit()
			self.couriers.append(courier)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise
	
	def __contains__(self, name):
		for e in self.couriers:
			if name == e.description:
				return True
		return False


class SpecListModel(QtCore.QAbstractListModel):
	
	def __init__(self):
		super().__init__()
		self.specs = [s for s in db.session.query(db.Spec)]
	
	def rowCount(self, index=QModelIndex()):
		return len(self.specs)
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid(): return
		if role == Qt.DisplayRole:
			return self.specs[index.row()].description
	
	def delete(self, row):
		try:
			spec = self.specs[row]
		except IndexError:
			return
		else:
			if spec == 'Mix': return
			db.session.delete(spec)
			try:
				db.session.commit()
				del self.specs[row]
				self.layoutChanged.emit()
			except:
				db.session.rollback()
				raise
	
	def add(self, spec_name):
		if spec_name in self.specs:
			raise ValueError
		
		spec = db.Spec(spec_name)
		db.session.add(spec)
		try:
			db.session.commit()
			self.specs.append(spec)
			self.layoutChanged.emit()
		except:
			db.session.rollback()
			raise


from sqlalchemy import text

# new query, limit the space to the current document.
sister_query = text(
	"""
		select exists(
		select 
			expedition_series.serie as serie_expedicion,
			reception_series.serie  as serie_reception
			,receptions.id 
		from expedition_series 
		inner join reception_series
		on reception_series.serie = expedition_series.serie and 
			reception_series.created_on = expedition_series.created_on
		inner JOIN reception_lines
			ON reception_lines.id = reception_series.line_id
		inner JOIN receptions
			ON reception_lines.reception_id = receptions.id
		where expedition_series.serie = :serie
			AND receptions.id = :current_document
		) as result
	"""
)

def has_sister_serie(serie, reception_id):
	return db.session.execute(sister_query, {'serie': serie, 'current_document':reception_id}).scalar()

class ReceptionSeriesModel:
	
	def __init__(self, reception):
		self.reception = reception
		self.reception_series = db.session.query(db.ReceptionSerie). \
			join(db.ReceptionLine).join(db.Reception). \
			where(db.Reception.id == reception.id).all()
	
	def add(self, line, serie, description, condition, spec):
		
		if has_sister_serie(serie, self.reception.id):
			raise ValueError('Serie already processed in expedition, please check there.')
		
		if serie.lower() in [o.serie.lower() for o in self.reception_series]:
			raise ValueError('Serie already processed in this reception order')
		
		if any((
				description not in utils.description_id_map.keys(),
				condition not in utils.conditions.difference({'Mix'}),
				spec not in utils.specs.difference({'Mix'})
		)): raise ValueError('Invalid description, condition or spec')
		
		reception_serie = db.ReceptionSerie(
			utils.description_id_map[description],
			line, serie, condition, spec
		)
		db.session.add(reception_serie)
		try:
			db.session.commit()
			self.reception_series.append(reception_serie)
		except:
			db.session.rollback()
			raise
	
	def add_invented(self, line, qnt, description, condition, spec):
		
		import uuid
		reception_series = []
		if qnt > 0:
			for i in range(qnt):
				reception_series.append(
					db.ReceptionSerie(
						utils.description_id_map[description],
						line, str(uuid.uuid4()), condition, spec
					)
				)
			db.session.add_all(reception_series)
			try:
				db.session.commit()
				self.reception_series.extend(reception_series)
			except:
				db.session.rollback()
				raise
		
		elif qnt < 0:
			# No podemos hacer LIMIT en sqlalchemy, damos este rodeo
			qnt = abs(qnt)
			counter, targets = 0, []
			for serie in self.reception_series:
				if all((
						serie.item_id == utils.description_id_map[description],
						serie.line_id == line.id,
						serie.condition == condition,
						serie.spec == spec
				)):
					counter += 1
					targets.append(serie)
					if counter == qnt:
						break
			
			for target in targets:
				db.session.delete(target)
			
			try:
				db.session.commit()
				for t in targets:
					self.reception_series.remove(t)
			except:
				db.session.rollback()
				raise
	
	def delete(self, series):
		delete_targets = [r for r in self.reception_series if r.serie in series]
		# Prevent deleting:
		if self.reception.auto and \
				any(has_sister_serie(serie.serie) for serie in delete_targets):
			raise ValueError('Serie already processed in expedition, please check there.')
		
		for t in delete_targets:
			db.session.delete(t)
		try:
			db.session.commit()
			for t in delete_targets:
				self.reception_series.remove(t)
		except:
			db.session.rollack()
			raise
	
	def processed_in(self, line):
		return sum(1 for o in self.reception_series if o.line_id == line.id)
	
	def __len__(self):
		return len(self.reception_series)


class DefinedSeriesModel(QtCore.QAbstractListModel):
	
	def __init__(self, rs_model, \
	             line, description, condition, spec):
		super().__init__()
		self.rs_model = rs_model
		
		self.series = [r.serie for r in rs_model.reception_series \
		               if all((
				r.line_id == line.id,
				r.item_id == utils.description_id_map[description],
				r.condition == condition,
				r.spec == spec
			))]
	
	def clear(self):
		self.layoutAboutToBeChanged.emit()
		self.series = []
		self.layoutChanged.emit()
	
	def rowCount(self, index):
		return len(self.series)
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		if role == Qt.DisplayRole:
			return self.series[index.row()]
	
	def index_of(self, key):
		for index, serie in enumerate(self.series):
			if key in serie:
				return index
	
	def __contains__(self, key):
		for serie in self.series:
			if key in serie:
				return True
		return False


Group = namedtuple('Group', 'description condition spec quantity')


class GroupModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SPEC, QUANTITY = 0, 1, 2, 3
	
	def __init__(self, rs_model, line):
		super().__init__()
		self._headerData = ['Description', 'Condition', 'Spec', 'Quantity']
		self.name = 'groups'
		key = attrgetter('item_id', 'condition', 'spec')
		
		series = rs_model.reception_series
		_filter = lambda o: o.line_id == line.id
		series = list(filter(_filter, series))
		series = sorted(series, key=key)
		
		self.groups = []
		for key, group in groupby(iterable=series, key=key):
			item_id, condition, spec = key
			self.groups.append(
				Group(utils.description_id_map.inverse[item_id],
				      condition, spec, len(list(group)))
			)
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid(): return
		row, col = index.row(), index.column()
		group = self.groups[row]
		if role == Qt.DisplayRole:
			return {
				self.__class__.DESCRIPTION: group.description,
				self.__class__.CONDITION: group.condition,
				self.__class__.SPEC: group.spec,
				self.__class__.QUANTITY: group.quantity
			}.get(col)
	
	def sort(self, section, order):
		attr = {
			self.DESCRIPTION: 'description',
			self.CONDITION: 'condition',
			self.SPEC: 'spec',
			self.QUANTITY: 'quantity'
		}.get(section)
		
		if attr:
			reverse = True if order == Qt.AscendingOrder else False
			self.layoutAboutToBeChanged.emit()
			self.groups.sort(key=operator.attrgetter(attr), reverse=reverse)
			self.layoutChanged.emit()
	
	def group_exists(self, description, condition, spec):
		pass
		for group in self.groups:
			if all((
					group.description == description,
					group.condition == condition,
					group.spec == spec
			)):
				return True
		return False
	
	def __len__(self):
		return len(self.groups)


class EditableGroupModel(GroupModel):
	
	def __init__(self, rs_model, line, form):
		self.line = line
		self.rs_model = rs_model
		self.form = form
		super().__init__(rs_model, line)
	
	def flags(self, index):
		if not index.isValid(): return Qt.ItemIsEnabled
		if index.column() == self.__class__.QUANTITY:
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid(): return False
		try:
			v = int(value)
			if v < 0:
				raise ValueError
		except ValueError:
			return False
		
		if role == Qt.EditRole:
			row, column = index.row(), index.column()
			if index.column() == self.__class__.QUANTITY:
				try:
					new_processed = int(value)
					if new_processed < 0:
						raise ValueError
				except ValueError:
					return False
				try:
					group = self.groups[row]
					old_processed = self.groups[row].quantity
					actual_item = group.description
					condition = group.condition
					spec = group.spec
					self.rs_model.add_invented(self.line,
					                           new_processed - old_processed,
					                           actual_item, condition, spec
					                           )
					self.form.populate_body()
					return True
				
				except:
					raise  # debug
					return False
		return True


class AdvancedStockModel(StockModel):
	DESCRIPTION, CONDITION, SPEC, QUANTITY, REQUEST = 0, 1, 2, 3, 4
	
	def __init__(self, warehouse_id=None, line=None):
		super(QtCore.QAbstractTableModel, self).__init__()
		
		self._headerData = ['Description', 'Condition', 'Spec',
		                    'Quantity avail. ', 'Requested quant.']
		self.name = 'stocks'
		
		if line is None or warehouse_id is None:
			self.stocks = []
		else:
			self.stocks = self.computeStock(warehouse_id, line)
	
	@property
	def requested_stocks(self):
		return list(filter(lambda s: s.request > 0, self.stocks))
	
	def computeStock(self, warehouse_id, line):
		
		query = db.session.query(
			db.ImeiMask.item_id, db.ImeiMask.condition,
			db.ImeiMask.spec, func.count(db.ImeiMask.imei).label('quantity')
		).where(
			db.ImeiMask.origin_id == line.origin_id,
			db.ImeiMask.warehouse_id == warehouse_id,
		).group_by(
			db.ImeiMask.item_id, db.ImeiMask.condition, db.ImeiMask.spec
		)
		
		if line.condition != 'Mix':
			query = query.where(db.ImeiMask.condition == line.condition)
		
		if line.spec != 'Mix':
			query = query.where(db.ImeiMask.spec == line.spec)
		
		if line.item_id:
			query = query.where(db.ImeiMask.item_id == line.item_id)
		
		elif line.mixed_description:
			item_ids = utils.get_itemids_from_mixed_description(line.mixed_description)
			query = query.where(db.ImeiMask.item_id.in_(item_ids))
		
		imeis = {
			StockEntry(r.item_id, r.condition, r.spec, r.quantity) for r in query
		}
		
		query = db.session.query(
			db.AdvancedLineDefinition.item_id,
			db.AdvancedLineDefinition.condition,
			db.AdvancedLineDefinition.spec,
			func.sum(db.AdvancedLineDefinition.quantity).label('quantity')
		).where(
			db.AdvancedLineDefinition.local_count_relevant == True
		).group_by(
			db.AdvancedLineDefinition.item_id,
			db.AdvancedLineDefinition.condition,
			db.AdvancedLineDefinition.spec
		)
		
		defined = {
			StockEntry(r.item_id, r.condition, r.spec, r.quantity)
			for r in query
		}
		
		for d in defined:
			for imei in imeis:
				if imei == d:
					imei -= d
		
		return list(filter(lambda s: s.quantity > 0, imeis))


class DefinedStockModel(BaseTable, QtCore.QAbstractTableModel):
	DESCRIPTION, CONDITION, SPEC, REQUESTED = 0, 1, 2, 3
	
	def __init__(self, line):
		super().__init__()
		self.line = line
		self._headerData = ['Description', 'Condition', 'Spec', 'Requested']
		self.name = 'definitions'
		self.definitions = line.definitions
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		
		if role == Qt.DisplayRole:
			definition = self.definitions[row]
			
			if col == self.DESCRIPTION:
				return definition.item.clean_repr
			elif col == self.CONDITION:
				return definition.condition
			elif col == self.SPEC:
				return definition.spec
			elif col == self.REQUESTED:
				return definition.quantity
	
	def add(self, *requested_stocks,
	        showing_condition=None):
		for stock in requested_stocks:
			self.line.definitions.append(
				db.AdvancedLineDefinition(
					item_id=stock.item_id,
					condition=stock.condition,
					spec=stock.spec,
					quantity=stock.request,
					showing_condition=showing_condition
				)
			)
		
		db.session.flush()
	
	def __iter__(self):
		return iter(self.definitions)
	
	def __getitem__(self, item):
		return self.definitions[item]
	
	def reset(self):
		self.layoutAboutToBeChanged.emit()
		self.line.definitions.clear()
		self.definitions = self.line.definitions
		db.session.flush()
		self.layoutChanged.emit()


class BankAccountModel(BaseTable, QtCore.QAbstractTableModel):
	NAME, IBAN, SWIFT, ADDRESS, POSTCODE, CITY, STATE, COUNTRY, ROUTING, CURRENCY = \
		0, 1, 2, 3, 4, 5, 6, 7, 8, 9
	
	def __init__(self, partner):
		super().__init__()
		self.accounts = partner.accounts
		self.name = 'accounts'
		self._headerData = ['Name', 'Iban', 'Swift', 'Address', \
		                    'Post code', 'City', 'State', 'Country', 'Routing', 'Currency']
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid(): return
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			account = self.accounts[row]
			return {
				self.__class__.NAME: account.bank_name,
				self.__class__.IBAN: account.iban,
				self.__class__.SWIFT: account.swift,
				self.__class__.ADDRESS: account.bank_address,
				self.__class__.POSTCODE: account.bank_postcode,
				self.__class__.CITY: account.bank_city,
				self.__class__.STATE: account.bank_state,
				self.__class__.COUNTRY: account.bank_country,
				self.__class__.ROUTING: account.bank_routing,
				self.__class__.CURRENCY: account.currency
			}.get(col)
	
	def add(self, name, iban, swift, address, postcode, city, state, country, \
	        routing, currency):
		self.layoutAboutToBeChanged.emit()
		account = db.PartnerAccount(name, iban, swift, address, postcode, \
		                            city, state, country, routing, currency)
		self.accounts.append(account)
		self.layoutChanged.emit()
	
	def delete(self, row):
		
		self.layoutAboutToBeChanged.emit()
		
		del self.accounts[row]
		
		self.layoutChanged.emit()
	
	def setData(self, index, value, role=Qt.EditRole):
		if not index.isValid(): return False
		row, col = index.row(), index.column()
		account = self.accounts[row]
		if role == Qt.EditRole:
			if col == self.__class__.NAME:
				account.bank_name = value
			elif col == self.__class__.IBAN:
				account.iban = value
			elif col == self.__class__.ADDRESS:
				account.bank_address = value
			elif col == self.__class__.SWIFT:
				account.swift = value
			elif col == self.__class__.POSTCODE:
				account.bank_postcode = value
			elif col == self.__class__.CITY:
				account.bank_city = value
			elif col == self.__class__.STATE:
				account.bank_state = value
			elif col == self.__class__.COUNTRY:
				account.bank_country = value
			elif col == self.__class__.ROUTING:
				account.bank_routing = value
			elif col == self.__class__.CURRENCY:
				account.currency = value
			return True
		return False
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)


from db import Spec, Condition, Warehouse


class ChangeModel(BaseTable, QtCore.QAbstractTableModel):
	SN, DESCRIPTION, HINT = 0, 1, 2
	
	def __init__(self, hint=None):
		
		from importlib import reload
		global utils
		utils = reload(utils)
		
		super().__init__()
		self.hint = hint
		self._headerData = ['IMEI/SN', 'Description', hint.capitalize()]
		
		self.sns = []
		self.name = 'sns'
		self.hint = hint
		
		self.set_things()
	
	def set_things(self):
		if self.hint == 'warehouse':
			self.attrname = 'warehouse.description'
			self.orm_class = Warehouse
			self.orm_change_class = db.WarehouseChange
		elif self.hint == 'spec':
			self.attrname = 'spec'
			self.orm_class = Spec
			self.orm_change_class = db.SpecChange
		elif self.hint == 'condition':
			self.attrname = 'condition'
			self.orm_class = Condition
			self.orm_change_class = db.ConditionChange
	
	def data(self, index, role=Qt.DisplayRole):
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		if role == Qt.DisplayRole:
			imei_object = self.sns[row]
			if column == self.SN:
				return imei_object.imei
			elif column == self.DESCRIPTION:
				return imei_object.item.clean_repr
			elif column == self.HINT:
				if self.hint == 'warehouse':
					return imei_object.warehouse.description
				return getattr(imei_object, self.attrname)
	
	def search(self, sn):
		if sn in self:
			return
		
		imeis = db.session.query(db.Imei).where(db.Imei.imei == sn).all()
		mask = db.session.query(db.ImeiMask).all()
		
		aux = []
		for imei in imeis:
			for imei_mask in mask:
				if imei.imei == imei_mask.imei:
					break
			else:
				aux.append(imei)
		
		self.sns.extend(aux)
		self.layoutChanged.emit()
	
	def apply(self, name, comment):
		
		if self.hint == 'warehouse':
			
			for imei_object in self.sns:
				before = imei_object.warehouse.description
				after = name
				
				if after != before:
					db.session.add(db.WarehouseChange(imei_object.imei, before, after, comment))
					imei_object.warehouse_id = utils.warehouse_id_map.get(name)
		else:
			
			for imei_object in self.sns:
				before = getattr(imei_object, self.attrname)
				after = name
				if after != before:
					db.session.add(self.orm_change_class(imei_object.imei, before, after, comment))
					setattr(imei_object, self.attrname, name)
		
		try:
			db.session.commit()
			self.layoutChanged.emit()
		except:
			raise
	
	def __len__(self):
		return len(self.sns)
	
	def __contains__(self, sn):
		sn = sn.lower()
		for e in self.sns:
			if e.imei.lower() == sn:
				return True
		else:
			return False


def export_invoices_sales_excel(invoice, file_path):
	book = Workbook()
	sheet = book.active
	header = ['Imei/SN', 'Description', 'Condition', 'Spec']
	sheet.append([
		'Document Date = ' + str(invoice.date),
		'Document Number = ' + invoice.doc_repr,
		'Customer = ' + invoice.partner_name,
		'Supplier = ' + 'Euromedia Investment Group S.L. ',
		'Agent = ' + invoice.agent
	])
	
	sheet.append([])
	sheet.append(header)
	for proforma in invoice.proformas:
		for row in generate_excel_rows(proforma):
			sheet.append(list(row))
	
	book.save(file_path)


def export_proformas_sales_excel(proforma, file_path):
	book = Workbook()
	sheet = book.active
	
	header = ['Imei/SN', 'Description', 'Condition', 'Spec']
	
	# Document Date
	# Customer
	# Supplier
	# Agente primera parte
	
	sheet.append([
		'Document Date = ' + str(proforma.date),
		'Document Number = ' + proforma.doc_repr,
		'Customer = ' + proforma.partner_name,
		'Supplier = ' + 'Euromedia Investment Group S.L. ',
		'Agent = ' + proforma.agent.fiscal_name.split()[0]
	])
	
	sheet.append([])
	
	sheet.append(header)
	
	for row in generate_excel_rows(proforma):
		sheet.append(list(row))
	
	book.save(file_path)


def generate_excel_rows(proforma):
	# Handle credit notes whose main property is that
	# warehouse_id is NULL on SQL side.
	if proforma.is_credit_note:
		for line in proforma.credit_note_lines:
			yield line.sn, line.item.clean_repr, line.public_condition or line.condition, line.spec
	
	else:
		for eline in proforma.expedition.lines:
			condition = get_condition(eline)
			spec = get_spec(eline)
			description = eline.item.clean_repr
			
			for serie in eline.series:
				yield serie.serie, description, condition, spec


def get_condition(eline: db.ExpeditionLine):
	lines_iter = eline.expedition.proforma.lines or eline.expedition.proforma.advanced_lines
	for pline in lines_iter:
		if pline == eline:
			return pline.showing_condition or pline.condition


def get_spec(eline: db.ExpeditionLine):
	for pline in eline.expedition.proforma.lines:
		if all((
				eline.item_id == pline.item_id,
				eline.condition == eline.condition,
				eline.spec == eline.spec
		)):
			if pline.ignore_spec:
				return ''
			else:
				return pline.spec


class WhRmaIncomingModel(BaseTable, QtCore.QAbstractTableModel):
	ID, PARTNER, DATE, QUANTITY, STATUS, INVOICE, DUMPED = 0, 1, 2, 3, 4, 5, 6
	
	def __init__(self, search_key=None, filters=None, last=10):
		super().__init__()
		self._headerData = ['ID', 'Partner', 'Date', 'Quantity', 'Status', 'Invoice', 'Exported']
		self.name = 'orders'
		
		query = db.session.query(db.WhIncomingRma).join(db.IncomingRma).join(
			db.IncomingRmaLine).join(db.WhIncomingRmaLine)
		
		query = query.where(db.IncomingRma.date > utils.get_last_date(last))
		
		if search_key:
			query = query.where(
				or_(
					db.WhIncomingRmaLine.sn.contains(search_key),
					db.IncomingRmaLine.cust.contains(search_key)
				)
			)
		
		self.orders = query.all()
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		order = self.orders[row]
		
		if role == Qt.DisplayRole:
			if column == self.ID:
				return str(order.incoming_rma.id).zfill(6)
			elif column == self.PARTNER:
				return order.incoming_rma.lines[0].cust
			elif column == self.DATE:
				return order.incoming_rma.date.strftime('%d/%m/%Y')
			elif column == self.STATUS:
				s = {line.accepted for line in order.lines}
				if s == {'y'}:
					return 'Accepted'
				elif s == {'n'}:
					return 'Rejected'
				elif 'p' in s:
					return 'Pending'
				else:
					return 'Partially Accepted'
			
			elif column == self.INVOICE:
				invoices = order.invoice
				if invoices:
					return invoices
				else:
					return 'Not yet created'
			
			elif column == self.QUANTITY:
				return str(len(order.lines)) + ' pcs '
			
			elif column == self.DUMPED:
				return 'Yes' if order.dumped else 'No'
		
		elif role == Qt.DecorationRole:
			
			if column == self.STATUS:
				
				s = {line.accepted for line in order.lines}
				if s == {'y'}:
					color = GREEN
				elif s == {'n'}:
					color = RED
				elif 'p' in s:
					color = ORANGE
				else:
					color = YELLOW
				
				return QtGui.QColor(color)
			elif column == self.PARTNER:
				return QtGui.QIcon(':\partners')
			elif column == self.DATE:
				return QtGui.QIcon(':\calendar')


from utils import warehouse_id_map


def exists_credit_line(imei):
	return db.session.query(exists().where(db.CreditNoteLine.sn == imei)).scalar()


def rma_credit_difference(imei):
	credit_count = db.session.query(func.count(db.CreditNoteLine.id)). \
		where(db.CreditNoteLine.sn == imei).scalar()
	
	rma_count = db.session.query(func.count(db.WhIncomingRmaLine.id)).where(
		db.WhIncomingRmaLine.accepted == 'y',
		db.WhIncomingRmaLine.sn == imei
	).scalar()
	
	return credit_count < rma_count


class CreditNoteLineModel(BaseTable, QtCore.QAbstractTableModel):
	ITEM, CONDITION, SPEC, QUANTITY, PRICE, TAX, IMEI = 0, 1, 2, 3, 4, 5, 6
	
	def __init__(self, lines):
		super().__init__()
		self._headerData = ['Description', 'Condition', 'Spec', 'Quantity', 'Price', 'Tax', 'IMEI']
		self.name = 'lines'
		self.lines = lines
	
	@property
	def subtotal(self):
		return round(sum(line.price * line.quantity for line in self.lines), 2)
	
	@property
	def tax(self):
		return round(sum(line.quantity * line.price * line.tax / 100 for line in self.lines), 2)
	
	@property
	def total(self):
		return round(self.subtotal + self.tax, 2)
	
	@property
	def quantity(self):
		return sum(line.quantity for line in self.lines)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		if role == Qt.DisplayRole:
			line = self.lines[row]
			if column == self.ITEM:
				return description_id_map.inverse[line.item_id]
			elif column == self.IMEI:
				return line.sn
			elif column == self.CONDITION:
				return line.condition
			elif column == self.SPEC:
				return line.spec
			elif column == self.QUANTITY:
				return str(line.quantity)
			elif column == self.PRICE:
				return str(line.price)
			elif column == self.TAX:
				return str(line.tax)


class WhRmaIncomingLineModel(BaseTable, QtCore.QAbstractTableModel):
	SN, DESCRIPTION, CONDITION, SPEC, PROBLEM, ACCEPTED, WHY, WAREHOUSE, TARGET_CONDITION = 0, 1, 2, 3, 4, 5, 6, 7, 8
	
	def __init__(self, lines, block_target=False):
		super().__init__()
		self._headerData = ['SN/IMEI', 'Description', 'Condt.', 'Spec',
		                    'Problem', 'Accepted(y/n/p)', 'Why', 'Target WH', 'Target Condt.']
		self.name = 'lines'
		self.lines = lines
		self.block_target = block_target
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		if role == Qt.DisplayRole:
			line = self.lines[row]
			if column == self.SN:
				return line.sn
			elif column == self.PROBLEM:
				return line.problem
			elif column == self.WHY:
				return line.why
			elif column == self.ACCEPTED:
				return line.accepted
			elif column == self.WAREHOUSE:
				return warehouse_id_map.inverse[line.warehouse_id]
			elif column == self.DESCRIPTION:
				return line.item.clean_repr
			elif column == self.CONDITION:
				return line.condition
			elif column == self.SPEC:
				return line.spec
			elif column == self.TARGET_CONDITION:
				return line.target_condition
		
		
		elif role == Qt.BackgroundRole:
			if column == self.TARGET_CONDITION:
				''' Return yellow color '''
				return QtGui.QColor(255, 255, 0)
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		editables = [self.WHY, self.ACCEPTED, self.WAREHOUSE]
		
		if not self.block_target:
			editables.append(self.TARGET_CONDITION)
		
		if index.column() in editables:
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	''' Simplify the method above '''
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return False
		
		row, column = index.row(), index.column()
		if role == Qt.EditRole:
			line = self.lines[row]
			if column == self.WHY:
				line.why = value
				return True
			elif column == self.ACCEPTED:
				value = value.lower()
				if value in ('y', 'n', 'p'):
					line.accepted = value
					return True
			elif column == self.WAREHOUSE:
				try:
					line.warehouse_id = warehouse_id_map[value]
					return True
				except KeyError:
					return False
			
			elif column == self.TARGET_CONDITION:
				if value not in utils.conditions:
					return False
				line.target_condition = value
				return True
			return False
		return False


class RmaIncomingModel(BaseTable, QtCore.QAbstractTableModel):
	ID, PARTNER, DATE, QNT, ACCEPTED, INWH = 0, 1, 2, 3, 4, 5
	
	def __init__(self, search_key=None, filters=None, last=10):
		
		super().__init__()
		
		self.name = 'orders'
		
		self._headerData = ['ID', 'Partner', 'Date', 'Total ', 'Accepted', 'In WH']
		
		query = db.session.query(db.IncomingRma).join(db.IncomingRmaLine). \
			where(db.IncomingRma.date > utils.get_last_date(last))
		
		if search_key:
			query = query.filter(
				or_(
					db.IncomingRmaLine.sn.contains(search_key),
					db.IncomingRmaLine.cust.contains(search_key)
				)
			)
		
		self.orders = query.all()
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		
		if not index.isValid():
			return
		
		row, column = index.row(), index.column()
		rma_order = self.orders[row]
		if role == Qt.DisplayRole:
			if column == self.ID:
				return str(rma_order.id).zfill(6)
			elif column == self.DATE:
				return rma_order.date.strftime('%d/%m/%Y')
			elif column == self.PARTNER:
				try:
					return rma_order.lines[0].cust
				except IndexError:
					return 'Unknown'
			
			elif column == self.ACCEPTED:
				return str(sum(line.accepted for line in rma_order.lines)) + ' pcs'
			
			elif column == self.INWH:
				return 'Yes' if rma_order.wh_incoming_rma is not None else 'No'
			elif column == self.QNT:
				return str(len(rma_order.lines)) + ' pcs'
		
		elif role == Qt.DecorationRole:
			
			if column == self.DATE:
				
				return QtGui.QIcon(':\calendar')
			
			elif column == self.ACCEPTED:
				if all((line.accepted for line in rma_order.lines)):
					return QtGui.QColor(GREEN)
				
				elif all((not line.accepted for line in rma_order.lines)):
					return QtGui.QColor(RED)
				
				elif len({line.accepted for line in rma_order.lines}) == 2:  # We found both
					return QtGui.QColor(ORANGE)
			
			elif column == self.PARTNER:
				return QtGui.QIcon(':\partners')
	
	def sort(self, column: int, order: Qt.SortOrder = ...) -> None:
		reverse = True if order == Qt.AscendingOrder else False
		attr = {
			self.ID: 'id',
			self.DATE: 'date'
		}.get(column)
		if attr:
			self.layoutAboutToBeChanged.emit()
			self.orders.sort(key=operator.attrgetter(attr), reverse=reverse)
			self.layoutChanged.emit()
	
	def to_warehouse_old(self, row):
		rma_order: db.IncomingRma = self.orders[row]
		wh_rma_order = db.WhIncomingRma(rma_order)
		
		if all(not line.accepted for line in rma_order.lines):
			raise ValueError('Rma not accepted. I will not send to WH')
		
		# Raises value error if invoice type not found
		lines = []
		for line in rma_order.lines:
			if line.accepted:
				wh_rma_order.lines.append(db.WhIncomingRmaLine(line))
		
		for line in rma_order.lines:
			if line.accpeted:
				lines.append(db.WhIncomingRmaLine(line))
		
		try:
			db.session.commit()
		except IntegrityError as ex:
			db.session.rollback()
			raise ValueError('Wh order already exists')
	
	def to_warehouse(self, row):
		rma_order = self.orders[row]
		
		if all(not line.accepted for line in rma_order.lines):
			raise ValueError('Rma not accepted. I will not send to WH')
		
		lines = []
		for line in rma_order.lines:
			if line.accepted:
				lines.append(db.WhIncomingRmaLine(line))
		
		wh_order = db.WhIncomingRma(rma_order)
		wh_order.lines.extend(lines)
		db.session.commit()  # No posible error raising
	
	def __getitem__(self, item):
		return self.orders[item]


class RmaIncomingLineModel(BaseTable, QtCore.QAbstractTableModel):
	IMEI, SUPP, RCPT, WTYENDSUPP, PURCHAS, DEFAS, SOLD_AS, PUBLIC, CUST, SALE_DATE, \
		EXPED, WTYENDC, PROBLEM, ACCEPTED, WHY = 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14
	
	def __init__(self, lines):
		super().__init__()
		self._headerData = [
			'Imei/SN', 'Supp.', 'Rcpt.', 'Wty. End(S)', 'Pur. as', 'Def. as', 'Sold as', 'Public condt.', 'Cust.',
			'Sale Date', 'Exped.', 'Wty End(C)', 'Problem', 'Accepted (y/n)', 'Why'
		]
		self.name = 'lines'
		self.lines = lines
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		line: db.IncomingRmaLine = self.lines[row]
		if role == Qt.DisplayRole:
			if column == self.IMEI:
				return line.sn
			elif column == self.SUPP:
				return line.supp
			elif column == self.RCPT:
				return line.recpt.strftime('%d-%m-%Y')
			elif column == self.WTYENDSUPP:
				return str(line.wtyendsupp)
			elif column == self.PURCHAS:
				return line.purchas
			elif column == self.DEFAS:
				return line.defas
			elif column == self.SOLD_AS:
				return line.soldas
			elif column == self.PUBLIC:
				return line.public
			elif column == self.CUST:
				return line.cust
			elif column == self.SALE_DATE:
				return str(line.saledate)
			elif column == self.EXPED:
				return line.exped.strftime('%d-%m-%Y')
			elif column == self.WTYENDC:
				return str(line.wtyendcust)
			elif column == self.PROBLEM:
				return line.problem
			elif column == self.ACCEPTED:
				return 'y' if line.accepted else 'n'
			elif column == self.WHY:
				return line.why
	
	def flags(self, index):
		if not index.isValid():
			return Qt.ItemIsEnabled
		if index.column() in (self.PROBLEM, self.ACCEPTED, self.WHY):
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		if role == Qt.EditRole:
			line = self.lines[row]
			if col == self.WHY:
				line.why = value
			elif col == self.PROBLEM:
				line.problem = value
			elif col == self.ACCEPTED:
				value = value.lower()
				if value in ('y', 'n', 'p'):
					line.accepted = value == 'y'
			
			return True
		return False
	
	def serie_processed(self, sn):
		for line in self.lines:
			if line.sn == sn:
				return True
		else:
			return False
	
	def add(self, sn):
		if self.serie_processed(sn):
			raise ValueError('Imei/Sn already processed')
		
		if db.session.query(exists().where(db.Imei.imei == sn)).scalar():
			raise ValueError(f"Imei : {sn} in inventory")
		
		try:
			self.layoutAboutToBeChanged.emit()
			self.lines.append(db.IncomingRmaLine.from_sn(sn))
			self.layoutChanged.emit()
			db.session.flush()
		except ValueError:
			raise
	
	def update_warehouse(self):
		# Not necessary to check index error
		wh_order = self.lines[0].incoming_rma.wh_incoming_rma
		
		if wh_order is not None:
			
			wh_lines = wh_order.lines
			rma_lines = self.lines
			
			for line in self.rma_warehouse_difference(rma_lines, wh_lines):
				if line.accepted:
					wh_order.lines.append(db.WhIncomingRmaLine(line))
			
			db.session.commit()
			
			for line in self.warehouse_rma_difference(wh_lines, rma_lines):
				db.session.delete(line)
			
			db.session.commit()
	
	@staticmethod
	def warehouse_rma_difference(wh_lines, rma_lines):
		aux = []
		for wh_line in wh_lines:
			for rma_line in rma_lines:
				if wh_line.sn == rma_line.sn:
					break
			else:
				aux.append(wh_line)
		return aux
	
	@staticmethod
	def rma_warehouse_difference(rma_lines, wh_lines):
		aux = []
		for rma_line in rma_lines:
			for wh_line in wh_lines:
				if rma_line.sn == wh_line.sn:
					break
			else:
				aux.append(rma_line)
		return aux
	
	def delete(self, row):
		self.layoutAboutToBeChanged.emit()
		self.lines.pop(row)
		db.session.flush()
		self.layoutChanged.emit()
	
	def __iter__(self):
		return iter(self.lines)
	
	def __bool__(self):
		return bool(self.lines)


class ChangeModelTrace(BaseTable, QtCore.QAbstractTableModel):
	FROM, TO, WHEN, COMMENT = 0, 1, 2, 3
	
	def __init__(self, Sqlalchemy_cls, imei):
		super().__init__()
		self._headerData = ['FROM', 'TO', 'WHEN', 'COMMENT']
		self.registers = db.session.query(Sqlalchemy_cls). \
			where(Sqlalchemy_cls.sn == imei).all()
		self.name = 'registers'
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		r = self.registers[row]
		
		if role == Qt.DisplayRole:
			if col == self.FROM:
				return r.before
			elif col == self.TO:
				return r.after
			elif col == self.WHEN:
				return str(r.created_on)
			elif col == self.COMMENT:
				return str(r.comment)


class TraceEntry:
	
	def __str__(self):
		return str(self.__dict__)


class OperationModel(BaseTable, QtCore.QAbstractTableModel):
	OPERATION, DOC, DATE, PARTNER, PICKING = 0, 1, 2, 3, 4
	
	def __init__(self, imei):
		super().__init__()
		self._headerData = ['Operation', 'Doc', 'Date', 'Partner', 'Picking']
		self.name = 'entries'
		self.entries = []
		
		query = db.session.query(db.ReceptionSerie). \
			join(db.ReceptionLine).join(db.Reception). \
			join(db.PurchaseProforma). \
			where(db.ReceptionSerie.serie == imei)
		
		for r in query:
			te = TraceEntry()
			te.operation = 'Purchase'
			# Get doc
			try:
				te.doc = 'FR ' + r.line.reception.proforma.invoice.doc_repr
			except AttributeError:
				te.doc = 'PR ' + r.line.reception.proforma.doc_repr
			
			# Get date:
			try:
				te.date = r.line.reception.proforma.invoice.date
			except AttributeError:
				te.date = r.line.reception.proforma.date
			
			te.partner = r.line.reception.proforma.partner_name
			te.picking = r.created_on
			
			self.entries.append(te)
		
		query = db.session.query(db.ExpeditionSerie).join(db.ExpeditionLine). \
			join(db.Expedition).join(db.SaleProforma).where(db.ExpeditionSerie.serie == imei)
		
		for r in query:
			te = TraceEntry()
			te.operation = 'Sale'
			try:
				te.doc = 'FR ' + r.line.expedition.proforma.invoice.doc_repr
			except AttributeError:
				te.doc = 'PR ' + r.line.expedition.proforma.doc_repr
			
			try:
				te.date = r.line.expedition.proforma.invoice.date
			except AttributeError:
				te.date = r.line.expedition.proforma.date
			
			te.partner = r.line.expedition.proforma.partner_name
			
			te.picking = r.created_on
			
			self.entries.append(te)
		
		# query = db.session.query(db.WhIncomingRmaLine).join(db.WhIncomingRma).join(db.SaleInvoice) \
		#     .where(db.WhIncomingRmaLine.sn == imei)\
		#     .where(db.WhIncomingRmaLine.accepted == 'y')
		#
		# for r in query:
		#     te = TraceEntry()
		#     te.operation = 'Incoming Rma'
		#
		#     # Find concrete sale
		#     for sale in r.wh_incoming_rma.invoices:
		#         for line in sale.proformas[0].credit_note_lines:
		#             if line.sn == imei:
		#                 break
		#
		#     te.doc = 'FR ' + sale.doc_repr
		#     te.date = sale.date
		#     te.partner = r.wh_incoming_rma.incoming_rma.lines[0].cust
		#     te.picking = 'Not Registered'
		#
		#     self.entries.append(te)
		
		query = db.session.query(db.CreditNoteLine).join(db.SaleProforma).join(db.SaleInvoice). \
			where(db.CreditNoteLine.sn == imei)
		
		for r in query:
			te = TraceEntry()
			te.operation = 'Incoming Rma'
			
			invoice = r.proforma.invoice
			
			te.doc = 'FR ' + invoice.doc_repr
			te.date = invoice.date
			te.partner = invoice.partner_name
			te.picking = 'Not registered'
			
			self.entries.append(te)
		
		self.entries.sort(key=lambda te: te.date)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			te = self.entries[row]
			if col == self.OPERATION:
				return te.operation
			elif col == self.DOC:
				return te.doc
			elif col == self.DATE:
				return str(te.date)
			elif col == self.PARTNER:
				return te.partner
			elif col == self.PICKING:
				return str(te.picking)


def find_last_description(sn):
	try:
		obj = db.session.query(db.Imei).where(db.Imei.imei == sn).one()
		obj: db.Imei
	except NoResultFound:
		obj = None
	
	if obj:
		return ', '.join((obj.item.clean_repr, obj.condition + ' Condt.',
		                  obj.spec + ' Spec'))
	else:
		exp_series = db.session.query(db.ExpeditionSerie). \
			join(db.ExpeditionLine).where(db.ExpeditionSerie.serie == sn).all()
		
		try:
			exp_serie = exp_series[-1]
		except IndexError:
			return "Not found neither in stock nor in sales."
		else:
			line = exp_serie.line
			return ', '.join(
				(
					line.item.clean_repr,
					line.condition + ' Condt.',
					line.spec + ' Spec'
				)
			)


def where_was_it_sold(serie):
	try:
		obj = db.session.query(db.ExpeditionSerie).where(db.ExpeditionSerie.serie == serie).all()[-1]
		return obj.line.expedition.proforma.invoice.doc_repr
	except NoResultFound:
		pass


def build_credit_note_and_commit(partner_id, agent_id, order, candidates):
	from datetime import datetime
	
	proforma = db.SaleProforma()
	proforma.type = order.lines[0].invoice_type
	
	proforma.number = get_next_num(db.SaleProforma, proforma.type)
	
	proforma.partner_id = partner_id
	proforma.warehouse_id = None
	proforma.date = datetime.now().date()
	proforma.cancelled = False
	proforma.agent_id = agent_id
	proforma.courier_id = 1
	proforma.shipping_address_id = (
		db.session.query(db.ShippingAddress)
		.where(db.ShippingAddress.partner_id == partner_id)
		.first().id
	)
	
	text = ''
	
	credit_notes_where_sold = set(where_was_it_sold(line.sn) for line in candidates)
	
	for doc_repr in credit_notes_where_sold:
		text += doc_repr + '/'
	
	for wh_line in candidates:
		proforma.credit_note_lines.append(db.CreditNoteLine(wh_line))
	
	proforma.note = text
	db.session.add(proforma)
	db.session.commit()
	return proforma


class AvailableEntry:
	
	def __init__(self, sale_invoice, credit_note):
		self.sale_id = sale_invoice.id
		self.credit_id = credit_note.id
		self.doc_repr = credit_note.doc_repr
		self.total = credit_note.total_debt
		self.applied = credit_note.applied
		self.paid = credit_note.paid
		self._applying = 0
	
	@property
	def applying(self):
		return self._applying
	
	@applying.setter
	def applying(self, v):
		if v == 't':
			self._applying = self.total - self.applied - self.paid
			return
		
		v = float(v)
		if v > 0:
			raise ValueError
		if v < self.total - self.applied - self.paid:
			raise ValueError
		
		self._applying = v


class AvailableCreditNotesModel(BaseTable, QtCore.QAbstractTableModel):
	DOC_REPR, TOTAL, APPLIED, PAID, APPLYING = 0, 1, 2, 3, 4
	
	def __init__(self, invoice):
		super().__init__()
		self.invoice = invoice
		self._data = self._get_data()
		
		self.name = '_data'
		self._headerData = ['Document', 'Total', 'Applied', 'Returned', 'Applying(Editable)']
	
	def sort(self, column: int, order: Qt.SortOrder = ...) -> None:
		
		if column == self.DOC_REPR:
			self.layoutAboutToBeChanged().emit()
			self._data = sorted(self._data, key=lambda o: o.doc_repr,
			                    reverse=True if order == Qt.DescendingOrder else False)
			self.layoutChanged()
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		if role == Qt.DisplayRole:
			entry = self._data[row]
			return [
				entry.doc_repr,
				entry.total,
				entry.applied,
				entry.paid,
				entry.applying
			][column]
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		entry = self._data[row]
		if role == Qt.EditRole:
			if column == self.APPLYING:
				try:
					entry.applying = value
					return True
				except ValueError as ex:
					return False
			return False
		return False
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return
		col = index.column()
		if col == self.APPLYING:
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def _get_data(self):
		data = []
		
		sq = db.session.query(db.ManyManySales.credit_id).distinct()
		
		q = db.session.query(db.SaleInvoice).join(db.SaleProforma).where(
			db.SaleProforma.warehouse_id == None,
			db.SaleProforma.partner_id == self.invoice.partner_object.id,
			db.SaleInvoice.id.not_in(sq)
		)
		
		data.extend([AvailableEntry(self.invoice, c) for c in q])
		
		credit_applied_query = db.session.query(db.ManyManySales.credit_id,
		                                        func.sum(db.ManyManySales.fraction).label('applied')) \
			.group_by(db.ManyManySales.credit_id)
		
		credit_applied_dict = dict(r for r in credit_applied_query)
		
		invoice_id_line_query = db.session.query(
			db.SaleInvoice.id,
			db.CreditNoteLine.price * db.CreditNoteLine.quantity * (1 + db.CreditNoteLine.tax / 100)
		).join(
			db.SaleProforma, db.SaleProforma.sale_invoice_id == db.SaleInvoice.id
		).join(
			db.CreditNoteLine, db.SaleProforma.id == db.CreditNoteLine.proforma_id
		)
		
		key = lambda r: r.id
		invoice_id_lines = sorted(invoice_id_line_query.all(), key=key)
		
		invoice_id_total_dict = dict((id, sum(r[1] for r in group)) for id, group in groupby(invoice_id_lines, key=key))
		
		candidate_ids = set()
		for credit_id in credit_applied_dict:
			if invoice_id_total_dict[credit_id] < credit_applied_dict[credit_id]:
				candidate_ids.add(credit_id)
		
		assert credit_applied_dict.keys() & invoice_id_total_dict.keys() == credit_applied_dict.keys()
		
		data.extend(
			[
				AvailableEntry(self.invoice, c)
				for c in db.session.query(db.SaleInvoice).where(db.SaleInvoice.id.in_(candidate_ids))
			]
		)
		
		return data
	
	def add(self):
		for obj in [o for o in self._data if o.applying != 0]:
			db.session.add(db.ManyManySales(obj.sale_id, obj.credit_id, obj.applying))
		try:
			db.session.commit()
		except sqlalchemy.exc.IntegrityError:
			db.session.rollback()
			raise ValueError('You cant apply same credit note to same invoice twice.')


class AppliedCreditNotesModel(BaseTable, QtCore.QAbstractTableModel):
	DOC_REPR, FRACTION = 0, 1
	
	def __init__(self, invoice: db.SaleInvoice):
		super().__init__()
		self.invoice = invoice
		self._data = invoice.return_discounts
		self.name = '_data'
		self._headerData = ['Document', 'Rate Applied']
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			obj = self._data[row]
			if col == self.DOC_REPR:
				return obj.credit_note.doc_repr
			elif col == self.FRACTION:
				return obj.fraction
	
	def delete(self, rows):
		for row in rows:
			db.session.delete(self._data[row])
		db.session.commit()


class WhereCreditNotesModel(BaseTable, QtCore.QAbstractTableModel):
	DOC_REPR, FRACTION = 0, 1
	
	def __init__(self, credit_note: db.SaleInvoice):
		super().__init__()
		self.credit_note = credit_note
		self._data = credit_note.wasted_discounts
		self.name = '_data'
		self._headerData = ['Document', 'Rate Applied']
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		if role == Qt.DisplayRole:
			obj = self._data[row]
			if column == self.DOC_REPR:
				return obj.sale.doc_repr
			elif column == self.FRACTION:
				return obj.fraction
	
	@property
	def total(self):
		return sum(o.fraction for o in self._data)


class SIIInvoice:
	
	def __init__(self, invoice):
		self.invoice_number = invoice.doc_repr
		self.partner_name = invoice.partner_name
		self.partner_ident = invoice.partner_object.fiscal_number
		self.country_code = utils.get_country_code(invoice.partner_object.billing_country)
		self.invoice_date = invoice.date.strftime('%d-%m-%Y')
		
		self.lines = []
		
		lines = []
		for proforma in invoice.proformas:
			lines.extend(proforma.lines or proforma.advanced_lines or proforma.credit_note_lines)
		
		for line in lines:
			self.lines.append(SIILine(line))
		
		self.ambos_tax = len({line.tax for line in self.lines}) == 2
	
	def __repr__(self):
		name = self.__class__.__name__
		return f"{name}({self.invoice_number}, {self.partner_name}, \"" \
		       f"{self.partner_ident}, {self.country_code})"


class SIILine:
	
	def __init__(self, line):
		self.quantity = line.quantity
		self.price = line.price
		self.tax = line.tax
		self.is_stock = False
		if isinstance(line, db.SaleProformaLine):
			if line.item_id is not None:
				self.is_stock = True
		elif isinstance(line, db.AdvancedLine):
			if line.item_id is not None or line.mixed_description is not None:
				self.is_stock = True
		elif isinstance(line, db.CreditNoteLine):
			self.is_stock = True
		self.line_type = self.resolve_line_type()
	
	def resolve_line_type(self):
		booleans = (self.is_stock, self.tax != 0, self.price > 0)
		return int(''.join(str(int(b)) for b in booleans), base=2)
	
	def __repr__(self):
		name = self.__class__.__name__
		return f"{name}({self.quantity}, {self.price}, {self.tax}, {self.is_stock}) "


def do_sii(_from=None, to=None, series=None):
	import os
	jsonfeed = os.path.abspath(r'.\data.json')
	jsonresponse = os.path.abspath(r'.\response.json')

	print(jsonfeed)
	print(jsonresponse) 

	from db import (
		session,
		SaleInvoice,
		SaleProforma,
		Partner
	)
	
	siiinovices: list[SIIInvoice] = []
	for invoice in session.query(SaleInvoice).join(SaleProforma).join(Partner). \
			where(SaleInvoice.date >= _from). \
			where(SaleInvoice.date <= to). \
			where(SaleInvoice.type.in_(series)):
		siiinovices.append(SIIInvoice(invoice))
	
	import json
	import subprocess
	
	with open(jsonfeed, 'w') as fp:
		print('with open jsonfeed as fp')
		json.dump(siiinovices, default=lambda o: o.__dict__, fp=fp, indent=4)

	completed_subprocess = subprocess.run(['sii.exe', jsonfeed, jsonresponse], shell=True)
	
	print(completed_subprocess)

	if completed_subprocess.returncode == 0:
		with open(jsonresponse, 'r') as fp:
			return json.load(fp)
	else:
		raise ValueError('Subprocess failed')


class SIILogModel(BaseTable, QtCore.QAbstractTableModel):
	NUMBER, DATE, STATUS, MESSAGE = 0, 1, 2, 3
	
	def __init__(self, registers):
		super().__init__()
		self._headerData = ['Invoice Number', 'Date', 'Status', 'Message']
		self.registers = registers
		self.name = 'registers'
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		
		if not index.isValid():
			return
		
		if role == Qt.DisplayRole:
			row, column = index.row(), index.column()
			reg = self.registers[row]
			if column == self.NUMBER:
				return reg['invoice_number']
			elif column == self.DATE:
				return reg['invoice_date']
			elif column == self.STATUS:
				return reg['invoice_status']
			elif column == self.MESSAGE:
				return reg['message']
	
	def sort(self, column: int, order: Qt.SortOrder = ...) -> None:
		item = {
			self.NUMBER: 'invoice_number',
			self.STATUS: 'invoice_status',
			self.DATE: 'invoice_date',
			self.MESSAGE: 'message'
		}.get(column)
		
		if item:
			self.layoutAboutToBeChanged.emit()
			self.registers = sorted(self.registers, key=operator.itemgetter(item),
			                        reverse=True if order == Qt.DescendingOrder else False)
			self.layoutChanged.emit()


candidates = [
	'ship',
	'envio',
	'envío',
	'dhl',
	'courier',
	'freight',
	'freigth',
	'service fee',
	'Auto/Imported',

]

candidates.extend([c.description for c in db.session.query(db.Courier)])


# Works for both purchases and sales
@functools.cache
def get_avg_rate(proforma):
	base, rated_base = 0, 0
	for p in proforma.payments:
		base += p.amount
		rated_base += p.amount / p.rate
	try:
		return base / rated_base
	except ZeroDivisionError:
		return 1


@functools.cache
def get_purchase_expenses_breakdown(proforma):
	shipping_cost = 0
	remaining_cost = 0
	
	avg_rate = get_avg_rate(proforma)
	
	# Search in lines:
	for line in proforma.lines:
		if line.item_id is not None or line.description in utils.descriptions:
			continue
		else:
			for candidate in candidates:
				if line.description.lower().find(candidate.lower()) != -1:
					shipping_cost += line.price * line.quantity * (1 + line.tax / 100)
					break
			else:
				remaining_cost += line.price * line.quantity * (1 + line.tax / 100)
	
	# Search in associated expenses:
	for expense in proforma.expenses:
		for candidate in candidates:
			if expense.note.lower().find(candidate.lower()) != -1:
				shipping_cost += expense.amount
				break
		else:
			remaining_cost += expense.amount
	
	return remaining_cost, shipping_cost


@functools.cache
def get_purchase_stock_value(proforma):
	stock_value = 0
	for line in proforma.lines:
		if line.item_id is not None or line.description in utils.descriptions:
			stock_value += line.price * line.quantity * (1 + line.tax / 100)
	
	return stock_value


fields = """
    pdoc_type pdoc_number pdate ppartner pagent pitem 
    pcond pspec pserie pdollar prate peuro pfinancial_status pship pexpenses 
    ptotal
    """

defaults = ('',) * len(fields.split())
PurchaseRow = namedtuple('PurchaseRow', field_names=fields, defaults=defaults)


def do_cost_price(imei):
	try:
		rec_serie = db.session.query(db.ReceptionSerie).join(db.ReceptionLine) \
			.join(db.Reception).join(db.PurchaseProforma) \
			.join(db.Partner, db.Partner.id == db.PurchaseProforma.partner_id) \
			.join(db.Agent, db.Agent.id == db.PurchaseProforma.agent_id) \
			.where(db.ReceptionSerie.serie == imei).all()[-1]  # take last
	
	except IndexError:
		return PurchaseRow()
	else:
		
		proforma = rec_serie.line.reception.proforma
		rec_line = rec_serie.line
		
		proforma_line = None
		for aux in proforma.lines:
			if aux == rec_line:
				proforma_line = aux
				break
		
		if not proforma_line:
			raise ValueError('Fatal error: could not match proforma purchase with warehouse')
		
		avg_rate = get_avg_rate(proforma)
		base_price = proforma_line.price
		no_shipping_expense, shipping_expense = get_purchase_expenses_breakdown(proforma)  # already rate applied
		shipping_delta = shipping_expense / proforma.total_quantity
		remaining_expense_delta = base_price * no_shipping_expense / get_purchase_stock_value(proforma)  # formula
		
		try:
			doc_number = proforma.invoice.doc_repr
			doc_type = 'Invoice'
			date = proforma.invoice.date.strftime('%d/%m/%Y')
		except AttributeError:
			doc_number = proforma.doc_repr
			doc_type = 'Proforma'
			date = proforma.date.strftime('%d/%m/%Y')
		
		if math.isclose(avg_rate, 1):
			dollar = 'Unknown'
			euro = base_price
		else:
			dollar = base_price
			euro = base_price / avg_rate
		
		try:
			item = proforma_line.item.clean_repr
		except AttributeError:
			item = proforma_line.description
		
		return PurchaseRow(
			pdoc_type=doc_type,
			pdoc_number=doc_number,
			pdate=date,
			ppartner=proforma.partner_name,
			pagent=proforma.agent.fiscal_name.split()[0],
			pitem=item,
			pcond=rec_serie.line.condition,
			pspec=rec_serie.line.spec,
			pserie=imei,
			pdollar=dollar,
			prate=avg_rate,
			peuro=euro,
			pfinancial_status=proforma.financial_status_string,
			pship=shipping_delta,
			pexpenses=remaining_expense_delta,
			ptotal=euro + shipping_delta + remaining_expense_delta
		)


@functools.cache
def get_sale_stock_key(proforma):
	if proforma.lines:
		return lambda line: line.item_id is not None
	elif proforma.advanced_lines:
		return lambda line: line.item_id is not None or line.mixed_description is not None


@functools.cache
def get_sale_shipping_key(proforma):
	if proforma.lines:
		return lambda line, candidate: line.description.lower().find(candidate.lower()) != -1
	
	elif proforma.advanced_lines:
		return lambda line, candidate: line.free_description.lower().find(candidate.lower()) != -1


@functools.cache
def get_sale_terms_key(proforma):
	if proforma.lines:
		return lambda line: line.description.lower().find('term') != -1
	elif proforma.advanced_lines:
		return lambda line: line.free_description.lower().find('term') != -1


@functools.cache
def get_sale_breakdown(proforma):
	shipping_cost = 0
	remaining_cost = 0
	terms = 0
	
	avg_rate = get_avg_rate(proforma)
	
	stock_key = get_sale_stock_key(proforma)
	shipping_key = get_sale_shipping_key(proforma)
	term_key = get_sale_terms_key(proforma)
	
	for line in proforma.lines or proforma.advanced_lines:
		
		if stock_key(line):
			continue
		
		elif term_key(line):
			terms += line.price * line.quantity * (1 + line.tax / 100) / avg_rate
		
		else:
			for candidate in candidates:
				if shipping_key(line, candidate):
					shipping_cost += line.price * line.quantity * (1 + line.tax / 100) / avg_rate
					break
			else:
				remaining_cost += line.price * line.quantity * (1 + line.tax / 100) / avg_rate
	
	for expense in proforma.expenses:
		for candidate in candidates:
			if expense.note.lower().find(candidate.lower()) != -1:
				shipping_cost += expense.amount
				break
		else:
			remaining_cost += expense.amount
	
	return remaining_cost, shipping_cost, terms


@functools.cache
def get_sale_proforma_stock_value(proforma):
	stock_value = 0
	stock_key = get_sale_stock_key(proforma)
	for line in proforma.lines or proforma.advanced_lines:
		if stock_key(line):
			stock_value += line.price * line.quantity * (1 + line.tax / 100)
	
	return stock_value


fields = """ sdoc_type sdoc_number sdate spartner sagent 
          sitem scond sspec sserie sdollar srate seuro sfinancial_status sship 
          sterm sexpenses stotal """

defaults = ('',) * (len(fields.split()) - 1)
defaults += (0,)

SaleRow = namedtuple('SaleRow', field_names=fields, defaults=defaults)


def get_rma_expenses(imei):
	try:
		proforma = db.session.query(db.SaleProforma).join(db.SaleInvoice).join(db.WhIncomingRma) \
			.join(db.WhIncomingRmaLine).where(db.WhIncomingRmaLine.sn == imei).all()[-1]
	except IndexError:
		return 0
	else:
		qnt = sum(line.quantity for line in proforma.credit_note_lines)
		expenses = sum(exp.amount for exp in proforma.expenses)
		return expenses / qnt


def do_sale_price(imei):
	try:
		exp = db.session.query(db.ExpeditionSerie).join(db.ExpeditionLine) \
			.join(db.Expedition).join(db.SaleProforma) \
			.join(db.Partner, db.Partner.id == db.SaleProforma.partner_id) \
			.join(db.Agent, db.Agent.id == db.SaleProforma.agent_id) \
			.where(db.ExpeditionSerie.serie == imei).all()[-1]  # take last
	
	except IndexError:  # No result found
		
		return SaleRow()
	
	else:
		
		proforma = exp.line.expedition.proforma
		exp_line = exp.line
		proforma_line = None
		
		# We need several process for each type of lines sale contains hence the switch
		if proforma.credit_note_lines:  # Rma check, Provisional, We need to define how to process rma
			return SaleRow()
		# We need several processes for each type of lines sale contains hence the switch
		elif proforma.lines:
			for aux in proforma.lines:
				if aux == exp_line:
					proforma_line = aux
		elif proforma.advanced_lines:
			for aux in proforma.advanced_lines:
				if aux.item_id is not None and aux == exp_line:
					proforma_line = aux
				elif aux.mixed_description is not None:
					if any(definition == exp_line for definition in aux.definitions):
						proforma_line = aux
		
		for aux in proforma.lines or proforma.advanced_lines:
			
			if aux == exp_line:
				proforma_line = aux
				break
		
		if not proforma_line:
			return SaleRow()
		
		# raise ValueError('Fatal error could not match warehouse with sale proforma')
		
		avg_rate = get_avg_rate(proforma)
		base_price = proforma_line.price
		
		remaining_expense, shipping_expense, terms = get_sale_breakdown(proforma)
		
		# Compute deltas:
		shipping_delta = shipping_expense / proforma.total_quantity
		terms_delta = terms / proforma.total_quantity
		remaining_expense_delta = base_price * remaining_expense / get_sale_proforma_stock_value(proforma)
		
		try:
			doc_number = proforma.invoice.doc_repr
			doc_type = 'Invoice'
			date = proforma.invoice.date.strftime('%d/%m/%Y')
		except AttributeError:
			doc_number = proforma.doc_repr
			doc_type = 'Proforma'
			date = proforma.date.strftime('%d/%m/%Y')
		
		if math.isclose(avg_rate, 1):
			dollar = 'Unknown'
			euro = base_price
		else:
			dollar = base_price
			euro = base_price / avg_rate
		
		return SaleRow(
			sdoc_type=doc_type,
			sdoc_number=doc_number,
			sdate=date,
			spartner=proforma.partner_name,
			sagent=proforma.agent.fiscal_name.split()[0],
			sitem=exp.line.item.clean_repr,
			scond=exp.line.condition,
			sspec=exp.line.spec,
			sserie=imei,
			srate=avg_rate,
			sdollar=dollar,
			seuro=euro,
			sfinancial_status=proforma.financial_status_string,
			sship=shipping_delta,
			sterm=terms_delta,
			sexpenses=remaining_expense_delta,
			stotal=euro - shipping_delta + terms_delta - remaining_expense_delta
		)


import typing as tp 

class HarvestModel(QtCore.QAbstractListModel):
	
	def __init__(self):
		super(HarvestModel, self).__init__()
		self._elements = []
	
	def rowCount(self, parent: QModelIndex = ...) -> int:
		return len(self._elements)
	
	def columnCount(self, parent: QModelIndex = ...) -> int:
		return 1
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		if role == Qt.DisplayRole:
			return self._elements[index.row()]
	
	def add(self, element):
		self.layoutAboutToBeChanged.emit()
		self._elements.append(element)
		self.layoutChanged.emit()
	
	def delete(self, rows):
		self.layoutAboutToBeChanged.emit()
		for row in sorted(rows, reverse=True):
			try:
				del self._elements[row]
			except IndexError:
				pass
		self.layoutChanged.emit()
	
	def delete_all(self):
		self.layoutAboutToBeChanged.emit()
		self._elements.clear()
		self.layoutChanged.emit()
	
	@property
	def elements(self):
		return self._elements


class OutputRegister:
	
	def __init__(self, purchase_row, sale_row):
		self.__dict__.update(purchase_row._asdict())
		self.__dict__.update(sale_row._asdict())
	
	@property
	def astuple(self):
		return tuple(self.__dict__.values())


def zipped(doc_numbers):
	for doc_number in doc_numbers:
		type, number = doc_number.split('-')
		yield type, number


def append_registers(obj, query=None, series=None):
	iterator = series or query  # Exclusive
	for register in iterator:
		try:
			serie = register.serie
		except AttributeError:
			serie = register
		obj._registers.append(OutputRegister(do_cost_price(serie), do_sale_price(serie)))


class OutputModel(BaseTable, QtCore.QAbstractTableModel):
	PDOCTYPE, PNUMBER, PDATE, PPARTNER, PAGENT, PDESCRIPTION, PCONDITION, PSPEC, PSERIAL, PDOLAR, PRATE, PEURO, \
		PFINANCIAL_STATUS, PSHIPPING, PEXPENSES, PTOTAL, SDOCTYPE, SDOCNUMBER, SDATE, SPARTNER, SAGENT, SDESCRIPTION, SCONDITION, \
		SSPEC, SSERIAL, SDOLAR, SRATE, SEURO, SFINANCIAL_STATUS, SSHIPPING, STERMS, SEXPENSES, STOTAL, SHARVEST = \
		0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, \
			21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33
	
	def __init__(self):
		super().__init__()
		self.name = '_registers'
		self._registers = []
		self._headerData = [
			'Doc. Type', 'Doc. Nº', 'Date', 'Partner', 'Agent', 'Product', 'Condition', 'Spec.', 'Serial',
			'$', 'Rate', '€', 'Status', 'Shipping', 'Expenses', 'Total Cost €', 'Doc. Type', 'Doc. Nº', 'Date',
			'Partner',
			'Agent', 'Product', 'Condition', 'Spec', 'Serial', '$', 'Rate', '€', 'Status', 'Shipping', 'Terms',
			'Expenses', 'Total Income', 'Harvest'
		]
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		register = self._registers[row]
		
		if role == Qt.DisplayRole:
			if col == self.PDOCTYPE:
				return register.pdoc_type
			elif col == self.PNUMBER:
				return register.pdoc_number
			elif col == self.PDATE:
				return register.pdate
			elif col == self.PPARTNER:
				return register.ppartner
			elif col == self.PAGENT:
				return register.pagent
			elif col == self.PDESCRIPTION:
				return register.pitem
			elif col == self.PCONDITION:
				return register.pcond
			elif col == self.PSPEC:
				return register.pspec
			elif col == self.PSERIAL:
				return register.pserie
			elif col == self.PDOLAR:
				return register.pdollar
			elif col == self.PRATE:
				return register.prate
			elif col == self.PEURO:
				return register.peuro
			elif col == self.PFINANCIAL_STATUS:
				return register.pfinancial_status
			elif col == self.PSHIPPING:
				return register.pship
			elif col == self.PEXPENSES:
				return register.pexpenses
			elif col == self.PTOTAL:
				return register.ptotal
			elif col == self.SDOCTYPE:
				return register.sdoc_type
			elif col == self.SDOCNUMBER:
				return register.sdoc_number
			elif col == self.SDATE:
				return register.sdate
			elif col == self.SPARTNER:
				return register.spartner
			elif col == self.SAGENT:
				return register.sagent
			elif col == self.SDESCRIPTION:
				return register.sitem
			elif col == self.SCONDITION:
				return register.scond
			elif col == self.SSPEC:
				return register.sspec
			elif col == self.SSERIAL:
				return register.sserie
			elif col == self.SDOLAR:
				return register.sdollar
			elif col == self.SRATE:
				return register.srate
			elif col == self.SEURO:
				return register.seuro
			elif col == self.SFINANCIAL_STATUS:
				return register.sfinancial_status
			elif col == self.SSHIPPING:
				return register.sship
			elif col == self.STERMS:
				return register.sterm
			elif col == self.SEXPENSES:
				return register.sexpenses
			elif col == self.STOTAL:
				return register.stotal
			elif col == self.SHARVEST:
				if register.stotal == 0:
					return 0
				else:
					return register.stotal - register.ptotal
	
	def export(self, file):
		from openpyxl import Workbook
		wb = Workbook()
		ws = wb.active
		ws.append(self._headerData)
		for register in self._registers:
			try:
				ws.append(register.astuple + (register.stotal - register.ptotal,))
			except TypeError:
				ws.append(register.astuple + ('Could not find a result',))
		
		wb.save(file)
	
	@classmethod
	def by_period(cls, _from, to, _input=False, agent_id=None, exclude_at_capital=False):
		self = cls()
		if _input:
			query = db.session.query(db.ReceptionSerie.serie).join(db.ReceptionLine). \
				join(db.Reception).join(db.PurchaseProforma). \
				where(db.PurchaseProforma.date >= _from,
			          db.PurchaseProforma.date <= to)
			
			if exclude_at_capital:
				query = query.where(db.PurchaseProforma.partner_id != 19)
		else:
			query = db.session.query(db.ExpeditionSerie.serie).join(db.ExpeditionLine). \
				join(db.Expedition).join(db.SaleProforma). \
				where(db.SaleProforma.date >= _from,
			          db.SaleProforma.date <= to)
			
			if exclude_at_capital:
				query = query.where(db.SaleProforma.partner_id != 19)
		
		if agent_id:
			if _input:
				query = query.where(db.PurchaseProforma.agent_id == agent_id)
			else:
				query = query.where(db.SaleProforma.agent_id == agent_id)
		
		append_registers(self, query=query)
		
		return self
	
	@classmethod
	def by_document(cls, type_dict, doc_numbers: tp.Optional[list], partner_id=tp.Optional[int], 
		year:tp.Optional[int]=None):

		self = cls()
		predicates = []
		
		sales_query = db.session.query(db.ExpeditionSerie).join(db.ExpeditionLine).join(db.Expedition)
		sales_query = sales_query.join(db.SaleProforma)
		
		purchases_query = db.session.query(db.ReceptionSerie).join(db.ReceptionLine).join(db.Reception)
		purchases_query = purchases_query.join(db.PurchaseProforma)
		
		if type_dict['PurchaseProforma']:
			cls = db.PurchaseProforma
			query = purchases_query

			# Extend with partner:
			if partner_id:
				query = query.where(db.PurchaseProforma.partner_id == partner_id)
		
		elif type_dict['SaleProforma']:
			cls = db.SaleProforma
			query = sales_query
			
			# Extend with partner:
			if partner_id:
				query = query.where(db.SaleProforma.partner_id == partner_id)
		
		elif type_dict['SaleInvoice']:
			query = sales_query.join(db.SaleInvoice)
			cls = db.SaleInvoice
	
			# Extend with partner: 
			if partner_id:
				query = query.where(db.SaleProforma.partner_id == partner_id)
		
		elif type_dict['PurchaseInvoice']:
			query = purchases_query.join(db.PurchaseInvoice)
			cls = db.PurchaseInvoice
			
			# Extend with partner:
			if partner_id:
				query = query.where(db.PurchaseProforma.partner_id == partner_id)
		
		# for type, number in zipped(doc_numbers):
		# 	predicates.append(and_(cls.type == type, cls.number == number))
		
		if doc_numbers:
			query = query.where(or_(*[and_(cls.type == type, cls.number == number) for type, number in zipped(doc_numbers)]))

		if year: 
			query = query.where(func.year(cls.date) == year)

		print("Query Built:", "-" * 50)
		print(query)
		print("Query Built:","-" * 50)

		append_registers(self, query=query)

		
		return self
	
	@classmethod
	def by_serials(cls, serials):
		self = cls()
		append_registers(self, series=serials)
		return self


def extract_doc_repr(invoice_text):
	from re import search
	return invoice_text[slice(*re.search(r'[1-6]\-0*\d+\Z', invoice_text).span())]


def add_expense(doc_repr, amount):
	from datetime import datetime 
	year = datetime.now().year
	type, number = doc_repr.split('-')
	type, number = int(type), int(number)
	
	sale_proformas = db.session.query(db.SaleProforma).join(db.SaleInvoice) \
		.where(
			db.SaleInvoice.type == type, 
			db.SaleInvoice.number == number, 
			func.year(db.SaleInvoice.date) == year).all()
	
	if len(sale_proformas) != 1:
		raise ValueError('Multiple Proformas. Dont know where to put the expenses')
	
	sale_proforma = sale_proformas[0]
	
	from datetime import datetime
	db.SaleExpense(datetime.now(), amount, 'Auto/Imported', sale_proforma)
	
	db.session.commit()


def resolve_dhl_expenses(file_path):
	exc_mss = "No reconozco esta estructura de archivo, notificar a Andrei si quieres que funcione."
	resolved = []
	unresolved = []
	
	with open(file_path, 'r') as fp:
		reader = csv.DictReader(fp)
		for dict_row in reader:
			try:
				company = dict_row.get('Senders Name', dict_row.get('Detalles Remitente'))
				if not company:
					raise ValueError(exc_mss)

				if company.lower().find('euromedia') != -1:
					invoice_text = dict_row.get('Shipment Reference 1', dict_row.get('Referencia'))
					if not invoice_text:
						raise ValueError(exc_mss)
					
					try:
						doc_repr = extract_doc_repr(invoice_text)

					except AttributeError:
						unresolved.append(invoice_text)
					else:
						amount = dict_row.get('Total amount (excl. VAT)', dict_row.get('Importe Bruto'))
						add_expense(doc_repr, amount)
						resolved.append(doc_repr)
			except KeyError as ex:

				raise ValueError(exc_mss)

		return resolved, unresolved


def update_description(imei, to_item_id):
	db.session.query(db.CreditNoteLine).filter(db.CreditNoteLine.sn == imei) \
		.update({db.CreditNoteLine.item_id: to_item_id})
	
	db.session.query(db.Imei).filter(db.Imei.imei == imei) \
		.update({db.Imei.item_id: to_item_id})
	
	db.session.query(db.ImeiMask).filter(db.ImeiMask.imei == imei) \
		.update({db.ImeiMask.item_id: to_item_id})
	
	db.session.query(db.IncomingRmaLine).filter(db.IncomingRmaLine.sn == imei) \
		.update({db.IncomingRmaLine.item_id: to_item_id})
	
	db.session.query(db.ReceptionSerie).filter(db.ReceptionSerie.serie == imei) \
		.update({db.ReceptionSerie.item_id: to_item_id})
	
	db.session.query(db.WhIncomingRmaLine).filter(db.WhIncomingRmaLine.sn == imei) \
		.update({db.WhIncomingRmaLine.item_id: to_item_id})
	
	try:
		line = db.session.query(db.ExpeditionLine).join(db.ExpeditionSerie) \
			.where(db.ExpeditionSerie.serie == imei).all()[-1]
	except IndexError:
		pass
	else:
		line.item_id = to_item_id
	
	try:
		db.session.commit()
	except Exception as ex:
		raise ValueError(str(ex))


def find_item_id_from_serie(serie):
	return db.session.query(db.Imei.item_id).where(db.Imei.imei == serie).scalar()


class Fuck:
	__slots__ = ['serie', 'from_', 'to']
	
	def __init__(self, serie, from_, to):
		self.serie = serie
		self.from_ = from_
		self.to = to
	
	def __repr__(self):
		cls_name = self.__class__.__name__
		return f"{cls_name}(serie={self.serie}, from={self.from_}, to={self.to})"


class FuckExcel:
	
	def __init__(self, sale_invoice: db.SaleInvoice):
		partner = sale_invoice.partner_object
		self.serie = sale_invoice.type
		self.number = sale_invoice.number
		self.doc_repr = sale_invoice.doc_repr
		self.date = sale_invoice.date.strftime('%d/%m/%Y')
		self.code = self.get_code(partner)
		self.customer = sale_invoice.partner_name
		self.nif = partner.fiscal_number
		self.address = ' '.join((partner.billing_line1, partner.billing_line2))
		self.postcode = partner.billing_postcode
		self.state = partner.billing_state
		self.payment = 'Transferencia'
		self.total = sale_invoice.total_debt
		self.tax = sale_invoice.tax
		if sale_invoice.is_credit_note:
			self.obs = sale_invoice.cn_repr
		else:
			self.obs = sale_invoice.note
	
	def get_code(self, partner):
		from io import StringIO
		
		if not partner.fiscal_number:
			return ""
		
		buffer = StringIO()
		for s in partner.fiscal_number:
			if s.isdigit():
				buffer.write(s)
		
		try:
			return "4300" + buffer.getvalue()[len(buffer.getvalue()) - 5:]
		except IndexError:
			return "4300" + buffer.getvalue()
	
	@property
	def as_tuple(self):
		return tuple(self.__dict__.values())


def get_month_name(month):
	return {
		1: 'Enero',
		2: 'Febrero',
		3: 'Marzo',
		4: 'Abril',
		5: 'Mayo',
		6: 'Junio',
		7: 'Julio',
		8: 'Agosto',
		9: 'Septiembre',
		10: 'Octubre',
		11: 'Noviembre',
		12: 'Diciembre'
	}.get(month)


def month_key(sale):
	return sale.date.month


class FucksModel(BaseTable, QtCore.QAbstractTableModel):
	SERIE, FROM, TO = 0, 1, 2
	
	def __init__(self):
		super().__init__()
		self._headerData = ['Serie', 'From', 'To']
		self.name = '_fucks'
		self._fucks = [Fuck(i, '', '') for i in range(1, 7)]
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, column = index.row(), index.column()
		if role == Qt.DisplayRole:
			fuck = self._fucks[row]
			if column == self.SERIE:
				return fuck.serie
			elif column == self.FROM:
				return fuck.from_
			elif column == self.TO:
				return fuck.to
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return
		
		column = index.column()
		if column in (self.FROM, self.TO):
			return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
		else:
			return Qt.ItemFlags(~Qt.ItemIsEditable)
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return
		try:
			number = int(value)
		except ValueError:
			return False
		
		row, column = index.row(), index.column()
		fuck = self._fucks[row]
		if role == Qt.EditRole:
			if column == self.FROM:
				fuck.from_ = number
				return True
			elif column == self.TO:
				fuck.to = number
			return False
		return False
	
	def export(self, year):
		import os
		from openpyxl import Workbook
		
		header = [
			'Serie Factura',
			'Nº Factura',
			'Factura',
			'Fecha',
			'Código contable',
			'Cliente',
			'NIF',
			'Dirección',
			'Código postal',
			'Población',
			'Forma pago',
			'Total Factura',
			'IVA',
			'Observaciones',
		]
		
		for fuck in self.fucks:
			query = self.get_query(fuck.serie, fuck.from_, fuck.to, year)
			sales = query.all()
			sales = sorted(sales, key=month_key)
			for month, sales_group in groupby(sales, key=month_key):
				target_dir = self.get_target_dir(month, year)
				wb = Workbook()
				ws = wb.active
				ws.append(header)
				for sale in list(sales_group):
					ws.append(FuckExcel(sale).as_tuple)
				
				filename = ''.join(('serie', str(fuck.serie), '.xlsx'))
				path = os.path.join(target_dir, filename)
				try:
					os.remove(path)
				except (FileNotFoundError, OSError) as ex:
					print(f'Logging: Exception raised: {ex}')
				finally:
					wb.save(os.path.join(target_dir, filename))
	
	@staticmethod
	def get_query(serie, _from, to, year):
		
		serie, _from, to, year = int(serie), int(_from), int(to), int(year)
		
		return db.session.query(db.SaleInvoice).where(
			db.SaleInvoice.type == serie,
			db.SaleInvoice.number >= _from,
			db.SaleInvoice.number <= to,
			func.year(db.SaleInvoice.date) == year
		)
	
	@property
	def fucks(self):
		return [fuck for fuck in self._fucks if fuck.from_ != '']
	
	@staticmethod
	def get_target_dir(month, year):
		path = os.path.join(
			os.environ['USERPROFILE'], 'Dropbox', 'Spain', 'Gestoria', 'Euromedia',
			'Facturas Ingreso', 'Automaticos', year, get_month_name(month)
		)
		if not os.path.exists(path):
			os.makedirs(path)
		return path

class Tupable:
	
	@property
	def as_tuple(self):
		return tuple(self.__dict__.values())


class StockValuationEntryDocument(Tupable):
	pass


class StockValuationEntryImei(Tupable):
	
	def __init__(self, imei:str, purchase_row: PurchaseRow):
		self.imei = imei
		self.description = purchase_row.pitem
		self.condition = purchase_row.pcond
		self.spec = purchase_row.pspec
		self.serial = purchase_row.pserie
		self.date = purchase_row.pdate
		self.partner = purchase_row.ppartner
		
		prefix = 'INV ' if purchase_row.pdoc_type.startswith('Inv') else 'PI '
		self.doc = prefix + purchase_row.pdoc_number
		
		self.cost = purchase_row.ptotal


def get_external_from_imei(imei):
	try:
		return db.session.query(db.PurchaseProforma.external).join(db.Reception). \
			join(db.ReceptionLine).join(db.ReceptionSerie). \
			where(db.ReceptionSerie.serie == imei).all()[-1].external
	except IndexError:
		return ''


class Exportable:
	
	def export(self, filepath):
		from openpyxl import Workbook
		wb = Workbook()
		ws = wb.active
		ws.append(self._headerData)
		for entry in self.entries:
			if isinstance(entry, tuple):
				ws.append(entry)
			else:
				ws.append(entry.as_tuple)
		
		wb.save(filepath)


class StockValuationModelDocument(Exportable, BaseTable, QtCore.QAbstractTableModel):
	DESC, COND, SPEC, QNT, DATE, PARTNER, DOC_REPR, COST = 0, 1, 2, 3, 4, 5, 6, 7
	
	def __init__(self, type, number, year, proforma=False):
		super().__init__()
		self.name = 'entries'
		self.entries = []
		self._headerData = [
			'Description',
			'Condition',
			'Spec',
			'Qnt',
			'Purchase Date',
			'Partner',
			'Nº Doc',
			'Cost'
		]
		
		self.init_data(type, number, year, proforma=proforma)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			reg = self.entries[row]
			if col == self.DESC:
				return reg.description
			elif col == self.COND:
				return reg.condition
			elif col == self.SPEC:
				return reg.spec
			elif col == self.QNT:
				return reg.qnt
			elif col == self.DATE:
				return reg.date
			elif col == self.PARTNER:
				return reg.partner
			elif col == self.DOC_REPR:
				return reg.doc
			elif col == self.COST:
				return reg.cost
	
	def init_data(self, type, number, year, proforma=False):
		if not proforma:
			try:
				purchase_proforma = db.session.query(db.PurchaseProforma). \
					join(db.PurchaseInvoice).where(db.PurchaseInvoice.type == type). \
					where(db.PurchaseInvoice.number == number). \
					where(func.year(db.PurchaseInvoice.date) == year).one()
			except sqlalchemy.exc.NoResultFound:
				raise ValueError("No results found")
		else:
			
			try:
				purchase_proforma = db.session.query(db.PurchaseProforma). \
					where(db.PurchaseProforma.type == type). \
					where(db.PurchaseProforma.number == number). \
					where(func.year(db.PurchaseProforma.date) == year).one()
			
			except sqlalchemy.exc.NoResultFound:
				raise ValueError("No results found")
		
		for line in purchase_proforma.lines:
			
			if line.item_id is None and line.description not in utils.descriptions:
				continue
			
			entry = StockValuationEntryDocument()
			entry.description = line.description or line.item.clean_repr
			entry.condition = line.condition
			entry.spec = line.spec
			entry.qnt = line.quantity
			if not proforma:
				entry.date = purchase_proforma.invoice.date.strftime('%d/%m/%Y')
			else:
				entry.date = purchase_proforma.date.strftime('%d/%m/%Y')
			
			entry.partner = purchase_proforma.partner_name
			entry.doc = purchase_proforma.doc_repr if proforma else purchase_proforma.invoice.doc_repr
			
			avg_rate = get_avg_rate(purchase_proforma)
			if isinstance(avg_rate, str):
				raise ValueError('I could not find mean rate')
			
			base_price = line.price
			remaining_expense, shipping_expense = get_purchase_expenses_breakdown(
				purchase_proforma)  # already rate applied
			shipping_delta = shipping_expense / purchase_proforma.total_quantity
			remaining_expense_delta = base_price * remaining_expense / get_purchase_stock_value(purchase_proforma)
			
			entry.cost = base_price / avg_rate + shipping_delta + remaining_expense_delta
			
			self.entries.append(entry)


class StockValuationEntryWarehouse(Tupable):
	
	def __init__(self, purchase_row: PurchaseRow):
		self.description = purchase_row.pitem
		self.condition = purchase_row.pcond
		self.spec = purchase_row.pspec
		self.serial = purchase_row.pserie
		self.qnt = 1
		self.date = purchase_row.pdate
		self.partner = purchase_row.ppartner
		
		prefix = 'INV ' if purchase_row.pdoc_type.startswith('Inv') else 'PI '
		self.doc = prefix + purchase_row.pdoc_number
		
		self.cost = purchase_row.ptotal
	
	@property
	def as_tuple(self):
		return tuple(self.__dict__.values())


class StockValuationEntryWarehouseNoSerie(Tupable):
	
	def __init__(self, description, condition, spec, date, partner, doc, cost, qnt):
		self.description = description
		self.condition = condition
		self.spec = spec
		self.date = date
		self.partner = partner
		self.doc = doc
		self.cost = cost
		self.serial = ''
		self.qnt = qnt


class StockValuationModelWarehouse(Exportable, BaseTable, QtCore.QAbstractTableModel):
	DESC, COND, SPEC, SERIAL, QNT, DATE, PARTNER, DOC_REPR, COST = 0, 1, 2, 3, 4, 5, 6, 7, 8
	
	def __init__(self, warehouse_id, date=None):
		super().__init__()
		self.name = 'entries'
		self.entries = []
		self._headerData = [
			'Description',
			'Condition',
			'Spec',
			'Serial',
			'Qnt',
			'Purchase Date',
			'Partner',
			'Nº Doc',
			'Cost'
		]
		
		query = db.session.query(db.Imei.imei).where(db.Imei.warehouse_id == warehouse_id)
		
		registers = []
		for register in db.session.query(db.Imei.imei).where(db.Imei.warehouse_id == warehouse_id):
			registers.append(StockValuationEntryWarehouse(do_cost_price(register.imei)))
		
		has_not_serie = lambda object: utils.valid_uuid(object.imei)
		has_serie = lambda object: not utils.valid_uuid(object.imei)
		
		uuid_group = list(filter(has_not_serie, registers))
		sn_group = filter(has_serie, registers)
		
		key = operator.attrgetter('description', 'condition', 'spec', 'cost')
		uuid_group = sorted(uuid_group, key=key)
		
		for entry in sn_group:
			self.entries.append(entry)
		
		key = operator.attrgetter('description', 'condition', 'spec', 'date', 'partner', 'doc', 'cost')
		
		for key, group in groupby(iterable=uuid_group, key=key):
			description, condition, spec, date, partner, doc, cost = key
			
			self.entries.append(StockValuationEntryWarehouseNoSerie(
				description, condition, spec, date, partner, doc, cost, len(list(group)))
			)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			reg = self.entries[row]
			if col == self.DESC:
				return reg.description
			elif col == self.COND:
				return reg.condition
			elif col == self.SPEC:
				return reg.spec
			elif col == self.SERIAL:
				return reg.serial
			elif col == self.DATE:
				return reg.date
			elif col == self.PARTNER:
				return reg.partner
			elif col == self.DOC_REPR:
				return reg.doc
			elif col == self.COST:
				return reg.cost
			elif col == self.QNT:
				return reg.qnt


class WarehouseSimpleValueEntry:
	__slots__ = ('description', 'condition', 'spec', 'serial',
	             'purchase_date', 'partner', 'doc', 'currency', 'cost', 'qnt')
	
	def __init__(self, serie):
		
		rec_serie = db.session.query(
			db.ReceptionSerie
		).where(
			db.ReceptionSerie.serie == serie
		).order_by(
			db.ReceptionSerie.id.desc()
		).first()
		
		proforma = rec_serie.line.reception.proforma
		
		# Itera la factura si encuentra hace cosas
		# si no, hace blancos para que se
		# vean en el excel.
		
		for line in proforma.lines:
			if line == rec_serie.line:
				self.description = line.description or line.item.clean_repr
				self.condition = line.condition
				self.spec = line.spec
				self.serial = serie
				self.purchase_date = proforma.date.strftime('%d/%m/%Y')
				self.partner = proforma.partner_name
				self.doc = proforma.doc_repr
				self.qnt = 1
				
				if proforma.eur_currency:
					self.currency = 'EUR'
					self.cost = line.price
				
				else:
					if proforma.financial_status_string == 'Paid':
						self.currency = 'EUR'
						self.cost = line.price / get_avg_rate(proforma)
					else:
						self.currency = 'USD'
						self.cost = line.price
				
				break  # Muy importante, si no, el código del bloque else se ejecuta.
		
		else:
			self.description = self.condition = self.spec = \
				self.purchase_date = self.partner = self.doc = self.currency = ""
			self.serial = serie
			self.qnt = self.cost = 0
	
	@property
	def as_tuple(self):
		return (
			self.description, self.condition, self.spec,
			self.serial, self.purchase_date, self.partner, self.doc,
			self.currency, self.cost, self.qnt
		)
	
	def __repr__(self):
		return repr(self.as_tuple)


class WarehouseSimpleValueModel(Exportable, BaseTable, QtCore.QAbstractTableModel):
	
	def __init__(self, warehouse_id, date=None):
		super().__init__()
		self._headerData = [
			'Description',
			'Condition',
			'Spec',
			'Serial',
			'Purchase Date',
			'Partner',
			'Currency',
			'Nº Doc',
			'Cost'
		]
		
		self.name = 'entries'
		
		self.entries = []
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		
		if role == Qt.DisplayRole:
			return self.entries[index.row()].as_tuple[index.column()]


class StockValuationModelImei(Exportable, BaseTable, QtCore.QAbstractTableModel):
	IMEI, DESC, COND, SPEC, SERIAL, DATE, PARTNER, DOC_REPR, COST, QNT = 0, 1, 2, 3, 4, 5, 6, 7, 8, 10
	
	# As a consequence of the new batch feature, we now receive an iterable of imeis. 
	# if the client class wants just, 1, for the simple case, 
	# then it will send a list with imei. 

	# Note how easy was to generalize from 1 imei to n datasources of imeis. 

	# I love python!
	def __init__(self, *datasources): 
		super().__init__()
		self._headerData = [
			'Imei',
			'Description',
			'Condition',
			'Spec',
			'Serial',
			'Purchase Date',
			'Partner',
			'Nº Doc',
			'Cost'
		]

		self.name = 'entries'
		self.entries = [
			StockValuationEntryImei(imei, do_cost_price(imei)) 
			for imeis in datasources
			for imei in imeis
		]
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		
		row, col = index.row(), index.column()
		if role == Qt.DisplayRole:
			reg = self.entries[row]
			if col == self.IMEI:
				return reg.imei
			if col == self.DESC:
				return reg.description
			elif col == self.COND:
				return reg.condition
			elif col == self.SPEC:
				return reg.spec
			elif col == self.SERIAL:
				return reg.serial
			elif col == self.DATE:
				return reg.date
			elif col == self.PARTNER:
				return reg.partner
			elif col == self.DOC_REPR:
				return reg.doc
			elif col == self.COST:
				return reg.cost


class InvoicePaymentModel(BaseTable, QtCore.QAbstractTableModel):
	PROFORMA, TOTAL, PAID, PENDING = 0, 1, 2, 3
	
	def __init__(self, invoice):
		super().__init__()
		self.invoice = invoice
		self._headerData = ['Proforma', 'Total', 'Paid', 'Pending']
		self.proformas = invoice.proformas
		self.name = 'proformas'
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		proforma = self.proformas[row]
		paid = sum(p.amount for p in proforma.payments)
		total_debt = proforma.total_debt
		if role == Qt.DisplayRole:
			return [
				proforma.doc_repr,
				total_debt,
				paid,
				total_debt - paid
			][col]
	
	def complete_payments(self, rate, note):
		orm_class = db.SalePayment if isinstance(self.invoice, db.SaleInvoice) \
			else db.PurchasePayment
		for p in self.invoice.proformas:
			db.session.add(
				orm_class(
					date=datetime.now(),
					amount=p.total_debt,
					rate=rate,
					proforma=p,
					note=note
				)
			)
		
		db.session.commit()
	
	def __getitem__(self, item):
		return self.proformas[item]


class InvoiceExpensesModel(BaseTable, QtCore.QAbstractTableModel):
	DATE = 0
	
	def __init__(self, invoice):
		super().__init__()
		self.invoice = invoice
		self._headerData = ['Date', 'Amount', 'Info', 'Proforma']
		
		self.name = 'expenses'
		self.expenses = [
			expense
			for proforma in self.invoice.proformas
			for expense in proforma.expenses
		]
	
	def add(self, date, amount, info, proforma):
		orm_class = db.SaleExpense if isinstance(self.invoice, db.SaleInvoice) \
			else db.PurchaseExpense
		expense = orm_class(date, amount, info, proforma)
		db.session.add(expense)
		db.session.commit()
		self.layoutAboutToBeChanged.emit()
		self.expenses.append(expense)
		self.layoutChanged.emit()
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		expense = self.expenses[row]
		if role == Qt.DisplayRole:
			return [
				expense.date.strftime('%d/%m/%Y'),
				str(expense.amount),
				expense.note,
				expense.proforma.doc_repr
			][col]
		elif role == Qt.DecorationRole:
			if col == self.DATE:
				return QtGui.QIcon(':\calendar')
	
	def __getitem__(self, item):
		return self.expenses[item]


class SwitchModel(QtCore.QAbstractListModel):
	
	def __init__(self):
		super().__init__()
		self._data = list(db.name2db_map.keys())
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		
		if role == Qt.DisplayRole:
			return self._data[index.row()]
	
	def rowCount(self, parent: QModelIndex = ...) -> int:
		return len(self._data)
	
	def switch(self, row):
		fiscal_name = self._data[row]
		db.switch_database(fiscal_name)
		return fiscal_name


class InstrumentedList(list):
	
	def pop(self, index):
		item = super().pop(index)
		try:
			db.session.delete(item)
		except InvalidRequestError:
			''' Remove item from session '''
			db.session.expunge(item)
		
		return item
	
	def insert(self, index, item):
		super().insert(index, item)
		db.session.add(item)


class RepairsModel(BaseTable, QtCore.QAbstractTableModel):
	SN, ITEM, PARTNER, DATE, DESCRIPTION, COST = 0, 1, 2, 3, 4, 5
	
	def __init__(self):
		super().__init__()
		self._headerData = ['SN', 'Item', 'Repairer', 'Date', 'Description', 'Cost']
		self.name = 'repairs'
		self.repairs = InstrumentedList(db.session.query(db.Repair).all())
	
	@property
	def valid(self):
		return all(r.valid for r in self.repairs)
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		repair = self.repairs[row]
		if role == Qt.DisplayRole:
			try:
				clean_repr = repair.item.clean_repr
			except AttributeError:
				clean_repr = ''
			try:
				partner = repair.partner.fiscal_name
			except AttributeError:
				partner = ''
			
			return [
				repair.sn,
				clean_repr,
				partner,
				repair.date.strftime('%d/%m/%Y'),
				repair.description or '',
				str(repair.cost or 0)
			][col]
		elif role == Qt.DecorationRole:
			if col == self.DATE:
				return QtGui.QIcon(':\calendar')
			elif col == self.PARTNER:
				return QtGui.QIcon(':\partners')
	
	def insertRows(self, row: int, count: int, parent: QModelIndex = ...) -> bool:
		self.beginInsertRows(parent, row, row + count - 1)
		self.repairs.insert(row, db.Repair())
		self.endInsertRows()
		return True
	
	def removeRows(self, row: int, count: int, parent: QModelIndex = ...) -> bool:
		self.beginRemoveRows(parent, row, row + count - 1)
		self.repairs.pop(row)
		self.endRemoveRows()
		return True
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		
		from utils import description_id_map, partner_id_map
		
		if not index.isValid():
			return False
		row, col = index.row(), index.column()
		repair = self.repairs[row]
		if role == Qt.EditRole:
			if col == self.SN:
				repair.sn = value
			elif col == self.ITEM:
				try:
					repair.item_id = description_id_map[value]
					repair.item = db.session.query(db.Item).where(db.Item.id == repair.item_id).one()
					return True
				except KeyError:
					return False
			elif col == self.PARTNER:
				try:
					repair.partner_id = partner_id_map[value]
					repair.partner = db.session.query(db.Partner).where(db.Partner.id == repair.partner_id).one()
					return True
				except KeyError:
					return False
			elif col == self.DATE:
				try:
					repair.date = datetime.strptime(value, '%d%m%Y').date()
					return True
				except ValueError:
					return False
			elif col == self.DESCRIPTION:
				repair.description = value
			elif col == self.COST:
				try:
					repair.cost = decimal.Decimal(value)
					return True
				except decimal.InvalidOperation:
					return False
			return False
		
		return False
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return Qt.NoItemFlags
		return Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable
	
	def sort(self, column: int, order: Qt.SortOrder = ...) -> None:
		self.layoutAboutToBeChanged.emit()
		attr = {self.DATE: 'date', self.COST: 'cost'}.get(column)
		if attr:
			reverse = order == Qt.AscendingOrder
			self.layoutAboutToBeChanged.emit()
			self.repairs.sort(key=operator.attrgetter(attr), reverse=reverse)
			self.layoutChanged.emit()


class DiscountModel(BaseTable, QtCore.QAbstractTableModel):
	SN, INVOICE, ITEM, PARTNER, DISCOUNT = 0, 1, 2, 3, 4
	
	def __init__(self):
		super().__init__()
		self._headerData = ['SN', 'Invoice', 'Item', 'Partner', 'Discount']
		self.name = 'discounts'
		self.discounts = InstrumentedList(db.session.query(db.Discount).all())
	
	@property
	def valid(self):
		return all(d.valid for d in self.discounts)
	
	def editable_columns(self):
		return self.SN, self.DISCOUNT
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		discount = self.discounts[row]
		if role == Qt.DisplayRole:
			try:
				clean_repr = discount.item.clean_repr
			except AttributeError:
				clean_repr = ''
			try:
				partner = discount.invoice.partner_name
			except AttributeError:
				partner = ''
			try:
				invoice = discount.invoice.doc_repr_year
			except AttributeError:
				invoice = ''
			
			return [
				discount.sn,
				invoice,
				clean_repr,
				partner,
				str(discount.discount or 0)
			][col]
		elif role == Qt.DecorationRole:
			if col == self.PARTNER:
				return QtGui.QIcon(':\partners')
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return False
		row, col = index.row(), index.column()
		discount = self.discounts[row]
		if role == Qt.EditRole:
			if col == self.SN:
				try:
					reception_serie = (
						db.session.query(db.ReceptionSerie)
						.where(db.ReceptionSerie.serie == value).all()[-1]
					)
				except IndexError:
					return False
				
				if not reception_serie:
					return False
				
				discount.sn = value
				discount.invoice_id = reception_serie.line.reception.proforma.invoice_id
				discount.invoice = db.session.query(db.PurchaseInvoice).where(
					db.PurchaseInvoice.id == discount.invoice_id).one()
				
				discount.item_id = reception_serie.line.item_id
				discount.item = db.session.query(db.Item).where(db.Item.id == discount.item_id).one()
				return True
			
			elif col == self.DISCOUNT:
				try:
					discount.discount = decimal.Decimal(value)
					return True
				except decimal.InvalidOperation:
					return False
		return False
	
	def insertRows(self, row: int, count: int, parent: QModelIndex = ...) -> bool:
		self.beginInsertRows(parent, row, row + count - 1)
		self.discounts.insert(row, db.Discount())
		self.endInsertRows()
		return True
	
	def removeRows(self, row: int, count: int, parent: QModelIndex = ...) -> bool:
		self.beginRemoveRows(parent, row, row + count - 1)
		self.discounts.pop(row)
		self.endRemoveRows()
		return True
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return Qt.NoItemFlags
		if index.column() in self.editable_columns():
			return Qt.ItemIsEditable | Qt.ItemIsEnabled | Qt.ItemIsSelectable
		return Qt.ItemIsEnabled | Qt.ItemIsSelectable


class JournalEntriesModel(BaseTable, QtCore.QAbstractTableModel):
	ID, DATE, DESCRIPTION, TYPE, BALANCE, AUTOGENERATED = 0, 1, 2, 3, 4, 5
	
	def __init__(self):
		super().__init__()
		self._headerData = ['ID', 'Date', 'Description', 'Type', 'Amount', 'Autogenerated']
		self.name = 'entries'
		self.entries = InstrumentedList(db.session.query(db.JournalEntry).all())
	
	def __getitem__(self, item):
		return self.entries[item]
	
	def removeRows(self, position, rows=1, index=QModelIndex()) -> bool:
		self.beginRemoveRows(QModelIndex(), position, position + rows - 1)
		self.entries.pop(position)
		self.endRemoveRows()
		return True
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		entry = self.entries[row]
		if role == Qt.DisplayRole:
			if col == self.ID:
				return str(entry.id).zfill(6)
			elif col == self.DATE:
				return entry.date.strftime('%d/%m/%Y')
			elif col == self.DESCRIPTION:
				return entry.description.capitalize()
			elif col == self.TYPE:
				return entry.related_type.capitalize()
			elif col == self.BALANCE:
				return str(sum(line.debit for line in entry.lines))
			elif col == self.AUTOGENERATED:
				if db.AutoEnum.auto_no == entry.auto:
					return 'No'
				elif db.AutoEnum.auto_yes == entry.auto:
					return 'Yes'
				elif db.AutoEnum.auto_semi == entry.auto:
					return 'Semi'
		
		elif role == Qt.DecorationRole:
			if col == self.DATE:
				return QtGui.QIcon(':\calendar')
			elif col == self.AUTOGENERATED:
				if db.AutoEnum.auto_no == entry.auto:
					return QtGui.QColor(RED)
				elif db.AutoEnum.auto_yes == entry.auto:
					return QtGui.QColor(GREEN)
				elif db.AutoEnum.auto_semi == entry.auto:
					return QtGui.QColor(ORANGE)
	
	@property
	def editable_columns(self):
		return self.DESCRIPTION, self.TYPE, self.DATE
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return Qt.NoItemFlags
		if index.column() in self.editable_columns:
			return Qt.ItemIsEnabled | Qt.ItemIsSelectable | Qt.ItemIsEditable
		return Qt.ItemIsEnabled | Qt.ItemIsSelectable
	
	def setData(self, index: QModelIndex, value: typing.Any, role: int = ...) -> bool:
		if not index.isValid():
			return False
		row, col = index.row(), index.column()
		entry = self.entries[row]
		if role == Qt.EditRole:
			if col == self.DESCRIPTION:
				entry.description = value
				db.session.commit()
				return True
			elif col == self.TYPE:
				if value not in db.JournalEntry.RELATED_TYPES:
					return False
				entry.related_type = value
				db.session.commit()
				return True
			elif col == self.DATE:
				try:
					entry.date = datetime.strptime(value, '%d%m%Y')
				except ValueError:
					return False
				db.session.commit()
				return True


class JournalEntryLineModel(BaseTable, QtCore.QAbstractTableModel):
	ACCOUNT, DEBIT, CREDIT, DESCRIPTION = 0, 1, 2, 3
	
	def __init__(self, entry, form):
		super().__init__()
		self._headerData = ['Account', 'Debit', 'Credit', 'Description']
		self.name = 'lines'
		self.lines = entry.lines
		self.form = form
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		row, col = index.row(), index.column()
		line = self.lines[row]
		if role == Qt.DisplayRole:
			try:
				account = f'{line.account.code} - {line.account.name}'
			except AttributeError:
				account = ''
			return [
				account,
				str(line.debit),
				str(line.credit),
				line.description
			][col]
	
	@property
	def balanced(self):
		return sum(line.debit for line in self.lines) == \
			sum(line.credit for line in self.lines)
	
	def flags(self, index: QModelIndex) -> Qt.ItemFlags:
		if not index.isValid():
			return Qt.ItemIsEnabled
		return Qt.ItemFlags(super().flags(index) | Qt.ItemIsEditable)
	
	def setData(self, index, value, role=Qt.EditRole) -> bool:
		import decimal
		if not index.isValid():
			return False
		line = self.lines[index.row()]
		if role == Qt.EditRole:
			if index.column() == self.ACCOUNT:
				if value not in db.Account.get_leaf_accounts_map().keys():
					return False
				else:
					line.account = db.session.query(db.Account).where(
						db.Account.id == db.Account.get_leaf_accounts_map()[value]
					).one()
					line.account_id = line.account.id
					return True
			
			elif index.column() == self.DEBIT:
				try:
					line.debit = decimal.Decimal(value)
				except decimal.InvalidOperation:
					return False
				else:
					self.form.set_balanced()
					return True
			elif index.column() == self.CREDIT:
				try:
					line.credit = decimal.Decimal(value)
				except decimal.InvalidOperation:
					return False
				else:
					self.form.set_balanced()
					return True
			elif index.column() == self.DESCRIPTION:
				line.description = value
				return True
			return False
		return False
	
	def insertRows(self, position, rows=1, index=QModelIndex()):
		self.beginInsertRows(QModelIndex(), position, position + rows - 1)
		line = db.JournalEntryLine()
		self.lines.insert(position, line)
		self.endInsertRows()
		return True
	
	def removeRows(self, position, rows=1, index=QModelIndex()):
		self.beginRemoveRows(QModelIndex(), position, position + rows - 1)
		self.lines.pop(position)
		self.endRemoveRows()
		return True


class BankingTransactionModel(BaseTable, QtCore.QAbstractTableModel):
	ID, SOURCE, DESCRIPTION, TRANSACTION_DATE, VALUE_DATE, AMOUNT, POSTED = 0, 1, 2, 3, 4, 5, 7
	
	def __init__(self):
		super().__init__()
		self.name = 'banking_transactions'
		self.banking_transactions = db.session.query(db.BankingTransaction).all()
		
		self._headerData = ['ID', 'Description', 'Transaction Date', 'Value Date', 'Amount', 'Posted']
	
	def data(self, index: QModelIndex, role: int = ...) -> typing.Any:
		if not index.isValid():
			return
		
		row, col = index.row(), index.column()
		banking_transaction = self.banking_transactions[row]
		if role == Qt.DisplayRole:
			return [
				banking_transaction.id,
				banking_transaction.source,
				banking_transaction.description,
				banking_transaction.transaction_date.strftime('%d/%m/%Y'),
				banking_transaction.value_date.strftime('%d/%m/%Y'),
				banking_transaction.amount,
				'Yes' if banking_transaction.posted else 'No'
			][col]
		elif role == Qt.DecorationRole:
			if col in (self.TRANSACTION_DATE, self.VALUE_DATE):
				return QtGui.QIcon(':\calendar')
			elif col == self.POSTED:
				return QtGui.QColor('green') if banking_transaction.posted else QtGui.QColor('red')


def caches_clear():
	get_avg_rate.cache_clear()
	get_purchase_expenses_breakdown.cache_clear()
	get_purchase_stock_value.cache_clear()
	get_sale_stock_key.cache_clear()
	get_sale_shipping_key.cache_clear()
	get_sale_terms_key.cache_clear()
	get_sale_breakdown.cache_clear()
	get_sale_proforma_stock_value.cache_clear()


if __name__ == '__main__':
	m = FucksModel()
	m._fucks = [
		Fuck(1, 11, 12),
		Fuck(2, 210, 265),
		Fuck(3, 4, 5),
		Fuck(4, 96, 126)
	]
	
	m.export(2023)
