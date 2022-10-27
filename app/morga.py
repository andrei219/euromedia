from builtins import enumerate

from models import do_cost_price
from db import ReceptionSerie, session
from openpyxl import load_workbook

FILEPATH = r'Z:\fAULTY VESI25.10.xlsx'

def find_cost(imei):

    try:
        return do_cost_price(imei).ptotal
    except:
        return "Not found"


def find_doc_repr_and_partner(imei):

    try:
        obj = session.query(ReceptionSerie).where(ReceptionSerie.serie == imei).all()[-1]

    except IndexError:
        doc_repr = "Not found"
        partner = "Not found"
    else:
        doc_repr = obj.line.reception.proforma.doc_repr
        partner = obj.line.reception.proforma.partner_name

    return doc_repr, partner


if __name__ == '__main__':

    wb = load_workbook(FILEPATH)
    sheet = wb.active

    for i, row in enumerate(sheet.iter_rows(min_col=3, max_col=3, min_row=1), start=1):

        imei = row[0].value
        cost = find_cost(imei)
        doc_repr, partner = find_doc_repr_and_partner(imei)


        sheet.cell(i, 10).value = cost
        sheet.cell(i, 11).value = doc_repr
        sheet.cell(i, 12).value = partner

    wb.save(FILEPATH)

